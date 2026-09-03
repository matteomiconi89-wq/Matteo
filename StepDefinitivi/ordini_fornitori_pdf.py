# -*- coding: utf-8 -*-
"""Ordini PDF per FORNITORE dal file SCHEDA BASE_..._COMPLETO.xlsm.
Un PDF per fornitore (i fogli d'ordine creati da automazioni_excel_FINALE:
CODICE/DESCRIZIONE/Q.TA/EUR CAD/EUR TOT), con il riferimento COMMESSA."""
import datetime
import os

from PIL import Image, ImageDraw, ImageFont

PAG = (2338, 1653)               # A4 orizzontale 200 dpi
MARG = 90
_FONTDIR = r"C:\Windows\Fonts"
# fogli "di sistema" del modello: NON sono ordini fornitore
_BASE = {"GENERALE", "PALLET", "LAVORAZIONE MASSELLI", "BORDI", "PANNELLAME",
         "MASSELLO", "FERRAMENTA", "ILLUMINAZIONE", "SEZ. RAGGRUPPATA"}


def _font(px, bold=False):
    return ImageFont.truetype(
        os.path.join(_FONTDIR, "arialbd.ttf" if bold else "arial.ttf"), px)


def _eur(v):
    try:
        return f"{float(v):,.2f}".replace(",", "§").replace(".", ",").replace(
            "§", ".") + " €"
    except Exception:
        return ""


def _num(v):
    try:
        f = float(v)
        return f"{f:g}"
    except Exception:
        return str(v or "")


def _dx(dr, xr, y, txt, px, bold=False, col="black"):
    f = _font(px, bold)
    w = dr.textlength(txt, font=f)
    dr.text((xr - w, y), txt, fill=col, font=f)


def _disegna_ordine(fornitore, commessa, righe, path):
    img = Image.new("RGB", PAG, "white")
    dr = ImageDraw.Draw(img)
    dr.text((MARG, 44), f"ORDINE - {fornitore}", fill="black",
            font=_font(50, True))
    oggi = datetime.date.today().strftime("%d/%m/%Y")
    dr.text((MARG, 120), f"COMMESSA: {commessa}      Data: {oggi}",
            fill=(80, 80, 80), font=_font(30, True))
    dr.line([(MARG, 172), (PAG[0] - MARG, 172)], fill="black", width=3)
    # colonne
    x_cod = MARG + 6
    x_desc = MARG + 300
    r_qta = PAG[0] - MARG - 560
    r_cad = PAG[0] - MARG - 300
    r_tot = PAG[0] - MARG - 8
    y = 200
    dr.text((x_cod, y), "CODICE", fill="black", font=_font(26, True))
    dr.text((x_desc, y), "DESCRIZIONE", fill="black", font=_font(26, True))
    _dx(dr, r_qta, y, "Q.TA", 26, True)
    _dx(dr, r_cad, y, "€/CAD", 26, True)
    _dx(dr, r_tot, y, "€ TOT.", 26, True)
    y += 44
    dr.line([(MARG, y), (PAG[0] - MARG, y)], fill="black", width=2)
    y += 10
    passo = min(48, (PAG[1] - 140 - y) / max(len(righe), 1))
    fpx = 24 if passo >= 30 else (20 if passo >= 24 else 16)
    tot = 0.0
    for i, (cod, desc, qta, cad, _ttot) in enumerate(righe):
        if i % 2:
            dr.rectangle([MARG, y - 2, PAG[0] - MARG, y + passo - 4],
                         fill=(244, 246, 240))
        # il totale riga = q.ta x EUR/CAD: nel foglio fornitore e' una formula
        # che openpyxl NON calcola, quindi lo ricalcolo qui
        try:
            riga_tot = float(qta or 0) * float(cad or 0)
        except Exception:
            riga_tot = None
        dr.text((x_cod, y), str(cod or "")[:20], fill="black", font=_font(fpx))
        dr.text((x_desc, y), str(desc)[:62], fill="black", font=_font(fpx))
        _dx(dr, r_qta, y, _num(qta), fpx)
        _dx(dr, r_cad, y, _eur(cad), fpx)
        _dx(dr, r_tot, y, _eur(riga_tot) if riga_tot is not None else "", fpx)
        if riga_tot:
            tot += riga_tot
        y += passo
    y += 8
    dr.line([(r_cad - 120, y), (PAG[0] - MARG, y)], fill="black", width=2)
    y += 8
    _dx(dr, r_cad, y, "TOTALE ORDINE", 30, True)
    _dx(dr, r_tot, y, _eur(tot), 30, True)
    img.save(path, "PDF", resolution=200)


def genera_pdf_ordini(complet_path, commessa, out_dir=None, log=print):
    """Un PDF d'ordine per ogni foglio-fornitore del _COMPLETO. out_dir:
    default = sottocartella ORDINI accanto al file. Ritorna la lista dei PDF."""
    import openpyxl
    wb = openpyxl.load_workbook(complet_path, data_only=True)
    if out_dir is None:
        out_dir = os.path.join(os.path.dirname(complet_path), "ORDINI")
    os.makedirs(out_dir, exist_ok=True)
    fatti = []
    for ws in wb.worksheets:
        if ws.title in _BASE:
            continue
        hdr = [str(ws.cell(1, c).value or "").strip().upper()
               for c in range(1, 6)]
        if "CODICE" not in hdr or "DESCRIZIONE" not in hdr:
            continue                     # non e' un ordine (es. laccatura/sez)
        righe = []
        for r in range(2, ws.max_row + 1):
            cod = ws.cell(r, 1).value
            desc = ws.cell(r, 2).value
            if str(cod or "").strip().upper() == "TOTALE":
                break
            if desc:
                righe.append((cod, desc, ws.cell(r, 3).value,
                              ws.cell(r, 4).value, ws.cell(r, 5).value))
        if not righe:
            continue
        forn = "".join(ch if ch not in '\\/:*?"<>|' else "_"
                       for ch in str(ws.title))
        path = os.path.join(out_dir, f"ORDINE_{forn}_{commessa}.pdf")
        try:
            _disegna_ordine(ws.title, commessa, righe, path)
            fatti.append(path)
            log(f"  ordine PDF -> {os.path.basename(path)}")
        except Exception as ex:
            log(f"    [!] ordine {ws.title} non riuscito ({ex})")
    return fatti
