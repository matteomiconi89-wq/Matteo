# -*- coding: utf-8 -*-
"""scheda_base_da_filiera - compila la SCHEDA BASE dal giro di filiera.

Copia lo stampo SCHEDA_BASE_MODELLO_FILIERA.xlsm dentro la cartella madre
del lavoro (Programmi CNC\SCHEDA BASE) come "SCHEDA BASE_<lavoro>.xlsm"
e la riempie via Excel COM (tendine e macro
sopravvivono, i valori restano in cache per lo script fornitori):
  - GENERALE, distinta pezzi (righe sopra l'END): CODICE, MOBILE,
    MATERIALE (dal 3D quando c'e'), Q.TA, PEZZO, H/L/SP
  - GENERALE, blocco FERRAMENTA: dal conteggio 'Ferramenta -- codice'
    dello STP, agganciato per CODICE al catalogo interno (cosi' la
    DESCRIZIONE combacia esatta e il prezzo scatta da solo)
Non tocca: U.M. laccatura, LATI DA LACCARE, intestazioni bordi/finiture
(scelte di commessa), numero lastre. Tutte le formule restano vive.
"""

import os
import shutil

# lo STAMPO (listino/base di ferramenta, pannellame, bordi, masselli) sta su
# Dropbox\STEFANO\FORNITORI cosi' e' CONDIVISO: chiunque aggiunge materiali o
# codici lo modifica li' e la filiera li pesca da solo. Ripiego sul locale
# (Desktop\CLAUDE\Definitivi) se Dropbox non c'e'.
_STAMPO_DROPBOX = (r"C:\Users\User\Dropbox\STEFANO\FORNITORI"
                   r"\SCHEDA_BASE_MODELLO_FILIERA.xlsm")
_STAMPO_LOCALE = (r"C:\Users\User\Desktop\CLAUDE\Definitivi"
                  r"\SCHEDA_BASE_MODELLO_FILIERA.xlsm")
STAMPO = _STAMPO_DROPBOX if os.path.isfile(_STAMPO_DROPBOX) else _STAMPO_LOCALE
RIGA_DATI = 4          # prima riga distinta
RIGA_END = 338         # riga 'END' dello stampo (i dati stanno sopra)
FERR_START = 345       # prima riga blocco ferramenta
FERR_END = 530


def _leggi_riepilogo(pu):
    """Righe della distinta dal riepilogo PROGRAMMI_UNICI (valori)."""
    import openpyxl
    p = os.path.join(pu, "_RIEPILOGO_programmi_unici.xlsx")
    if not os.path.isfile(p):
        p = os.path.join(pu, "_RIEPILOGO_programmi_unici_NEW.xlsx")
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["PROGRAMMI_UNICI"]
    righe = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        # r = FILE, PEZZO, GRUPPI, QTA, L, H, SP, MATERIALE, ...
        righe.append({"file": r[0], "pezzo": r[1], "mobile": r[2],
                      "qta": r[3], "L": r[4], "H": r[5], "SP": r[6],
                      "mat": r[7] or ""})
    return righe


def _etichette_lavoro(base):
    """Mappa {file dxf: numero-programma} identica a quella dell'esploso
    e delle schede (da scheda_pdf._etichette sul manifest): cosi' il
    CODICE della scheda costo combacia coi palloncini e i titoli."""
    import json
    import sys
    try:
        d = r"C:\Users\User\Desktop\CLAUDE\SchedePDF"
        if d not in sys.path:
            sys.path.append(d)
        from scheda_pdf import _etichette
        mf = os.path.join(base, "_manifest.json")
        return _etichette(json.load(open(mf, encoding="utf-8")))
    except Exception:
        return {}


def _codice_num(nome):
    """Numero pezzo dal nome ('02_FIANCO...' -> 2), senno' progressivo."""
    import re
    m = re.match(r"\s*0*(\d+)", str(nome or ""))
    return int(m.group(1)) if m else None


def _tokens(s):
    import re
    out = set()
    for t in re.split(r"[\s_./]+", str(s).lower()):
        t = t.strip(".,")
        if len(t) >= 3 and not t.replace("x", "").isdigit():
            out.add(t)
    return out


def _match_token(a, b):
    return a == b or (len(a) >= 4 and b.startswith(a)) \
        or (len(b) >= 4 and a.startswith(b))


def _e_sezionato(nome):
    u = str(nome).upper().rstrip(".")
    return "SEZIONAT" in u or u.endswith(" SEZ")


def abbina_lastra(mat, sp, voci):
    """Lastra migliore per un materiale. voci = [(nome, W, H, vsp)].
    Spessore con tolleranza +/-1,5 mm (grezzo 18 vs finito 19); a pari
    punteggio, se le candidate sono la STESSA lastra (stesse dimensioni)
    vince la versione intera (non 'SEZ'); lastre diverse = ambiguo, None.
    Ritorna (nome, W, H) oppure None."""
    tm = _tokens(mat)
    cands = []
    for nome, w, h, vsp in voci:
        if (sp is not None and vsp is not None
                and abs(float(vsp) - float(sp)) > 1.5):
            continue
        n = sum(1 for a in tm
                if any(_match_token(a, b) for b in _tokens(nome)))
        if n >= 1:
            cands.append((n, nome, w, h))
    if not cands:
        return None
    cands.sort(key=lambda c: -c[0])
    if cands[0][0] < 2:
        return None
    top = [c for c in cands if c[0] == cands[0][0]]
    if len(top) > 1:
        dims = {(round(c[2] or 0), round(c[3] or 0)) for c in top}
        if len(dims) != 1:
            return None                  # lastre diverse: non indovino
        interi = [c for c in top if not _e_sezionato(c[1])]
        top = interi or top
    return top[0][1], top[0][2], top[0][3]


def _blocco_pannellame(wb, ws, righe, log, fogli=None):
    """Compila L345:L368 (lastre) + formula MQ netti in S; L374 massello."""
    import re

    # materiali distinti della distinta con lo spessore piu' comune
    mats = {}
    for r in righe:
        m = (r["mat"] or "").strip()
        if not m:
            continue
        mats.setdefault(m, []).append(r["SP"])
    if not mats:
        return

    pan = wb.Worksheets("PANNELLAME")
    npan = pan.Cells(pan.Rows.Count, 1).End(-4162).Row
    voci = []
    for v in pan.Range(f"A2:D{max(npan, 2)}").Value or []:
        if isinstance(v[0], str) and v[0].strip():
            voci.append((v[0].strip(), v[1], v[2], v[3]))

    riga_l = 345
    abbinate = 0
    for mat, sps in sorted(mats.items()):
        if riga_l > 368:
            log("    [!] blocco lastre pieno: materiali in piu' saltati")
            break
        if "MASSELLO" in mat.upper():
            continue                     # ha il suo blocco sotto
        sp = max(set(sps), key=sps.count) if sps else None
        ab = abbina_lastra(mat, sp, voci)
        scelta = ab[0] if ab else None
        ws.Cells(riga_l, 12).Value = scelta          # L: nome lastra o vuoto
        if fogli and mat in fogli:                   # Q: n. fogli dal nesting
            ws.Cells(riga_l, 17).Value = fogli[mat]
        chiave = mat.replace('"', '""')
        ws.Cells(riga_l, 19).Formula = (             # S: MQ netti ESATTI
            f'=SUMPRODUCT((TRIM($C$4:$C$339)="{chiave}")'
            '*$AG$4:$AG$339)')
        # R (MC tot) e T (MQ): il template le ha su tutte le righe TRANNE la
        # prima (345) -> le scrivo io per OGNI materiale, cosi' anche il primo
        # esce con MC e MQ (senno' sembra che manchi)
        ws.Cells(riga_l, 18).Formula = (
            f'=M{riga_l}/1000*N{riga_l}/1000*O{riga_l}/1000*Q{riga_l}')
        ws.Cells(riga_l, 20).Formula = (
            f'=M{riga_l}/1000*N{riga_l}/1000*Q{riga_l}')
        if scelta:
            abbinate += 1
        riga_l += 1
    # svuoto le righe pannellame NON usate: R (MC), S (MQ netti) e T (MQ)
    # possono avere formule STANTIE del template (es. un MQ ripetuto sulle
    # righe vuote) -> le pulisco riga per riga (win32com non pulisce le range
    # multi-riga in colpo solo)
    for rr in range(riga_l, 369):
        try:
            ws.Range(f"R{rr}:T{rr}").ClearContents()
        except Exception:
            pass
    log(f"  SCHEDA BASE: lastre abbinate {abbinate}/{riga_l - 345} "
        "materiali (n. fogli da decidere)")

    # MASSELLO (sempre a MC): il MC lo somma gia' N374=U338 dalla colonna U
    # (SEARCH "MASSELLO" nei materiali); qui serve solo scegliere la TAVOLA
    # da comprare = la piu' sottile del catalogo che COPRE lo spessore max
    # dei pezzi massello (il grezzo si compra in tavole e si sfoglia).
    mas = [r for r in righe if "MASSELLO" in (r["mat"] or "").upper()]
    if mas:
        try:
            cat = wb.Worksheets("MASSELLO")
            ncat = cat.Cells(cat.Rows.Count, 1).End(-4162).Row
            tavole = [(v[0], float(v[1])) for v in
                      (cat.Range(f"A2:B{max(ncat, 2)}").Value or [])
                      if v[0] and v[1] is not None]
            sp_max = max((r["SP"] for r in mas if r["SP"]), default=0)
            coprono = [t for t in tavole if t[1] >= sp_max]
            scelta = (min(coprono, key=lambda t: t[1]) if coprono
                      else (max(tavole, key=lambda t: t[1]) if tavole
                            else None))
            if scelta:
                ws.Cells(374, 12).Value = scelta[0]
                log(f"  SCHEDA BASE: massello (MC) -> tavola {scelta[0]}")
        except Exception:
            pass


def compila_scheda_base(base, log=print, ferramenta=None, fogli=None,
                        fornitori=None):
    """base = cartella madre "Programmi CNC" del lavoro. ferramenta opzionale:
    [(codice, qta)]; se None niente blocco ferramenta. fornitori opzionale:
    {codice: FORNITORE} (dai layer del DWG d'assieme) -> scritto in colonna E
    (nascosta) del blocco ferramenta, esatto dal modello."""
    import pythoncom
    pythoncom.CoInitialize()
    import win32com.client

    if not os.path.isfile(STAMPO):
        log(f"  [!] stampo SCHEDA BASE non trovato: {STAMPO}")
        return None
    pu = os.path.join(base, "PROGRAMMI_UNICI")
    righe = _leggi_riepilogo(pu)
    if not righe:
        log("  [!] riepilogo vuoto: niente SCHEDA BASE")
        return None
    if len(righe) > RIGA_END - RIGA_DATI:
        log(f"  [!] {len(righe)} pezzi ma lo stampo ne tiene "
            f"{RIGA_END - RIGA_DATI}: tronco (dividere il lavoro)")
        righe = righe[:RIGA_END - RIGA_DATI]

    lavoro = os.path.basename(os.path.dirname(base)) or "lavoro"
    # TUTTO dentro la cartella madre del lavoro. Prima la scheda nasceva
    # FUORI (accanto ai file 3D) e con lei se ne andavano il _COMPLETO e la
    # cartella ORDINI, che vengono creati accanto alla scheda: era questo a
    # sparpagliare l'output del giro
    out_scheda = os.path.join(base, "SCHEDA BASE")
    os.makedirs(out_scheda, exist_ok=True)
    dst = os.path.join(out_scheda, f"SCHEDA BASE_{lavoro}.xlsm")
    try:
        shutil.copy2(STAMPO, dst)
    except PermissionError:
        log(f"  [!] {os.path.basename(dst)} aperta in Excel: "
            "chiudila e rilancia (scheda NON rigenerata)")
        return None

    xl = win32com.client.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    xl.EnableEvents = False
    try:
        wb = xl.Workbooks.Open(dst)
        xl.Calculation = -4135                      # manuale
        ws = wb.Worksheets("GENERALE")
        ws.Range("C1").Value = lavoro

        # --- distinta pezzi: blocco unico via matrice (veloce) ----------
        # CODICE = etichetta unica (come esploso e schede); ripiego sul
        # numero dal nome, poi progressivo
        eti = _etichette_lavoro(base)
        dati = []
        for i, r in enumerate(righe, 1):
            e = eti.get(r["file"])
            if e is not None:
                cod = int(e) if str(e).isdigit() else e
            else:
                cod = _codice_num(r["pezzo"]) or i
            dati.append([cod, r["mobile"], r["mat"] or None, r["qta"],
                         r["pezzo"], r["H"], r["L"], r["SP"]])
        n = len(dati)
        ws.Range(f"A{RIGA_DATI}:H{RIGA_DATI + n - 1}").Value = dati
        log(f"  SCHEDA BASE: {n} pezzi in distinta")

        # --- CORNICI a ML (colonna J, "ML CORNICE 19x9"): i pezzi il cui
        # materiale parla di CORNICE/GESSO si contano a METRO LINEARE
        # (lato lungo x q.ta), non a superficie; il massello resta a MC.
        n_corn = 0
        for i, r in enumerate(righe):
            mat = (r["mat"] or "").upper()
            if "CORNICE" in mat or "GESSO" in mat:
                lung = max(float(r["L"] or 0), float(r["H"] or 0)) / 1000.0
                ws.Cells(RIGA_DATI + i, 10).Value = round(
                    lung * float(r["qta"] or 1), 3)
                n_corn += 1
        if n_corn:
            log(f"  SCHEDA BASE: {n_corn} cornici a ML (colonna J)")

        # --- blocco acquisto PANNELLAME (sotto la distinta): abbina i
        # materiali ai nomi lastra del catalogo (esatti => prezzi vivi)
        # e scrive la formula MQ netti sul materiale; n. fogli resta
        # all'utente. Abbinamento prudente: spessore + parole chiave,
        # nel dubbio lascia vuoto.
        try:
            _blocco_pannellame(wb, ws, righe, log, fogli)
        except Exception as ex:
            log(f"    [!] blocco lastre non compilato: {ex}")

        # --- blocco ferramenta: scrivo SOLO il CODICE (col B) e la Q.TA
        # (col D). Il resto lo fanno le FORMULE del template (nuovo stampo):
        # C=descrizione, E=fornitore, F=prezzo -> INDEX/MATCH sul catalogo
        # FERRAMENTA per CODICE (col E, col trucco TEXT("00000000") sui codici
        # numerici); G=D*F. Cosi' i prezzi restano vivi e allineati al listino.
        if ferramenta:
            riga = FERR_START
            for cod, qta in ferramenta:
                if riga > FERR_END:
                    log("  [!] blocco ferramenta pieno: voci in piu' "
                        "tagliate")
                    break
                ws.Cells(riga, 1).Value = lavoro          # A  MOBILE
                ws.Cells(riga, 2).Value = str(cod)        # B  CODICE (dal DWG)
                ws.Cells(riga, 4).Value = qta             # D  Q.TA
                riga += 1
            log(f"  SCHEDA BASE: {len(ferramenta)} codici ferramenta in B "
                "(descrizione/fornitore/prezzo dalle formule del catalogo)")

        xl.Calculation = -4105                      # automatico
        xl.CalculateFull()
        wb.Save()                                   # valori in cache
        wb.Close(False)
        log(f"  -> {dst}")
        return dst
    finally:
        xl.EnableEvents = True
        try:
            xl.Quit()
        except Exception:
            pass


def genera_ordini_fornitori(scheda_path, log=print):
    """Dalla SCHEDA BASE compilata sforna il file '<scheda>_COMPLETO' coi fogli
    d'ORDINE per FORNITORE (ferramenta + illuminazione + laccatura) + la
    sezionatura raggruppata, ACCANTO alla scheda (quindi dentro la cartella
    madre del lavoro). Riusa HEADLESS (niente finestre)
    la logica del tool GUI automazioni_excel_FINALE, cosi' l'ordine si genera da
    solo in coda alla filiera."""
    import sys as _sys
    d = r"C:\Users\User\Desktop\CLAUDE\Definitivi"
    if d not in _sys.path:
        _sys.path.insert(0, d)
    try:
        import automazioni_excel_FINALE as AF
    except Exception as ex:
        log(f"    [!] automazione fornitori non disponibile ({ex})")
        return None

    class _NoGui:                       # zittisce i popup della GUI
        def showinfo(self, *a, **k):
            pass

        def showerror(self, *a, **k):
            pass

        def showwarning(self, *a, **k):
            pass
    AF.messagebox = _NoGui()

    class _Holder:
        def __init__(self, p):
            self._p = p

        def get(self):
            return self._p

    try:
        import openpyxl
        app = AF.AutomazioniFornitori.__new__(AF.AutomazioniFornitori)
        app.file_path = _Holder(scheda_path)
        out_dir = os.path.dirname(scheda_path)
        base = os.path.splitext(os.path.basename(scheda_path))[0]
        ext = os.path.splitext(scheda_path)[1]

        def _copia_accanto(suffisso):   # accanto al lavoro, non in Downloads
            fn = f"{base}_{suffisso}{ext}"
            dest = os.path.join(out_dir, fn)
            shutil.copy2(scheda_path, dest)
            wb = openpyxl.load_workbook(dest, data_only=True,
                                        keep_vba=(ext.lower() == ".xlsm"))
            return wb, dest, fn
        app.crea_copia_completa = _copia_accanto

        app.elaborazione_completa()
        comp = os.path.join(out_dir, f"{base}_COMPLETO{ext}")
        if os.path.isfile(comp):
            log(f"  ordini fornitori -> {os.path.basename(comp)}")
            # PDF d'ordine PER FORNITORE (un file ciascuno) col riferimento
            # COMMESSA preso dal percorso (es. 25-A019)
            try:
                import re
                m = re.search(r"\d{2}-[A-Za-z]\d{2,4}", scheda_path)
                commessa = m.group(0) if m else base
                from ordini_fornitori_pdf import genera_pdf_ordini
                genera_pdf_ordini(comp, commessa, log=log)
            except Exception as ex:
                log(f"    [!] PDF ordini non creati ({ex})")
            # RIEPILOGO COSTI per mobile (foglio Excel nel _COMPLETO): 6
            # categorie d'acquisto + manodopera (45 EUR/h, cella editabile)
            try:
                from riepilogo_costi import genera_riepilogo_costi
                genera_riepilogo_costi(comp, eur_ora=45.0, log=log)
            except Exception as ex:
                log(f"    [!] riepilogo costi non creato ({ex})")
            return comp
        log("    [!] file ordini fornitori non creato")
        return None
    except Exception as ex:
        log(f"    [!] ordini fornitori falliti ({ex})")
        return None
