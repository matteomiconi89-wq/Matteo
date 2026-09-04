"""
step_dxf_definitivi - dai file STEP alla cartella "Programmi CNC", a mani
libere. Tutto il giro scrive li' dentro: niente file sparsi nel lavoro.

Per ogni .stp/.step selezionato:
  1. nuovo disegno nell'AutoCAD GIA' APERTO (non chiude ne' riavvia nulla)
  2. IMPORT dello STEP (aspetta la traduzione in background)
  3. salva il DWG col NOME REALE (nome del file step) accanto allo step
  4. ESPLODINOMI automatico (tutti i blocchi + etichette) e ri-salva il DWG
  5. ESPSOL nella cartella  Programmi CNC\\<nome file>\\  accanto allo step
     (i frontali cassetto SX/CX/DX escono gia' in file unico)
  6. chiude il disegno

Si puo' selezionare anche un .dwg gia' pronto: si salta l'import, il DWG
viene aperto diretto e NON viene mai salvato (originale intatto);
ESPLODINOMI + ESPSOL lavorano in memoria, il resto e' identico.

Richiede i LISP in LISP_DIR (nomina-solididef, esplodi-nomi, esporta-solidi
v9 con espnomi-auto / espsol-core) e AutoCAD aperto con la cartella dei LISP
nei percorsi attendibili (TRUSTEDPATHS), senno' SECURELOAD blocca il load.
"""

import os
import sys
import time
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox

LISP_DIR = r"C:\Users\User\Desktop\CLAUDE\autocad"
LISP_FILES = ["nomina-solididef.lsp", "esplodi-nomi.lsp", "esporta-solidi.lsp"]
# cartella MADRE del giro: dentro ci finisce TUTTO (pezzi, programmi
# macchina, tavole cliente, libretti, scheda base, ordini, nesting)
OUT_DIRNAME = "Programmi CNC"

# False = modo 21032: frontali cassetto uniti in file unico, confronto col
#         modulo 21032 (regole CNC) e sezionatura finale.
# True  = modo GENERICO (step_dxf_generico): frontali NON uniti, confronto
#         con file_unici_dxf_generico, NIENTE sezionatura.
MODO_GENERICO = False

# moduli delle altre automazioni (confronto, sezionatura, TLF, MPR/MPRX)
for _d in (r"C:\Users\User\Desktop\CLAUDE\ConfrontoProgrammi",
           r"C:\Users\User\Desktop\CLAUDE\bom_exe",
           r"C:\Users\User\Desktop\CLAUDE\Dxf2Tlf",
           r"C:\Users\User\Desktop\CLAUDE\Dxf2Mpr"):
    if _d not in sys.path:
        sys.path.append(_d)
FLAG = os.path.join(os.environ.get("TEMP", r"C:\Windows\Temp"),
                    f"_step_dxf_definitivi_{os.getpid()}.flag")

TIMEOUT_LOAD = 90        # s per il caricamento dei lisp
TIMEOUT_IMPORT = 900     # s per la traduzione dello STEP
TIMEOUT_NOMI = 900       # s per ESPLODINOMI
TIMEOUT_ESPSOL = 3600    # s per ESPSOL (tanti pezzi = tanto tempo)


def fwd(p):
    return p.replace("\\", "/")


def com_retry(fn, tries=200, delay=0.3, descr=""):
    """AutoCAD rifiuta le chiamate COM mentre e' occupato: riprova.
    (A volte il rifiuto arriva come AttributeError del dispatch dinamico.)
    'Input non valido' invece e' permanente: inutile ritentarlo 60 s."""
    import pywintypes
    last = None
    for _ in range(tries):
        try:
            return fn()
        except (pywintypes.com_error, AttributeError) as e:
            info = getattr(e, "excepinfo", None)
            if info and info[5] == -2145386493:      # Input non valido
                raise RuntimeError(
                    f"AutoCAD ha rifiutato il comando ({descr}): "
                    "input non valido")
            last = e
            time.sleep(delay)
    raise RuntimeError(f"AutoCAD occupato troppo a lungo ({descr}): {last}")


def get_acad():
    import win32com.client
    acad = win32com.client.Dispatch("AutoCAD.Application")
    acad.Visible = True
    ricorda_finestra(acad)
    return acad


_ACAD_FIN = {"hwnd": None}


def ricorda_finestra(acad):
    """Salva l'HWND di AutoCAD quando la COM risponde: nelle emergenze
    (LISP sospeso su un comando a meta') la COM e' morta e resta solo
    la via win32."""
    try:
        _ACAD_FIN["hwnd"] = int(acad.HWND)
    except Exception:
        pass


def send(doc, s, descr=""):
    com_retry(lambda: doc.SendCommand(s), descr=descr or s[:40])


def invia_esc(app=None, n=3):
    """ESC come TASTO vero alle finestre giuste di AutoCAD (PostMessage,
    mai furto di focus): annulla un comando a meta' anche quando AutoCAD
    e' sospeso dentro un LISP e rifiuta la COM. Bersagli: la finestra col
    FOCUS del thread UI (GetGUIThreadInfo), le top-level del thread (la
    riga di comando flottante e' una owned top-level, non una child) e le
    child visibili; la cornice per ultima. NB: SendCommand con \\x03 su
    AutoCAD 2027 alza 'Input non valido': il cancel via COM non esiste."""
    try:
        import ctypes
        from ctypes import wintypes

        import win32con
        import win32gui
        import win32process

        hwnd = _ACAD_FIN.get("hwnd")
        if not (hwnd and win32gui.IsWindow(hwnd)):
            if app is not None:
                ricorda_finestra(app)
                hwnd = _ACAD_FIN.get("hwnd")
        if not (hwnd and win32gui.IsWindow(hwnd)):
            return

        class GUITHREADINFO(ctypes.Structure):
            _fields_ = [("cbSize", wintypes.DWORD),
                        ("flags", wintypes.DWORD),
                        ("hwndActive", wintypes.HWND),
                        ("hwndFocus", wintypes.HWND),
                        ("hwndCapture", wintypes.HWND),
                        ("hwndMenuOwner", wintypes.HWND),
                        ("hwndMoveSize", wintypes.HWND),
                        ("hwndCaret", wintypes.HWND),
                        ("rcCaret", wintypes.RECT)]

        tid = win32process.GetWindowThreadProcessId(hwnd)[0]
        bersagli = []

        gti = GUITHREADINFO(cbSize=ctypes.sizeof(GUITHREADINFO))
        if ctypes.windll.user32.GetGUIThreadInfo(tid, ctypes.byref(gti)):
            for h in (gti.hwndFocus, gti.hwndCaret, gti.hwndActive):
                if h and h not in bersagli:
                    bersagli.append(h)

        finestre = []
        try:
            win32gui.EnumThreadWindows(
                tid, lambda h, _: (finestre.append(h), True)[1], None)
        except Exception:
            pass
        for w in finestre:
            if w not in bersagli:
                bersagli.append(w)
            try:
                win32gui.EnumChildWindows(
                    w,
                    lambda h, _: ((bersagli.append(h)
                                   if (h not in bersagli
                                       and win32gui.IsWindowVisible(h))
                                   else None), True)[1],
                    None)
            except Exception:
                pass
        if hwnd not in bersagli:
            bersagli.append(hwnd)
        bersagli = bersagli[:80]

        giu = (0x01 << 16) | 1                      # scancode ESC, repeat 1
        su = giu | (1 << 30) | (1 << 31)
        for _ in range(n):
            for h in bersagli:
                win32gui.PostMessage(h, win32con.WM_KEYDOWN,
                                     win32con.VK_ESCAPE, giu)
                win32gui.PostMessage(h, win32con.WM_CHAR, 0x1B, giu)
                win32gui.PostMessage(h, win32con.WM_KEYUP,
                                     win32con.VK_ESCAPE, su)
            time.sleep(0.15)
    except Exception:
        pass


def sblocca_attivo(acad):
    """ESC sul disegno attivo: se l'utente ha un comando a meta' (una
    COPIA in attesa del secondo punto) i comandi della filiera finirebbero
    come risposte a quel prompt. Non fatale se non c'e' nessun disegno."""
    invia_esc(acad)


_flag_seq = [0]


def _flag_nuovo():
    """Sentinella con nome sempre nuovo: un ping in ritardo scrive un file
    che nessuno sta piu' aspettando (mai scambiarlo per il blocco vero).
    Bonifica preventiva: un residuo di una corsa vecchia (stesso PID
    riciclato da Windows) farebbe sembrare gia' fatto un lavoro mai partito."""
    _flag_seq[0] += 1
    p = FLAG[:-5] + f"_{_flag_seq[0]}.flag"
    try:
        os.remove(p)
    except OSError:
        pass
    return p


_HR_MORTO = {-2147023174, -2147417848}   # RPC spento / oggetto sganciato


def _controlla_vivo(doc, err_msg):
    """Durante un LISP lungo AutoCAD RIFIUTA le chiamate COM (normale);
    se invece il processo e' proprio morto l'errore e' diverso: meglio
    accorgersene subito che dopo un timeout da un'ora."""
    try:
        doc.Name
    except Exception as e:
        if getattr(e, "hresult", None) in _HR_MORTO:
            raise RuntimeError(
                f"AutoCAD si e' chiuso durante il lavoro [{err_msg}]")


def _scrivi_flag(p, testo='"ok"'):
    return (f'(setq _ff (open "{fwd(p)}" "w"))'
            f'(write-line {testo} _ff)(close _ff) ')


def _leggi_flag(p, attesa=2.0):
    """Contenuto della sentinella (aspetta che il LISP chiuda il file)."""
    t0 = time.time()
    while time.time() - t0 < attesa:
        try:
            with open(p) as f:
                t = f.read().strip()
            if t:
                return t
        except OSError:
            pass
        time.sleep(0.1)
    return ""


def send_flag(doc, lisp_code, timeout, err_msg):
    """Accoda i comandi + scrittura di un file-sentinella: quando il file
    compare, tutto cio' che precede e' stato eseguito davvero.
    Prima un PING breve preceduto da ESC: libera un eventuale comando a
    meta' e verifica che la riga di comando risponda, cosi' un blocco si
    scopre in 45 s e non dopo un timeout da un'ora. Mai ESC durante
    l'attesa del blocco vero: abortirebbe un lavoro sano."""
    for _ in range(3):
        ping = _flag_nuovo()
        invia_esc()
        # il ping scrive CMDACTIVE: le espressioni LISP girano ANCHE in
        # trasparenza sopra un comando a meta', quindi la sola esistenza
        # del file non basta a dire che la riga di comando e' libera
        send(doc, _scrivi_flag(ping, '(itoa (getvar "CMDACTIVE"))'),
             descr="ping riga di comando")
        t0 = time.time()
        while not os.path.exists(ping) and time.time() - t0 < 15:
            time.sleep(0.5)
        if os.path.exists(ping):
            attivo = _leggi_flag(ping)
            try:
                os.remove(ping)
            except OSError:
                pass
            if attivo == "0":
                break
    else:
        raise RuntimeError(
            "la riga di comando di AutoCAD non risponde o c'e' un comando "
            "a meta' che gli ESC automatici non riescono a chiudere: "
            "premi ESC in AutoCAD, chiudi eventuali finestre di dialogo, "
            f"poi riprova. [{err_msg}]")
    ff = _flag_nuovo()
    send(doc, lisp_code + _scrivi_flag(ff))
    t0 = time.time()
    probe = time.time()
    while not os.path.exists(ff):
        if time.time() - t0 > timeout:
            raise RuntimeError(err_msg)
        if time.time() - probe > 10:
            probe = time.time()
            _controlla_vivo(doc, err_msg)
        time.sleep(0.5)
    try:
        os.remove(ff)
    except OSError:
        pass


def conta_entita(doc):
    return com_retry(lambda: doc.ModelSpace.Count, descr="ModelSpace.Count")


def attendi_import(doc, log):
    """Lo STEP viene tradotto in background: aspetta che compaia geometria
    e che il conteggio resti stabile per qualche secondo."""
    t0 = time.time()
    n = 0
    while True:
        n = conta_entita(doc)
        if n > 0:
            break
        if time.time() - t0 > TIMEOUT_IMPORT:
            raise RuntimeError("IMPORT non terminato (nessuna geometria "
                               "comparsa): step troppo grande o import fallito")
        time.sleep(2)
    stabile = 0
    while stabile < 3:
        time.sleep(2)
        m = conta_entita(doc)
        if m == n:
            stabile += 1
        else:
            n = m
            stabile = 0
    log(f"  import completato ({n} entita')")


def elabora_step(acad, step, log):
    base = os.path.splitext(os.path.basename(step))[0]
    e_dwg = os.path.splitext(step)[1].lower() == ".dwg"
    cartella = os.path.dirname(step)
    outdir = os.path.join(cartella, OUT_DIRNAME, base)
    os.makedirs(outdir, exist_ok=True)
    # cartella pezzi PULITA a ogni giro: i dxf di corse precedenti non
    # devono mescolarsi al confronto (i pezzi cambiano tra un giro e
    # l'altro: nomi, ferramenta, geometrie)
    for f in os.listdir(outdir):
        if f.lower().endswith(".dxf"):
            try:
                os.remove(os.path.join(outdir, f))
            except OSError:
                pass
    dwg = os.path.join(cartella, base + ".dwg")

    if e_dwg:
        # DWG gia' pronto: si apre diretto e NON si salva mai
        # (l'originale resta intatto; ESPLODINOMI lavora solo in memoria)
        doc = com_retry(lambda: acad.Documents.Open(step), descr="apri DWG")
    else:
        doc = com_retry(lambda: acad.Documents.Add(), descr="nuovo disegno")
    try:
        invia_esc()
        send(doc, '(setvar "FILEDIA" 0)(setvar "CMDECHO" 0) ')

        carichi = "".join(f'(load "{fwd(os.path.join(LISP_DIR, f))}") '
                          for f in LISP_FILES)
        send_flag(doc, carichi, TIMEOUT_LOAD,
                  "caricamento LISP fallito: controlla che "
                  f"{LISP_DIR} sia nei percorsi attendibili (TRUSTEDPATHS)")
        if MODO_GENERICO:
            send(doc, '(setq *ESP-FRONT-NO-UNIONE* T) ')   # frontali NON uniti
        log("  lisp caricati")

        if not e_dwg:
            send(doc, f'(command "_.IMPORT" "{fwd(step)}") ')
            attendi_import(doc, log)

            com_retry(lambda: doc.SaveAs(dwg), descr="SaveAs DWG")
            log(f"  DWG salvato: {os.path.basename(dwg)}")
        else:
            log("  DWG aperto diretto (niente import)")

        extra = [] if e_dwg else _ferr_extra_da_stp(step)
        lisp_extra = " ".join(
            '"' + n.replace("\\", "\\\\").replace('"', '\\"') + '"'
            for n in extra)
        send(doc, f"(setq *BN-FERR-EXTRA* (list {lisp_extra})) ")
        if extra:
            log(f"  ferramenta extra dallo STP: {len(extra)} nomi blocco")
        send_flag(doc, "(espnomi-auto) ", TIMEOUT_NOMI,
                  "ESPLODINOMI non terminato")
        if not e_dwg:
            com_retry(lambda: doc.Save(), descr="salva DWG coi nomi")
            log("  ESPLODINOMI fatto (DWG ri-salvato coi nomi)")
        else:
            log("  ESPLODINOMI fatto (in memoria, DWG originale intatto)")

        send_flag(doc, f'(espsol-core "{fwd(outdir)}/") ', TIMEOUT_ESPSOL,
                  "ESPSOL non terminato")
        n_dxf = len([f for f in os.listdir(outdir)
                     if f.lower().endswith(".dxf")])
        log(f"  ESPSOL fatto: {n_dxf} dxf in {OUT_DIRNAME}\\{base}")
    finally:
        # due try separati: la Close va tentata anche se la send fallisce
        # (senno' documento orfano + FILEDIA a 0 nella sessione dell'utente)
        try:
            invia_esc()
            send(doc, '(setvar "FILEDIA" 1) ')
        except Exception:
            pass
        try:
            com_retry(lambda: doc.Close(False), descr="chiudi disegno")
        except Exception:
            pass


def _ferr_extra_da_stp(stp):
    """Nomi-blocco della ferramenta che in AutoCAD NON si chiamano
    'Ferramenta -- ...': nello STEP il PRODUCT ha due nomi e l'import
    battezza il blocco col SECONDO (es. 'Ferramenta -- 4640.10' ->
    blocco 'PAVANELLO_464010_piedino_incasso_M8'). La lista finisce in
    *BN-FERR-EXTRA* per ESPLODINOMI."""
    import re as _re
    out = []
    try:
        testo = open(stp, encoding="utf-8", errors="replace").read()
        for m in _re.finditer(r"PRODUCT\('([^']*)','([^']*)'", testo):
            f1, f2 = m.group(1), m.group(2)
            if (f1.upper().startswith("FERRAMENTA") and f2
                    and not f2.upper().startswith("FERRAMENTA")
                    and f2.upper() not in out):
                out.append(f2.upper())
    except OSError:
        pass
    return out


def _codice_camera(nome):
    import re as _re
    for parte in _re.split(r"[ _\-.]+", nome):
        if _re.fullmatch(r"[A-Za-z]{1,3}\d{2,4}", parte):
            return parte.upper()
    return nome


def libretti_pdf(base, log):
    """LIBRETTI PDF per ogni mobile: esploso (dallo STEP o dal DWG
    sorgente accanto alla cartella del lavoro) + una scheda quotata per
    programma. Non blocca mai il giro: al peggio avvisa e va avanti."""
    import json as _json
    sys.path.insert(0, r"C:\Users\User\Desktop\CLAUDE\SchedePDF")
    from scheda_pdf import genera as genera_schede
    mf = os.path.join(base, "_manifest.json")
    if not os.path.isfile(mf):
        log("  [!] niente manifest: libretti saltati")
        return
    manifest = _json.load(open(mf, encoding="utf-8"))
    camere = set()
    for m in manifest.values():
        camere.update(m.get("rooms") or m.get("groups") or [])
    sopra = os.path.dirname(base)
    cartelle = [d for d in os.listdir(base)
                if os.path.isdir(os.path.join(base, d))
                and d != "PROGRAMMI_UNICI"]
    sorgenti = {f.lower(): f for f in os.listdir(sopra)}
    log(f"  LIBRETTI PDF ({len(camere)} mobili):")

    def _norm(s):
        return "".join(c for c in str(s).lower() if c.isalnum())

    def _trova(nome_base, est):
        """il 3D sorgente accanto al lavoro: prima il nome esatto, poi il
        confronto NORMALIZZATO (spazi, underscore, trattini e maiuscole non
        contano). Prima si cercava solo il nome esatto della cartella: se il
        file 3D era scritto anche di un solo carattere diverso, step e dwg
        restavano None e il libretto usciva senza esploso, senza dirlo."""
        f = sorgenti.get((nome_base + est).lower())
        if f:
            return os.path.join(sopra, f)
        atteso = _norm(nome_base) + _norm(est)
        for basso, vero in sorgenti.items():
            if _norm(basso) == atteso:
                return os.path.join(sopra, vero)
        return None

    for cam in sorted(camere):
        step = dwg = None
        # candidati: la cartella del mobile dentro il lavoro E il nome della
        # camera (il 3D sorgente puo' chiamarsi come l'una o come l'altra)
        cand = [d for d in cartelle if d == cam or _codice_camera(d) == cam]
        if cam not in cand:
            cand.append(cam)
        for nb in cand:
            for est in (".stp", ".step"):
                step = step or _trova(nb, est)
            # il DWG lo prendo SEMPRE (anche se c'e' lo step): lo step
            # serve all'esploso, il DWG agli SCREENSHOT 3D del finito
            dwg = dwg or _trova(nb, ".dwg")
        log(f"    {cam}: step={os.path.basename(step) if step else '-'}"
            f"  dwg={os.path.basename(dwg) if dwg else '-'}")
        try:
            genera_schede(base, cam, step=step, dwg=dwg,
                          ripiego_dwg=MODO_GENERICO,
                          log=lambda m: log("  " + str(m)))
        except Exception as ex:
            log(f"    [!] libretto {cam} fallito: {ex}")


def _codice_commessa(base):
    """codice commessa (es. 25-A018) dal nome cartella lavoro o dal padre."""
    import re as _re
    for parte in (os.path.basename(base),
                  os.path.basename(os.path.dirname(base))):
        m = _re.search(r"\d{2}-A\d{3}", str(parte))
        if m:
            return m.group(0)
    return os.path.basename(os.path.dirname(base)) or ""


def viste_cliente_pdf(base, log):
    """VISTE 2D PER IL CLIENTE dal 3D di ogni mobile: pianta, prospetto e 2
    sezioni verticali (in X e in Y), quotate (ingombri + vani interni +
    spessori) e coi materiali, salvate in PDF (stampa) e DWG (scala reale,
    quote AutoCAD vere) nella sottocartella VISTE. Non blocca mai il giro."""
    import json as _json
    try:
        sys.path.insert(0, r"C:\Users\User\Desktop\CLAUDE\Viste")
        from viste_cliente import genera_viste_cliente
    except Exception as ex:
        log(f"  [!] modulo viste_cliente non caricato: {ex}")
        return
    mf = os.path.join(base, "_manifest.json")
    if not os.path.isfile(mf):
        log("  [!] niente manifest: viste cliente saltate")
        return
    manifest = _json.load(open(mf, encoding="utf-8"))
    camere = set()
    for m in manifest.values():
        camere.update(m.get("rooms") or m.get("groups") or [])
    sopra = os.path.dirname(base)
    cartelle = [d for d in os.listdir(base)
                if os.path.isdir(os.path.join(base, d))
                and d != "PROGRAMMI_UNICI"]
    sorgenti = {f.lower(): f for f in os.listdir(sopra)}
    out_dir = os.path.join(base, "VISTE")
    commessa = _codice_commessa(base)
    log(f"  VISTE CLIENTE ({len(camere)} mobili) -> cartella VISTE:")
    for cam in sorted(camere):
        dwg = None
        for d in cartelle:
            if d == cam or _codice_camera(d) == cam:
                f = sorgenti.get((d + ".dwg").lower())
                if f:
                    dwg = os.path.join(sopra, f)
                break
        if not dwg:
            log(f"    [!] {cam}: niente DWG 3D d'assieme, vista saltata")
            continue
        try:
            genera_viste_cliente(dwg, out_dir, commessa=commessa, mobile=cam,
                                 fai_dwg=True,
                                 log=lambda m: log("  " + str(m)))
        except Exception as ex:
            log(f"    [!] viste {cam} fallite: {ex}")


def confronto_e_sezionatura(base, log):
    """Confronto + sezionatura + TLF/MPR/MPRX + libretti PDF."""
    res = _confronto_core(base, log)
    try:
        libretti_pdf(base, log)
    except Exception as ex:
        log(f"  [!] libretti PDF falliti: {ex}")
    try:
        ferramenta_da_stp(base, log)
    except Exception as ex:
        log(f"  [!] lista ferramenta fallita: {ex}")
    if MODO_GENERICO:
        fogli = {}
        try:
            from nesting_pannelli import nesting_lavoro
            fogli, _pdf = nesting_lavoro(base, log)
        except Exception as ex:
            log(f"  [!] nesting non riuscito: {ex}")
        try:
            from scheda_base_da_filiera import (compila_scheda_base,
                                                genera_ordini_fornitori)
            ferr = _ferramenta_per_scheda(base, log)
            dst = compila_scheda_base(base, log, ferramenta=ferr, fogli=fogli,
                                      fornitori=_ferramenta_fornitori(base))
            # in coda: ordini per FORNITORE (ferramenta/illuminazione/laccatura)
            # + sezionatura raggruppata nel file <scheda>_COMPLETO
            if dst:
                try:
                    genera_ordini_fornitori(dst, log)
                except Exception as ex:
                    log(f"  [!] ordini fornitori non creati: {ex}")
        except Exception as ex:
            log(f"  [!] SCHEDA BASE non compilata: {ex}")
        try:
            viste_cliente_pdf(base, log)
        except Exception as ex:
            log(f"  [!] viste cliente fallite: {ex}")
    # dove e' finito tutto: l'output sta SOLO qui dentro, niente file sparsi
    log(f"  TUTTO IN: {base}")
    log("     <mobile>\\        dxf dei pezzi + TLF / MPR / MPRX")
    log("     PROGRAMMI_UNICI\\ programmi senza doppioni + riepilogo")
    log("     VISTE\\           TAVOLE PER IL CLIENTE (pdf + dwg quotato)")
    log("     SCHEDE_PDF\\      libretti di officina, uno per mobile")
    if MODO_GENERICO:
        log("     SCHEDA BASE\\     scheda, _COMPLETO e ORDINI\\ fornitori")
        log("     NESTING_*.pdf    piano di taglio")
    return res


def _ferramenta_per_scheda(base, log=print):
    """Componenti sciolti totali (da ordinare) di tutti i mobili del lavoro:
    [(codice, q.ta)] per il blocco ferramenta della SCHEDA BASE.
    PREFERISCE il DWG d'assieme (variante coi solidi ferramenta su layer
    'Ferramenta -- CODICE -- FORNITORE': q.ta = n. solidi, ESATTA, e da' anche
    il FORNITORE); ripiego sul conteggio dallo STP. I fornitori trovati si
    leggono con _ferramenta_fornitori(base)."""
    sopra = os.path.dirname(base)
    tot = {}
    forn = {}
    cartelle = [d for d in os.listdir(base)
                if os.path.isdir(os.path.join(base, d))
                and d not in ("PROGRAMMI_UNICI", "SCHEDE_PDF")]
    sorgenti = {f.lower(): f for f in os.listdir(sopra)}
    ferramenta_da_dwg = None
    try:
        sys.path.insert(0, r"C:\Users\User\Desktop\CLAUDE\SchedePDF")
        from scheda_pdf import ferramenta_da_dwg
    except Exception:
        pass
    for d in cartelle:
        usato = False
        # 1) DWG annotato (variante): codice + FORNITORE + q.ta esatta
        if ferramenta_da_dwg:
            fdwg = sorgenti.get((d + ".dwg").lower())
            if fdwg:
                try:
                    ferr = ferramenta_da_dwg(os.path.join(sopra, fdwg), log)
                except Exception:
                    ferr = {}
                if ferr:
                    for cod, (qta, fn) in ferr.items():
                        tot[cod] = tot.get(cod, 0) + qta
                        if fn:
                            forn[cod] = fn
                    usato = True
        # 2) ripiego: conteggio dallo STP
        if not usato:
            for est in (".stp", ".step"):
                f = sorgenti.get((d + est).lower())
                if f:
                    try:
                        _, comp = _ferramenta_stp(os.path.join(sopra, f))
                        for cod, qta, _dentro in comp:
                            tot[cod] = tot.get(cod, 0) + qta
                    except OSError:
                        pass
                    break
    # SEMPRE (anche {}): niente fornitori "ereditati" da un lavoro precedente
    _ferramenta_per_scheda._fornitori = forn
    if forn:
        log(f"  ferramenta: {len(forn)} codici con FORNITORE dal dwg")
    return sorted(tot.items())


def _ferramenta_fornitori(base):
    """{codice: fornitore} dell'ultima chiamata a _ferramenta_per_scheda
    (dai layer del DWG d'assieme), {} se non disponibili."""
    return dict(getattr(_ferramenta_per_scheda, "_fornitori", {}) or {})


def _ferramenta_stp(stp):
    """Distinta ferramenta a DUE livelli da uno STEP CATIA.
    Legge PRODUCT / PRODUCT_DEFINITION(_FORMATION) / NAUO e MOLTIPLICA le
    quantita' lungo l'albero (una Giunzione montata 22 volte porta 22 di
    ognuno dei suoi componenti). Ritorna (montata, componenti):
      montata    = [(codice, q.ta)]           come si monta sul mobile
      componenti = [(codice, q.ta, dentro_a)] pezzi sciolti totali."""
    import re as _re
    testo = open(stp, encoding="utf-8", errors="replace").read()
    nomi = {}                            # id PRODUCT -> (nome1, nome2)
    for m in _re.finditer(r"#(\d+)\s*=\s*PRODUCT\('([^']*)','([^']*)'",
                          testo):
        nomi[m.group(1)] = (m.group(2), m.group(3))
    pdf2prod = {}
    for m in _re.finditer(
            r"#(\d+)\s*=\s*PRODUCT_DEFINITION_FORMATION[^(]*\("
            r"'[^']*'\s*,\s*'[^']*'\s*,\s*#(\d+)", testo):
        pdf2prod[m.group(1)] = m.group(2)
    pd2prod = {}
    for m in _re.finditer(
            r"#(\d+)\s*=\s*PRODUCT_DEFINITION\("
            r"'[^']*'\s*,\s*'[^']*'\s*,\s*#(\d+)", testo):
        pd2prod[m.group(1)] = pdf2prod.get(m.group(2))
    archi = {}                           # (prodotto padre, figlio) -> n
    for m in _re.finditer(
            r"NEXT_ASSEMBLY_USAGE_OCCURRENCE\('[^']*'\s*,\s*'[^']*'\s*,"
            r"\s*'[^']*'\s*,\s*#(\d+)\s*,\s*#(\d+)", testo):
        pa = pd2prod.get(m.group(1))
        fi = pd2prod.get(m.group(2))
        if pa and fi:
            archi[(pa, fi)] = archi.get((pa, fi), 0) + 1

    figli = {}
    padri = {}
    for (pa, fi), n in archi.items():
        figli.setdefault(pa, []).append((fi, n))
        padri.setdefault(fi, []).append((pa, n))

    def e_ferr(pid):
        return (nomi.get(pid, ("", ""))[0] or "").upper().startswith(
            "FERRAMENTA")

    def codice(pid):
        n1, n2 = nomi.get(pid, ("", ""))
        if "--" in n1:
            return n1.split("--", 1)[1].strip()
        return n1 or n2

    tot = {}                             # q.ta TOTALE moltiplicata

    def totale(pid, visti=()):
        if pid in tot:
            return tot[pid]
        if pid in visti:
            return 0
        pp = padri.get(pid)
        if not pp:
            t = 1                        # radice = l'assieme del mobile
        else:
            t = sum(totale(pa, visti + (pid,)) * n for pa, n in pp)
        tot[pid] = t
        return t

    montata = []                         # figlia diretta del mobile
    componenti = []                      # foglie: pezzi sciolti
    for pid in nomi:
        if not e_ferr(pid):
            continue
        q = totale(pid)
        if q <= 0:
            continue
        if any(not e_ferr(pa) for pa, _ in padri.get(pid, [])):
            montata.append((codice(pid), q))
        if not figli.get(pid):
            dentro = sorted({codice(pa) for pa, _ in padri.get(pid, [])
                             if e_ferr(pa)})
            componenti.append((codice(pid), q, " + ".join(dentro)))
    montata.sort()
    componenti.sort()
    return montata, componenti


def ferramenta_da_stp(base, log):
    """Foglio FERRAMENTA nel riepilogo di PROGRAMMI_UNICI: i blocchi
    'Ferramenta -- codice' contati dagli STP sorgente dei mobili."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    sopra = os.path.dirname(base)
    cartelle = [d for d in os.listdir(base)
                if os.path.isdir(os.path.join(base, d))
                and d not in ("PROGRAMMI_UNICI", "SCHEDE_PDF")]
    sorgenti = {f.lower(): f for f in os.listdir(sopra)}
    dati = []                            # (mobile, montata, componenti)
    for d in sorted(cartelle):
        for est in (".stp", ".step"):
            f = sorgenti.get((d + est).lower())
            if f:
                mon, comp = _ferramenta_stp(os.path.join(sopra, f))
                if mon or comp:
                    dati.append((d, mon, comp))
                break
    if not dati:
        return
    xls = os.path.join(base, "PROGRAMMI_UNICI",
                       "_RIEPILOGO_programmi_unici.xlsx")
    if not os.path.isfile(xls):
        xls = os.path.join(base, "PROGRAMMI_UNICI",
                           "_RIEPILOGO_programmi_unici_NEW.xlsx")
    if not os.path.isfile(xls):
        return
    wb = openpyxl.load_workbook(xls)
    if "FERRAMENTA" in wb.sheetnames:
        del wb["FERRAMENTA"]
    ws = wb.create_sheet("FERRAMENTA")
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9E1F2")
    fill2 = PatternFill("solid", fgColor="E2EFDA")
    for c, t in enumerate(("MOBILE", "CODICE FERRAMENTA", "Q.TA",
                           "DENTRO A"), 1):
        cl = ws.cell(1, c, t)
        cl.font = bold
        cl.fill = fill
    # SOLO i pezzi singoli da ordinare (foglie con le q.ta totali
    # moltiplicate lungo l'albero); "DENTRO A" ricorda da dove vengono
    i = 2
    n_voci = 0
    for mob, mon, comp in dati:
        for cod, qta, dentro in comp:
            ws.cell(i, 1, mob)
            ws.cell(i, 2, cod)
            ws.cell(i, 3, qta)
            ws.cell(i, 4, dentro)
            i += 1
            n_voci += 1
    for col, w in zip("ABCD", [44, 44, 7, 40]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    try:
        wb.save(xls)
        log(f"  FERRAMENTA dal 3D: {n_voci} pezzi da ordinare")
    except PermissionError:
        log("  [!] riepilogo aperto in Excel: foglio FERRAMENTA non "
            "scritto (rilancia a file chiuso)")


def _confronto_core(base, log):
    """Sulle sottocartelle di <base> (= Programmi CNC, una per mobile):
    - modo 21032:   confronta_file_unici_dxf (regole CNC) -> PROGRAMMI_UNICI
                    + _RIEPILOGO, poi genera_sezionatura -> _LAVORATO.xlsx
    - modo GENERICO: file_unici_dxf_generico -> PROGRAMMI_UNICI + _RIEPILOGO
                    semplice, NIENTE sezionatura."""
    if MODO_GENERICO:
        from file_unici_dxf_generico import detect_groups, run_comparison
        groups = detect_groups(base)
        if not groups:
            raise RuntimeError(f"nessuna cartella con dxf dentro {base}")
        log(f"  confronto GENERICO su {len(groups)} cartelle: "
            + ", ".join(g["label"] for g in groups))
        res = run_comparison(base, groups, log=lambda m: log("  " + str(m)))
        log(f"  programmi unici: {res['unici']} (da {res['file_letti']} file)")
        if res.get("bloccati"):
            log(f"  [!] file bloccati non aggiornati: "
                + ", ".join(res["bloccati"]))

        # --- lavori generici: niente regole CNC, quindi OGNI programma
        #     unico esce sia in TLF (Masterwood) sia in MPR+MPRX (HOMAG);
        #     in officina si usa quello della macchina giusta
        from dxf2tlf_masterwood import converti as converti_tlf
        from dxf2mpr_homag import converti as converti_mpr
        out_dir = res["out"]
        dxfs = sorted(f for f in os.listdir(out_dir)
                      if f.lower().endswith(".dxf"))
        if dxfs:
            tlf_dir = os.path.join(out_dir, "TLF")
            mpr_dir = os.path.join(out_dir, "MPR")
            mprx_dir = os.path.join(out_dir, "MPRX")
            os.makedirs(tlf_dir, exist_ok=True)
            os.makedirs(mpr_dir, exist_ok=True)
            os.makedirs(mprx_dir, exist_ok=True)
            log(f"  TLF + MPR/MPRX per {len(dxfs)} programmi unici:")
            n2d = 0
            for fn in dxfs:
                p = os.path.join(out_dir, fn)
                try:
                    converti_tlf(p, tlf_dir, log)
                except Exception as e:
                    if "3DSOLID" in str(e):
                        n2d += 1
                        continue               # 2D: inutile tentare l'MPRX
                    log(f"    [ERRORE TLF] {fn}: {e}")
                try:
                    converti_mpr(p, mpr_dir, log, out_mprx=mprx_dir)
                except Exception as e:
                    log(f"    [ERRORE MPRX] {fn}: {e}")
            if n2d:
                log(f"  {n2d} dxf 2D senza solido: saltati "
                    "(TLF/MPRX si fanno solo dal 3D)")
        return res

    from confronta_file_unici_dxf import detect_rooms, run_comparison
    from genera_sezionatura_cp_21032 import process_source

    rooms = detect_rooms(base)
    if not rooms:
        raise RuntimeError(f"nessuna cartella con dxf dentro {base}")
    log(f"  confronto su {len(rooms)} cartelle: "
        + ", ".join(r["code"] for r in rooms))
    res = run_comparison(base, rooms, log=lambda m: log("  " + str(m)))
    log(f"  programmi unici: {res['unici']} (da {res['file_letti']} file)")
    if res.get("bloccati"):
        log(f"  [!] file bloccati non aggiornati: {', '.join(res['bloccati'])}")

    riepilogo = os.path.join(res["out"], res["report"])
    lavorato = os.path.splitext(riepilogo)[0] + "_LAVORATO.xlsx"
    try:
        if os.path.exists(lavorato):
            os.remove(lavorato)
        n, k = process_source(riepilogo, lavorato)
        log(f"  sezionatura: {os.path.basename(lavorato)} "
            f"({n} righe BOM, {k} righe sezionatura)")
    except PermissionError:
        raise RuntimeError(f"{os.path.basename(lavorato)} aperto in Excel: "
                           "chiudilo e rilancia")

    # --- TLF Masterwood: SOLO i pezzi con MASTER nella colonna CNC --------
    from dxf2tlf_masterwood import converti as converti_tlf
    import openpyxl
    wb = openpyxl.load_workbook(riepilogo, read_only=True)
    ws = wb.active
    hdr = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        i_file, i_cnc = hdr.index("FILE PROGRAMMA"), hdr.index("CNC")
    except ValueError:
        wb.close()
        log("  [!] riepilogo senza colonna CNC: TLF saltati")
        return res
    righe = [(str(r[i_file]), str(r[i_cnc] or "").upper())
             for r in ws.iter_rows(min_row=2, values_only=True)
             if r and r[i_file]]
    wb.close()
    da_fare = [f for f, cnc in righe if "MASTER" in cnc]
    da_homag = [f for f, cnc in righe if "HOMAG" in cnc]
    if da_fare:
        tlf_dir = os.path.join(res["out"], "TLF")
        os.makedirs(tlf_dir, exist_ok=True)
        log(f"  TLF Masterwood ({len(da_fare)} pezzi con CNC MASTER):")
        for fn in da_fare:
            p = os.path.join(res["out"], str(fn))
            if not os.path.exists(p):
                log(f"    [!] {fn}: file non trovato, saltato")
                continue
            try:
                converti_tlf(p, tlf_dir, log)
            except Exception as e:
                log(f"    [ERRORE TLF] {fn}: {e}")
    else:
        log("  nessun pezzo con CNC MASTER: niente TLF")

    # --- MPR + MPRX woodWOP: SOLO i pezzi con HOMAG nella colonna CNC -----
    # (conversione DIRETTA: .mpr scritto dalla geometria del solido +
    #  TechAutoX che genera il .mprx; niente GUI, niente mani sul PC)
    if da_homag:
        from dxf2mpr_homag import converti as converti_mpr
        mpr_dir = os.path.join(res["out"], "MPR")
        mprx_dir = os.path.join(res["out"], "MPRX")
        os.makedirs(mpr_dir, exist_ok=True)
        os.makedirs(mprx_dir, exist_ok=True)
        log(f"  MPR+MPRX woodWOP ({len(da_homag)} pezzi con CNC HOMAG):")
        for fn in da_homag:
            p = os.path.join(res["out"], fn)
            if not os.path.exists(p):
                log(f"    [!] {fn}: file non trovato, saltato")
                continue
            try:
                converti_mpr(p, mpr_dir, log, out_mprx=mprx_dir)
            except Exception as e:
                log(f"    [ERRORE MPRX] {fn}: {e}")
    else:
        log("  nessun pezzo con CNC HOMAG: niente MPRX")
    return res


# ============================== GUI ========================================
class App:
    def __init__(self, root):
        self.root = root
        root.title("FILIERA UN CLIC  |  "
                   + ("LAVORO GENERICO (qualsiasi commessa)" if MODO_GENERICO
                      else "21032 (frontali uniti + sezionatura)")
                   + "  |  STEP -> DWG + DXF + programmi macchina")
        root.geometry("720x480")
        self.files = []

        tk.Button(root, text="1)  Scegli i file STEP (o DWG gia' pronti)...",
                  font=("Segoe UI", 11),
                  command=self.scegli).pack(fill="x", padx=10, pady=(10, 4))
        self.lista = tk.Listbox(root, height=8)
        self.lista.pack(fill="both", expand=False, padx=10)
        self.btn = tk.Button(root, text="2)  AVVIA", font=("Segoe UI", 12, "bold"),
                             bg="#2e7d32", fg="white", command=self.avvia)
        self.btn.pack(fill="x", padx=10, pady=6)
        self.log_box = tk.Text(root, height=12, state="disabled",
                               font=("Consolas", 9))
        self.log_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.log("Seleziona gli STEP (o DWG gia' pronti) e premi AVVIA.")
        self.log("AutoCAD deve essere APERTO; durante il lavoro non toccarlo.")
        self.log("Coi DWG: niente import e niente salvataggio, "
                 "l'originale resta intatto.")

    def log(self, msg):
        def _do():
            self.log_box.config(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.root.after(0, _do)

    def scegli(self):
        sel = filedialog.askopenfilenames(
            title="Scegli i file STEP o DWG",
            filetypes=[("STEP o DWG", "*.stp *.step *.dwg"),
                       ("STEP", "*.stp *.step"), ("DWG", "*.dwg"),
                       ("Tutti i file", "*.*")])
        if sel:
            self.files = list(sel)
            self.lista.delete(0, "end")
            for f in self.files:
                self.lista.insert("end", f)

    def avvia(self):
        if not self.files:
            messagebox.showwarning("Attenzione", "Prima scegli i file STEP.")
            return
        self.btn.config(state="disabled", text="Elaborazione in corso...")
        threading.Thread(target=self.lavora, daemon=True).start()

    def lavora(self):
        import pythoncom
        pythoncom.CoInitialize()
        ok, errori = 0, []
        cartelle = []          # cartelle di lavoro degli step riusciti
        try:
            acad = get_acad()
            self.log("nota: AVVIA annulla eventuali comandi a meta' "
                     "in AutoCAD (ESC automatico)")
            sblocca_attivo(acad)
            for step in self.files:
                nome = os.path.basename(step)
                self.log(f"--- {nome} ---")
                try:
                    try:
                        elabora_step(acad, step, self.log)
                    except Exception as e1:
                        # AutoCAD a volte rifiuta le chiamate COM se ancora
                        # indaffarato: riaggancia e riprova una volta
                        self.log(f"  [AVVISO] {e1}")
                        self.log("  riaggancio AutoCAD e riprovo...")
                        time.sleep(5)
                        acad = get_acad()
                        sblocca_attivo(acad)
                        elabora_step(acad, step, self.log)
                    ok += 1
                    c = os.path.dirname(step)
                    if c not in cartelle:
                        cartelle.append(c)
                except Exception as e:
                    errori.append(f"{nome}: {e}")
                    self.log(f"  [ERRORE] {e}")
        except Exception as e:
            errori.append(f"AutoCAD: {e}")
            self.log(f"[ERRORE] {e}")
            self.log(traceback.format_exc())
        finally:
            pythoncom.CoUninitialize()

        # --- giro finale: confronto file unici + sezionatura ----------------
        for c in cartelle:
            base = os.path.join(c, OUT_DIRNAME)
            fase = "CONFRONTO" if MODO_GENERICO else "CONFRONTO + SEZIONATURA"
            self.log(f"--- {fase}: {base} ---")
            try:
                confronto_e_sezionatura(base, self.log)
            except Exception as e:
                errori.append(f"confronto/sezionatura ({base}): {e}")
                self.log(f"  [ERRORE] {e}")
        self.log(f"\n>>> Finito: {ok}/{len(self.files)} step elaborati. <<<")
        if errori:
            self.log("Errori:\n  " + "\n  ".join(errori))

        def _fine():
            self.btn.config(state="normal", text="2)  AVVIA")
            if errori:
                messagebox.showwarning(
                    "Completato con errori",
                    f"{ok}/{len(self.files)} step elaborati.\n\n"
                    + "\n".join(errori[:10]))
            else:
                messagebox.showinfo("Fatto",
                                    f"{ok}/{len(self.files)} step elaborati.")
        self.root.after(0, _fine)


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
