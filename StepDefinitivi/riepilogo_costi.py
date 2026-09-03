# -*- coding: utf-8 -*-
"""Foglio RIEPILOGO COSTI per MOBILE, diviso in macrocategorie, dentro il
_COMPLETO della SCHEDA BASE. Ripartizione ROBUSTA (per quota del driver):
- Pannellame  -> quota MQ del mobile        (totale = blocco pannellame)
- Bordi       -> quota ML bordi del mobile   (totale = blocco bordi)
- Massello    -> quota MC del mobile          (totale = blocco massello)
- Ferramenta  -> quota n. pezzi (approssimata; la ferramenta non e' in distinta)
- Laccatura   -> ESATTA per pezzo (colonne AL-AX del GENERALE x prezzo letto
                 dalla sezione VERNICIATURA; carteggiatura AW/AX inclusa)
- Manodopera  -> ore modello (0,442 h/pz + 0,448 h/MQ) x EUR/h (cella editabile)

Metodo Guido per la laccatura (formule AL-AX della distinta, gia' calcolate):
  MQ = (F*G*AK + (F+G)*H*2)/1e6*D  -> faccia x n.lati(AK) + sviluppo bordi (una
       sola volta anche su 2 lati: 1000x1000 sp20 -> 1 lato 1,08 / 2 lati 2,08)
  ML = MAX(F/1000, 1)*D            -> minimo 1 m per pezzo
       (eccezione M17-ML/colonna AV: minimo 0,25 m = 250 mm)
  CART.RA ML (AW) = somma colonne ml;  CART.RA MQ (AX) = somma colonne mq
Il minimo vive nella formula della colonna nello STAMPO (non in Python): il
riepilogo legge la quantita' gia' calcolata.
Il prezzo €/CAD di ogni voce sta nella sezione VERNICIATURA della commessa
(descrizione col C == intestazione riga 3 della colonna): lo si LEGGE da li',
cosi' i colleghi lo aggiornano nel file. I default sotto servono solo se una
voce ha quantita' ma la VERNICIATURA non ha ancora il prezzo.
"""
import os
from collections import defaultdict

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

LACC_COL_MIN, LACC_COL_MAX = 38, 50        # AL..AX (49=CART.RA ML, 50=CART.RA MQ)
LACC_DEFAULT = [                           # (parola chiave header, um|None, €)
    ("CART", "ML", 3.8), ("CART", "MQ", 7.7),
    ("M 17", "MQ", 85.0), ("M 17", "ML", 35.0),
    ("EFFETTO ROCCIA", None, 94.5),
    ("TINTO", "MQ", 25.8), ("TINTO", "ML", 4.7),
    ("ALLUMINIO", "ML", 6.6),
    ("ANTRACITE", "ML", 5.9),
    ("LACCATO", "MQ", 36.0), ("LACCATO", "ML", 6.6),
]
CATS = ["PANNELLAME", "FERRAMENTA", "BORDI", "MASSELLO/CORNICI",
        "LACCATURA", "ILLUMINAZIONE"]


def _f(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def _codici_mobile(nome):
    """Codici mobile (es. 'P','L2','T','T1') dal nome di distinta:
    MOBILE_P->P, PANNELLATURA_L2->L2, V-ARMADIO CAMERA->V, Pannello X->X,
    MOBILE_T+T1->[T,T1]."""
    import re
    s = str(nome).strip().upper()
    if "_" in s:
        cod = s.rsplit("_", 1)[1]
    elif "-" in s:
        cod = s.split("-", 1)[0]
    elif " " in s:
        cod = s.rsplit(" ", 1)[1]
    else:
        cod = s
    return [c for c in re.split(r"[+\s/]", cod) if c]


def _codice_illum(nome):
    """Codice mobile in testa alla voce illuminazione:
    'L2 - PANN...'->'L2', 'P-PARETE LAVABI'->'P', 'W-CABINA ARMADIO'->'W'."""
    import re
    s = str(nome).strip().upper()
    return re.split(r"[-\s]", s, maxsplit=1)[0].strip()


def _illuminazione_per_nome(G):
    """Somma €TOT (col 7) per nome-mobile nella sezione ILLUMINAZIONE del
    GENERALE (col A = mobile). Ritorna {} se assente o vuota."""
    titolo = None
    for r in range(1, G.max_row + 1):
        v = G.cell(r, 1).value
        if v and str(v).strip().upper() == "ILLUMINAZIONE":
            titolo = r
            break
    if not titolo:
        return {}
    out = {}
    r = titolo + 2                       # salta titolo + header
    fine = min(G.max_row, titolo + 80)
    while r <= fine:
        a = G.cell(r, 1).value
        if a and str(a).strip().upper().startswith("TOTALE"):
            break
        if a and str(a).strip():
            mob = str(a).strip()
            try:
                e = float(G.cell(r, 7).value)
            except (TypeError, ValueError):
                e = 0.0
            if e:
                out[mob] = out.get(mob, 0.0) + e
        r += 1
    return out


def _norm(s):
    return " ".join(str(s).split()).upper() if s not in (None, "") else ""


def _um_header(hdr):
    h = _norm(hdr)
    if h.endswith(" ML") or h.endswith("_ML") or h == "ML":
        return "ML"
    return "MQ"                       # 'effetto roccia' e simili -> mq


def _prezzo_default(hdr):
    h = _norm(hdr)
    um = _um_header(hdr)
    for kw, u, pr in LACC_DEFAULT:
        if kw in h and (u is None or u == um):
            return pr
    return 0.0


def _prezzi_laccatura(G):
    """Ritorna (prezzo{col->€}, vern{descr->€}) per le colonne AL..AX (38..50).
    Il prezzo si legge dalla sezione VERNICIATURA (descrizione col C ==
    intestazione riga 3 della colonna); se manca si usa il default."""
    vern = {}
    tr = None
    for r in range(1, G.max_row + 1):
        if _norm(G.cell(r, 1).value) == "VERNICIATURA":
            tr = r
            break
    if tr:
        r = tr + 1
        while r < tr + 60:
            if _norm(G.cell(r, 1).value).startswith("TOTALE") and r > tr + 1:
                break
            desc = _norm(G.cell(r, 3).value)
            pr = G.cell(r, 6).value
            if desc and isinstance(pr, (int, float)):
                vern[desc] = float(pr)
            r += 1
    prezzo = {}
    for c in range(LACC_COL_MIN, LACC_COL_MAX + 1):
        hdr = G.cell(3, c).value
        if not hdr:
            prezzo[c] = 0.0
            continue
        prezzo[c] = vern.get(_norm(hdr))
        if prezzo[c] is None:
            prezzo[c] = _prezzo_default(hdr)
    return prezzo, vern


def genera_riepilogo_costi(complet_path, eur_ora=45.0, log=print):
    """Aggiunge il foglio 'RIEPILOGO COSTI' al file _COMPLETO. Ritorna il
    path o None. Il _COMPLETO ha gia' i VALORI (l'automazione lo salva
    data_only), quindi si legge senza data_only."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(
            complet_path, keep_vba=complet_path.lower().endswith(".xlsm"))
    except Exception as ex:
        log(f"    [!] riepilogo costi: file non aperto ({ex})")
        return None
    if "GENERALE" not in wb.sheetnames:
        return None
    G = wb["GENERALE"]

    # --- driver per mobile dalla distinta (righe 4-338) --------------------
    mobili = []
    pz = defaultdict(float)
    mq = defaultdict(float)
    bordo = defaultdict(float)
    mc = defaultdict(float)
    lacc = defaultdict(float)                 # euro laccatura per mobile
    prezzo_lacc, vern = _prezzi_laccatura(G)  # €/CAD per colonna AL..AX
    qty_lacc = defaultdict(float)             # q.ta per colonna (per la tabella)
    for r in range(4, 339):
        mob = G.cell(r, 2).value
        if not mob:
            continue
        mob = str(mob).strip()
        if mob not in pz:
            mobili.append(mob)
        pz[mob] += _f(G.cell(r, 4).value)                 # D  q.ta
        mq[mob] += _f(G.cell(r, 33).value)                # AG MQ TOTALI
        for c in range(12, 19):                           # L..R bordi (ML)
            bordo[mob] += _f(G.cell(r, c).value)
        mc[mob] += _f(G.cell(r, 21).value)                # U  MC MASSELLO
        for c in range(LACC_COL_MIN, LACC_COL_MAX + 1):   # AL..AX laccat+cart.ra
            v = _f(G.cell(r, c).value)
            if v:
                lacc[mob] += v * prezzo_lacc.get(c, 0.0)
                qty_lacc[c] += v
    # diagnostica: voci con quantita' ma senza prezzo
    manca = [_norm(G.cell(3, c).value) for c in qty_lacc
             if qty_lacc[c] > 0 and not prezzo_lacc.get(c)]
    if manca:
        log("    [!] laccatura senza prezzo (metti €/CAD in VERNICIATURA): "
            + ", ".join(manca))
    if qty_lacc:
        log(f"  laccatura+carteggiatura: {sum(lacc.values()):.0f} EUR")

    # --- totali categoria: uso le CELLE "TOTALE" dei blocchi (senza sommare
    # le righe voce + la riga totale, che raddoppierebbe) -------------------
    tot_pann = _f(G.cell(369, 24).value)      # X369 (TOTALE pannellame)
    tot_mass = _f(G.cell(386, 19).value)      # S386 (TOTALE massello/cornici)
    tot_bordi = _f(G.cell(399, 19).value)     # S399 (TOTALE bordi)
    tot_ferr = sum(_f(G.cell(r, 7).value) for r in range(345, 369))    # G voci
    TMQ = sum(mq.values()) or 1.0
    TBO = sum(bordo.values()) or 1.0
    TMC = sum(mc.values()) or 1.0
    TPZ = sum(pz.values()) or 1.0

    # --- illuminazione: sezione per-mobile del GENERALE (col A nome, col G €),
    # abbinata al mobile di distinta per codice (es. 'P-...' -> MOBILE_P) -----
    illum_raw = _illuminazione_per_nome(G)
    code2mob = defaultdict(list)
    for m in mobili:
        for c in _codici_mobile(m):
            code2mob[c].append(m)
    ill_mob = defaultdict(float)
    ill_extra = 0.0
    for nome_ill, euro in illum_raw.items():
        cand = code2mob.get(_codice_illum(nome_ill), [])
        if len(cand) == 1:
            ill_mob[cand[0]] += euro
        else:
            ill_extra += euro
    if illum_raw:
        log(f"  illuminazione: {sum(illum_raw.values()):.0f} EUR "
            f"({len(ill_mob)} mobili abbinati, {ill_extra:.0f} EUR da abbinare)")

    # --- foglio ------------------------------------------------------------
    nome = "RIEPILOGO COSTI"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome, 1)
    grigio = PatternFill("solid", fgColor="D8E4BC")
    giallo = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)
    bordo_thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    ws["A1"] = "RIEPILOGO COSTI PER MOBILE (macrocategorie)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Manodopera €/h:"
    ws["A2"].font = bold
    ws["B2"] = eur_ora
    ws["B2"].fill = giallo                     # cella EDITABILE
    ws["B2"].number_format = '#,##0.00 "€"'
    ws["C2"] = "(modificabile: il costo manodopera si aggiorna)"
    ws["C2"].font = Font(italic=True, size=9, color="808080")

    hdr = ["MOBILE"] + CATS + ["MANOD. (h)", "MANOD. €", "TOTALE €"]
    r0 = 4
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r0, j, h)
        c.font = bold
        c.fill = grigio
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = bordo_thin
    ws.column_dimensions["A"].width = 24
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 14

    r = r0 + 1
    for mob in mobili:
        pann = tot_pann * mq[mob] / TMQ
        ferr = tot_ferr * pz[mob] / TPZ
        bor = tot_bordi * bordo[mob] / TBO
        mas = tot_mass * mc[mob] / TMC
        lac = lacc[mob]
        ill = ill_mob.get(mob, 0.0)            # illuminazione abbinata per codice
        ore = 0.442 * pz[mob] + 0.448 * mq[mob]
        ws.cell(r, 1, mob)
        for j, v in enumerate([pann, ferr, bor, mas, lac, ill], 2):
            cc = ws.cell(r, j, round(v, 2))
            cc.number_format = '#,##0.00 "€"'
        ws.cell(r, 8, round(ore, 1)).number_format = '#,##0.0 "h"'
        ws.cell(r, 9).value = f"=H{r}*$B$2"    # manodopera € = ore x €/h
        ws.cell(r, 9).number_format = '#,##0.00 "€"'
        ws.cell(r, 10).value = f"=SUM(B{r}:G{r})+I{r}"   # totale
        ws.cell(r, 10).number_format = '#,##0.00 "€"'
        for j in range(1, 11):
            ws.cell(r, j).border = bordo_thin
        r += 1

    # illuminazione non abbinata a un mobile (il totale resta corretto)
    if ill_extra > 0.01:
        ws.cell(r, 1, "ILLUMINAZIONE (da abbinare)").font = Font(italic=True)
        for j in range(2, 8):
            ws.cell(r, j, 0.0).number_format = '#,##0.00 "€"'
        ws.cell(r, 7).value = round(ill_extra, 2)         # colonna ILLUMINAZIONE
        ws.cell(r, 8, 0.0).number_format = '#,##0.0 "h"'
        ws.cell(r, 9).value = f"=H{r}*$B$2"
        ws.cell(r, 9).number_format = '#,##0.00 "€"'
        ws.cell(r, 10).value = f"=SUM(B{r}:G{r})+I{r}"
        ws.cell(r, 10).number_format = '#,##0.00 "€"'
        for j in range(1, 11):
            ws.cell(r, j).border = bordo_thin
        r += 1

    # riga TOTALE
    ws.cell(r, 1, "TOTALE").font = bold
    for j in range(2, 11):
        col = chr(64 + j)
        cc = ws.cell(r, j)
        cc.value = f"=SUM({col}{r0 + 1}:{col}{r - 1})"
        cc.number_format = ('#,##0.0 "h"' if j == 8 else '#,##0.00 "€"')
        cc.font = bold
        cc.fill = grigio
        cc.border = bordo_thin
    ws.cell(r, 1).fill = grigio
    ws.cell(r, 1).border = bordo_thin

    # tabella prezzi laccatura/carteggiatura usati (lettura VERNICIATURA)
    rr = r + 3
    ws.cell(rr, 1, "PREZZI LACCATURA/CARTEGGIATURA usati "
            "(letti dalla sezione VERNICIATURA del GENERALE; "
            "* = default perche' manca il prezzo)").font = bold
    rr += 1
    for j, h in enumerate(["FINITURA / VOCE", "UM", "€/UM", "Q.TÀ", "€ TOT."], 1):
        c = ws.cell(rr, j, h)
        c.font = bold
        c.fill = grigio
    for c in range(LACC_COL_MIN, LACC_COL_MAX + 1):
        if qty_lacc.get(c, 0) <= 0:
            continue
        rr += 1
        hdr = G.cell(3, c).value
        star = "" if _norm(hdr) in vern else " *"
        ws.cell(rr, 1, str(hdr) + star)
        ws.cell(rr, 2, _um_header(hdr))
        ws.cell(rr, 3, prezzo_lacc[c]).number_format = '#,##0.00 "€"'
        ws.cell(rr, 4, round(qty_lacc[c], 3))
        ws.cell(rr, 5,
                round(qty_lacc[c] * prezzo_lacc[c], 2)
                ).number_format = '#,##0.00 "€"'
    ws.cell(rr + 2, 1,
            "NOTE: pannellame/bordi/massello ripartiti per quota (MQ/ML/MC); "
            "laccatura+carteggiatura ESATTE per pezzo (metodo Guido: sviluppo "
            "bordi 1 volta, min 1 m sui ml), prezzo dalla VERNICIATURA; "
            "ferramenta per quota pezzi; illuminazione dalla sezione "
            "ILLUMINAZIONE, abbinata al mobile per codice.").font = Font(
                italic=True, size=9, color="808080")

    try:
        wb.save(complet_path)
        log(f"  riepilogo costi -> foglio '{nome}' ({len(mobili)} mobili)")
        return complet_path
    except Exception as ex:
        log(f"    [!] riepilogo costi non salvato ({ex})")
        return None
