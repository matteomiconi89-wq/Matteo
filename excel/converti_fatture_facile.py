#!/usr/bin/env python3
"""
Convertitore Fatture XML → Excel
Interfaccia semplice: scegli i file XML e dove salvare
"""

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import tkinter as tk
from tkinter import filedialog, messagebox

def add_invoice_sheet(wb, xml_path, sheet_name):
    """Aggiunge un foglio al workbook con i dati della fattura"""
    
    # Parse XML
    tree = ET.parse(xml_path)
    root = tree.getroot()
    
    # Estrai dati intestazione
    fornitore_elem = root.find('.//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}CedentePrestatore//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}Denominazione')
    fornitore = fornitore_elem.text if fornitore_elem is not None else 'N/D'
    
    cliente_elem = root.find('.//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}CessionarioCommittente//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}Denominazione')
    cliente = cliente_elem.text if cliente_elem is not None else 'N/D'
    
    # Dati generali
    numero_elem = root.find('.//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}DatiGeneraliDocumento/{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}Numero')
    numero = numero_elem.text if numero_elem is not None else os.path.basename(xml_path).split('_')[1].split('.')[0]
    
    data_elem = root.find('.//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}DatiGeneraliDocumento/{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}Data')
    data = data_elem.text if data_elem is not None else 'N/D'
    
    totale_elem = root.find('.//{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}DatiGeneraliDocumento/{http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2}ImportoTotaleDocumento')
    totale = totale_elem.text if totale_elem is not None else '0'
    
    # Crea nuovo foglio
    ws = wb.create_sheet(title=sheet_name)
    
    # Intestazione fattura
    ws['A1'] = f"Fattura n. {numero} del {data}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Fornitore: {fornitore}"
    ws['A3'] = f"Cliente: {cliente}"
    ws['A4'] = f"Totale documento: € {totale}"
    ws['A4'].font = Font(bold=True)
    
    # Header tabella (riga 6)
    headers = ['Cod. Articolo', 'Descrizione', 'Q.tà', 'U.M.', 'Prezzo unit.', 'Totale']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Estrai TUTTE le righe articolo
    righe = root.findall('.//DettaglioLinee')
    row = 7
    
    for riga in righe:
        # Estrai dati
        cod_art_elem = riga.find('.//CodiceValore')
        cod_articolo = cod_art_elem.text if cod_art_elem is not None else ''
        
        desc_elem = riga.find('.//Descrizione')
        descrizione = desc_elem.text if desc_elem is not None else ''
        
        qta_elem = riga.find('.//Quantita')
        quantita = float(qta_elem.text) if qta_elem is not None else None
        
        um_elem = riga.find('.//UnitaMisura')
        um = um_elem.text if um_elem is not None else ''
        
        prezzo_elem = riga.find('.//PrezzoUnitario')
        prezzo_unit = float(prezzo_elem.text) if prezzo_elem is not None else None
        
        totale_elem = riga.find('.//PrezzoTotale')
        totale_riga = float(totale_elem.text) if totale_elem is not None else None
        
        # Scrivi TUTTE le righe
        ws.cell(row=row, column=1, value=cod_articolo)
        
        # Descrizione (in corsivo se è una riga descrittiva senza importo)
        cell_desc = ws.cell(row=row, column=2, value=descrizione)
        if totale_riga is None or totale_riga == 0:
            cell_desc.font = Font(italic=True, color='666666')
        
        if quantita is not None:
            cell = ws.cell(row=row, column=3, value=quantita)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        
        ws.cell(row=row, column=4, value=um)
        
        if prezzo_unit is not None:
            cell = ws.cell(row=row, column=5, value=prezzo_unit)
            cell.number_format = '€ #,##0.00'
            cell.alignment = Alignment(horizontal='right')
        
        if totale_riga is not None:
            cell = ws.cell(row=row, column=6, value=totale_riga)
            if totale_riga == 0:
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '€ #,##0.00'
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Totale finale
    ws.cell(row=row+1, column=5, value='TOTALE IMPONIBILE:').font = Font(bold=True)
    ws.cell(row=row+1, column=5).alignment = Alignment(horizontal='right')
    imponibile_elem = root.find('.//ImponibileImporto')
    imponibile = float(imponibile_elem.text) if imponibile_elem is not None else 0
    cell = ws.cell(row=row+1, column=6, value=imponibile)
    cell.number_format = '€ #,##0.00'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right')
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    
    return numero, row-7

def main():
    # Nascondi la finestra principale di tkinter
    root = tk.Tk()
    root.withdraw()
    
    # Chiedi quali file XML caricare
    messagebox.showinfo("Convertitore Fatture XML → Excel", 
                       "Passo 1: Seleziona i file XML delle fatture")
    
    xml_files = filedialog.askopenfilenames(
        title="Seleziona i file XML delle fatture",
        filetypes=[("File XML", "*.xml"), ("Tutti i file", "*.*")]
    )
    
    if not xml_files:
        messagebox.showwarning("Annullato", "Nessun file selezionato. Operazione annullata.")
        return
    
    # Chiedi dove salvare il file Excel
    messagebox.showinfo("Convertitore Fatture XML → Excel", 
                       f"Passo 2: Scegli dove salvare il file Excel\n\n{len(xml_files)} fatture selezionate")
    
    output_file = filedialog.asksaveasfilename(
        title="Salva il file Excel come...",
        defaultextension=".xlsx",
        filetypes=[("File Excel", "*.xlsx"), ("Tutti i file", "*.*")]
    )
    
    if not output_file:
        messagebox.showwarning("Annullato", "Nessun file di output selezionato. Operazione annullata.")
        return
    
    # Crea workbook unico
    wb = Workbook()
    # Rimuovi il foglio di default
    wb.remove(wb.active)
    
    risultati = []
    errori = []
    
    for idx, xml_path in enumerate(sorted(xml_files), 1):
        try:
            # Nome foglio (max 31 caratteri)
            sheet_name = f"Fatt_{idx:02d}"
            numero, righe = add_invoice_sheet(wb, xml_path, sheet_name)
            risultati.append((numero, righe, os.path.basename(xml_path)))
        except Exception as e:
            errori.append((os.path.basename(xml_path), str(e)))
    
    # Salva file
    wb.save(output_file)
    
    # Mostra risultato
    messaggio = f"✓ Conversione completata!\n\n"
    messaggio += f"Fatture convertite: {len(risultati)}\n"
    messaggio += f"File salvato: {os.path.basename(output_file)}\n\n"
    
    if errori:
        messaggio += f"⚠ Errori ({len(errori)}):\n"
        for file, errore in errori[:5]:  # Mostra max 5 errori
            messaggio += f"  - {file}\n"
    
    messagebox.showinfo("Completato!", messaggio)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        messagebox.showerror("Errore", f"Si è verificato un errore:\n\n{str(e)}")
