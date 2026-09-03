# -*- coding: utf-8 -*-
"""Legge tutte le distinte Excel (foglio SOLIDI) in distinte.json:
{ mobile: [ {codice, materiale, qta, H, L, SP, codice_vecchio, nome_dxf} ] }
E il censimento fori globale da estratto_ferramenta.json."""
import glob
import json
import os
from collections import Counter, defaultdict

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = r"C:\Users\User\Dropbox\STEFANO\25-A019 GUIDO_St Paul (Master Bedroom)\Esecutivi\EXCEL"

distinte = {}
for f in sorted(glob.glob(os.path.join(SRC, "*.xlsx"))):
    nome = os.path.splitext(os.path.basename(f))[0]
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as ex:
        print(f"{nome}: ERRORE {ex}")
        continue
    ws = None
    for cand in ("SOLIDI", "LISTA VECCHIA"):
        if cand in wb.sheetnames:
            ws = wb[cand]
            break
    if ws is None:
        ws = wb.worksheets[0]
    head = [str(c.value).strip().upper() if c.value is not None else "" for c in ws[1]]
    def col(*names):
        for n in names:
            for i, h in enumerate(head):
                if h.startswith(n):
                    return i
        return None
    ic, im, iq = col("CODICE"), col("MATERIALE"), col("Q.TA", "QTA", "QUANT")
    ih, il, isp = col("H."), col("L."), col("SP.")
    icv, idx = col("CODICE VECCHIO"), col("NOME DXF")
    righe = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if ic is None or r[ic] is None:
            continue
        def g(i):
            return r[i] if i is not None and i < len(r) else None
        righe.append({
            "codice": g(ic), "materiale": g(im), "qta": g(iq),
            "H": g(ih), "L": g(il), "SP": g(isp),
            "codice_vecchio": g(icv), "nome_dxf": g(idx),
        })
    distinte[nome] = {"foglio": ws.title, "righe": righe}
    print(f"{nome}: {len(righe)} righe (foglio {ws.title})")

json.dump(distinte, open(os.path.join(BASE, "distinte.json"), "w", encoding="utf-8"),
          indent=1, ensure_ascii=False)

# ---- censimento fori globale ----
est = json.load(open(os.path.join(BASE, "estratto_ferramenta.json"), encoding="utf-8"))
cens = Counter()
prof = defaultdict(Counter)
for fk, v in est.items():
    if "error" in v:
        print("ERRORE", fk, v["error"][:80])
        continue
    for lab, p in v["pezzi"].items():
        for h in p["fori"]:
            cens[h["dia"]] += 1
            prof[h["dia"]][round(h["len"])] += 1

print("\n=== CENSIMENTO DIAMETRI (tutti i file, tutti i pezzi) ===")
for dia, n in sorted(cens.items()):
    top = ", ".join(f"prof{p}x{c}" for p, c in prof[dia].most_common(6))
    print(f"Dia {dia:6.2f}: {n:5d} fori | {top}")
