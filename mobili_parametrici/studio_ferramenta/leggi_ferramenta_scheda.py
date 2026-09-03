# -*- coding: utf-8 -*-
"""Legge il foglio FERRAMENTA di SCHEDA BASE.1.xlsm (sola lettura)."""
import json
import os

import openpyxl

F = r"C:\Users\User\Dropbox\STEFANO\25-A019 GUIDO_St Paul (Master Bedroom)\Preventivi\Costi\SCHEDA BASE.1.xlsm"
wb = openpyxl.load_workbook(F, data_only=True, read_only=True, keep_vba=False)
print("FOGLI:", wb.sheetnames)
nome = next((s for s in wb.sheetnames if "ferram" in s.lower()), None)
print("FOGLIO FERRAMENTA:", nome)
ws = wb[nome]
righe = []
for r in ws.iter_rows(values_only=True):
    vals = ["" if v is None else str(v).strip() for v in r]
    if any(vals):
        righe.append(vals)
print("RIGHE NON VUOTE:", len(righe))
for i, r in enumerate(righe[:80]):
    print(f"{i:3d} | " + " | ".join(v[:38] for v in r[:8]))

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ferramenta_scheda_base.json")
json.dump(righe, open(out, "w", encoding="utf-8"), indent=1, ensure_ascii=False)
print("SCRITTO", out)
