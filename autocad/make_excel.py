import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'PARAMETRI'

YEL = PatternFill('solid', fgColor='FFFF00')
GRY = PatternFill('solid', fgColor='D9D9D9')
BLU = PatternFill('solid', fgColor='BDD7EE')
GRN = PatternFill('solid', fgColor='E2EFDA')
HDR = PatternFill('solid', fgColor='2F5496')

f_title = Font(name='Arial', size=14, bold=True, color='1F3864')
f_sec   = Font(name='Arial', size=12, bold=True, color='FFFFFF')
f_col   = Font(name='Arial', size=9,  bold=True, color='FFFFFF')
f_norm  = Font(name='Arial', size=10)
f_note  = Font(name='Arial', size=9, italic=True, color='595959')
f_key   = Font(name='Arial', size=10, bold=True)

al_c   = Alignment(horizontal='center', vertical='center')
al_l   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
al_lnw = Alignment(horizontal='left',   vertical='center')

thin = Side(border_style='thin', color='BFBFBF')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

MAX_COL = 10  # larghezza massima per merge titoli / vano table

def section_header(row, title, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = f_sec
    c.fill = HDR
    c.alignment = Alignment(horizontal='left', vertical='center')

def col_headers_params(row, with_material=False):
    headers = ['CODICE', 'VALORE', 'DESCRIZIONE', 'VALORI ACCETTATI', 'DEFAULT']
    if with_material:
        headers.append('MATERIALE / LAYER')
    ncols = len(headers)
    for col, txt in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font = f_col
        c.fill = GRY
        c.alignment = al_c
        c.border = brd

def param_row(row, code, value, desc, accepted, default, material=None):
    c1 = ws.cell(row=row, column=1, value=code)
    c1.font = f_key;  c1.border = brd;  c1.alignment = al_lnw

    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = f_norm; c2.fill = YEL; c2.alignment = al_c; c2.border = brd

    c3 = ws.cell(row=row, column=3, value=desc)
    c3.font = f_norm; c3.border = brd; c3.alignment = al_lnw

    c4 = ws.cell(row=row, column=4, value=accepted)
    c4.font = f_note; c4.border = brd; c4.alignment = al_lnw

    c5 = ws.cell(row=row, column=5, value=default)
    c5.font = f_norm; c5.border = brd; c5.alignment = al_c

    if material is not None:
        c6 = ws.cell(row=row, column=6, value=material)
        c6.font = f_norm; c6.fill = YEL
        c6.alignment = al_lnw; c6.border = brd

r = 1

# ── TITOLO ────────────────────────────────────────────────────────
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
c = ws.cell(row=r, column=1, value='ARMADIO PARAMETRICO - Foglio parametri')
c.font = f_title; c.alignment = al_l
r += 1

ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
c = ws.cell(row=r, column=1, value=(
    'Modifica solo le celle GIALLE. '
    'Salva come CSV UTF-8 (File > Salva con nome > CSV UTF-8 delimitato da virgola). '
    'In AutoCAD digita CARICA-ARMADIO.'))
c.font = f_note; c.alignment = al_l
r += 2

# ── DIMENSIONI GLOBALI ────────────────────────────────────────────
section_header(r, '  DIMENSIONI GLOBALI', 5); r += 1
col_headers_params(r); r += 1
param_row(r, 'W',   2400, 'Larghezza totale armadio (mm)', 'es. 1200-3600', 2400); r += 1
param_row(r, 'H',   2400, 'Altezza totale (mm)',           'es. 2000-2800', 2400); r += 1
param_row(r, 'D',    600, 'Profondita (mm)',               'es. 500-700',    600); r += 1

r += 1
# ── SPESSORI PANNELLI + MATERIALI ────────────────────────────────
section_header(r, '  SPESSORI PANNELLI + MATERIALI', 6); r += 1
col_headers_params(r, with_material=True); r += 1

# Ogni riga ha il campo MATERIALE/LAYER (colonna F) editabile
param_row(r, 'SP',   18, 'Spessore struttura carcassa (mm)', '16 / 18 / 22',  18, 'STRUTTURA');    r += 1
param_row(r, 'SA',   22, 'Spessore ante (mm)',               '18 / 22 / 25',  22, 'ANTE');         r += 1
param_row(r, 'SF',    8, 'Spessore fondo carcassa (mm)',     '6 / 8 / 10',     8, 'FONDO');        r += 1
param_row(r, 'SP_C', 15, 'Spessore fianchi giro cassetto (mm)', '12 / 15 / 18', 15, 'CASSETTO');   r += 1
param_row(r, 'SF_C',  6, 'Spessore fondo cassetto (mm)',     '6 / 8',           6, 'FONDO-CASS');  r += 1

# Nota colonna materiale
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1,
    value='MATERIALE / LAYER: nome del layer AutoCAD da usare per ogni tipo di pannello. '
          'Puoi scrivere qualsiasi nome (es. TRUCIOLARE-18, MDF-22, PIOPPO-15, HDF-8). '
          'Il layer viene creato automaticamente in AutoCAD se non esiste.')
c.font = f_note; c.alignment = al_l
r += 2

# ── ANTE ──────────────────────────────────────────────────────────
section_header(r, '  ANTE', 5); r += 1
col_headers_params(r); r += 1
param_row(r, 'TIPO_BATTUTA', 1, 'Tipo sovrapposizione anta',
    '1=In battuta  2=A filo  3=Mezza battuta', 1); r += 1
param_row(r, 'LUCE_ANTE', 3, 'Spazio tra 2 ante (mm)', '2-5', 3); r += 1
param_row(r, 'BLUM_BLOCK', 0, 'Usa blocchi STP cerniere Blum importati?',
    '0=No (modello LISP)  1=Si (BLUM_INTERA_110 / BLUM_MEDIANA_110)', 0); r += 1
param_row(r, 'BLUM_MOVENTO', 0, 'Usa blocchi DWG reali Blum Movento 760?',
    '0=No (modello LISP)  1=Si (DWG: 433_24_984/985)', 0); r += 1

r += 1
# ── RIPIANI E CASSETTI ────────────────────────────────────────────
section_header(r, '  RIPIANI E CASSETTI', 5); r += 1
col_headers_params(r); r += 1
param_row(r, 'N_RIP',  4,   'Numero ripiani per vano', '2-8',      4); r += 1
param_row(r, 'N_CAS',  3,   'Numero cassetti per vano tipo cassetti', '2-6', 3); r += 1
param_row(r, 'H_CAS',  200, 'Altezza cassetto (mm)',   '150-250', 200); r += 1
param_row(r, 'D_BOX',  500, 'Profondita corsa cassetto - guida Blum Movento 760 (mm)',
    '400-550', 500); r += 1

r += 1
# ── VANI ──────────────────────────────────────────────────────────
section_header(r, '  VANI', 5); r += 1
col_headers_params(r); r += 1
param_row(r, 'N_VANI', 3, 'Numero totale vani', '1-10', 3); r += 1

r += 1
# ── TABELLA VANI ──────────────────────────────────────────────────
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
c = ws.cell(row=r, column=1,
    value='TABELLA VANI - Inserisci una riga per ogni vano (aggiungi righe se N_VANI > 3)')
c.font = Font(name='Arial', size=10, bold=True)
c.fill = BLU; c.alignment = al_l
r += 1

vano_cols = ['VANO', 'LARGHEZZA_mm', 'N_ANTE', 'LATO_CERNIERA',
             'TIPO_BATTUTA_VANO', 'TIPO_INTERNI', 'H_FIANCO_mm',
             'N_CAS_VANO', 'H_CAS_VANO', 'D_BOX_VANO']
for col, txt in enumerate(vano_cols, 1):
    c = ws.cell(row=r, column=col, value=txt)
    c.font = f_col; c.fill = GRY; c.alignment = al_c; c.border = brd
r += 1

vani = [
    [1, 800, 2, 0, 1, 1, 0, 0, 0, 0],
    [2, 800, 1, 1, 1, 3, 0, 0, 0, 0],
    [3, 800, 2, 0, 1, 1, 0, 0, 0, 0],
]
for vano in vani:
    for col, val in enumerate(vano, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = f_norm; c.alignment = al_c; c.border = brd
        if col > 1:
            c.fill = YEL
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
c = ws.cell(row=r, column=1, value=(
    'LEGENDA:  N_ANTE: 1=singola 2=doppia  |  LATO_CERNIERA: 0=auto(2 ante) 1=SX 2=DX  |  '
    'TIPO_BATTUTA_VANO: 0=usa globale 1=battuta 2=filo 3=mezza  |  '
    'TIPO_INTERNI: 1=Appendiabiti 2=Ripiani 3=Cassetti+ripiani  |  '
    'H_FIANCO_mm: altezza fianco centrale (0=pieno)  |  '
    'N_CAS/H_CAS/D_BOX_VANO: 0=usa valore globale'))
c.font = f_note; c.alignment = al_l

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
c = ws.cell(row=r, column=1, value='COME USARE:')
c.font = Font(name='Arial', size=10, bold=True)
r += 1

steps = [
    '1)  Compila le celle GIALLE con i tuoi valori',
    '2)  Nella colonna MATERIALE / LAYER scrivi il nome del layer AutoCAD per ogni tipo di pannello',
    '3)  File > Salva con nome > scegli formato CSV UTF-8 (delimitato da virgola)',
    '4)  In AutoCAD: digita APPLOAD e carica crea-armadio.lsp',
    '5)  Digita CARICA-ARMADIO e seleziona il file CSV appena salvato',
    '6)  Per i blocchi Blum: in AutoCAD digita PREPARA-BLUM per le istruzioni complete',
]
for step in steps:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
    c = ws.cell(row=r, column=1, value=step)
    c.font = f_note
    r += 1

# ── Larghezze colonne ─────────────────────────────────────────────
# Col A-F per sezioni param, A-J per tabella vani
col_widths = [16, 14, 42, 52, 10, 20, 14, 12, 12, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 22

out = r'C:\Users\User\Desktop\CLAUDE\autocad\armadio-parametri.xlsx'
wb.save(out)
print('Salvato OK:', out)
