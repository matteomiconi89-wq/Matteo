#!/usr/bin/env python3
"""
AUTOMAZIONI EXCEL - VERSIONE FINALE COMPLETA
Script Python con 3 pulsanti:
1. CREA FOGLI FORNITORI
2. RAGGRUPPA SEZIONE  
3. ELABORAZIONE COMPLETA (entrambe le operazioni in un solo file)

✅ Tutti i file vengono salvati in Downloads con timestamp
✅ Il file originale NON viene mai modificato
✅ Gestione sicura di formule, celle vuote e caratteri speciali
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import datetime
from pathlib import Path

class AutomazioniFornitori:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Automazioni Excel - Versione Finale")
        self.window.geometry("650x720")  # Aumentata altezza per 5 pulsanti
        self.window.resizable(False, False)
        
        self.file_path = tk.StringVar()
        self.setup_ui()
        
    def setup_ui(self):
        """Crea l'interfaccia grafica"""
        
        main_frame = tk.Frame(self.window, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Titolo
        title = tk.Label(
            main_frame,
            text="🔧 AUTOMAZIONI EXCEL",
            font=('Arial', 20, 'bold'),
            bg='#f0f0f0',
            fg='#1F3864'
        )
        title.pack(pady=(0, 20))
        
        # Frame selezione file
        file_frame = tk.LabelFrame(
            main_frame,
            text="📂 Seleziona file Excel",
            font=('Arial', 10, 'bold'),
            bg='#f0f0f0',
            padx=10,
            pady=10
        )
        file_frame.pack(fill=tk.X, pady=(0, 20))
        
        file_entry = tk.Entry(
            file_frame,
            textvariable=self.file_path,
            font=('Arial', 10),
            state='readonly',
            width=55
        )
        file_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        browse_btn = tk.Button(
            file_frame,
            text="Sfoglia...",
            command=self.browse_file,
            font=('Arial', 10),
            bg='#4472C4',
            fg='white',
            cursor='hand2',
            padx=20
        )
        browse_btn.pack(side=tk.LEFT)
        
        # Frame pulsanti
        buttons_frame = tk.Frame(main_frame, bg='#f0f0f0')
        buttons_frame.pack(fill=tk.BOTH, expand=True)
        
        # PULSANTE 1
        btn1_frame = tk.Frame(buttons_frame, bg='#4472C4', relief=tk.RAISED, bd=2)
        btn1_frame.pack(fill=tk.X, pady=8)
        
        btn1 = tk.Button(
            btn1_frame,
            text="📊 FERRAMENTA",
            command=self.crea_fogli_fornitori,
            font=('Arial', 13, 'bold'),
            bg='#4472C4',
            fg='white',
            cursor='hand2',
            height=2
        )
        btn1.pack(fill=tk.X, padx=2, pady=2)
        
        desc1 = tk.Label(
            buttons_frame,
            text="Crea un foglio per ogni fornitore (righe 121-202)",
            font=('Arial', 9, 'italic'),
            bg='#f0f0f0',
            fg='#666666'
        )
        desc1.pack()
        
        # PULSANTE 2 - ILLUMINAZIONE
        btn_ill_frame = tk.Frame(buttons_frame, bg='#FFC000', relief=tk.RAISED, bd=2)
        btn_ill_frame.pack(fill=tk.X, pady=8)
        
        btn_ill = tk.Button(
            btn_ill_frame,
            text="💡 ILLUMINAZIONE FORNITORI",
            command=self.crea_fogli_illuminazione,
            font=('Arial', 13, 'bold'),
            bg='#FFC000',
            fg='white',
            cursor='hand2',
            height=2
        )
        btn_ill.pack(fill=tk.X, padx=2, pady=2)
        
        desc_ill = tk.Label(
            buttons_frame,
            text="Crea fogli illuminazione (righe 222-240)",
            font=('Arial', 9, 'italic'),
            bg='#f0f0f0',
            fg='#666666'
        )
        desc_ill.pack()

        # PULSANTE 3 - LACCATURA
        btn_lacc_frame = tk.Frame(buttons_frame, bg='#7030A0', relief=tk.RAISED, bd=2)
        btn_lacc_frame.pack(fill=tk.X, pady=8)

        btn_lacc = tk.Button(
            btn_lacc_frame,
            text="🎨 LACCATURA / VERNICIATURA",
            command=self.crea_foglio_laccatura,
            font=('Arial', 13, 'bold'),
            bg='#7030A0',
            fg='white',
            cursor='hand2',
            height=2
        )
        btn_lacc.pack(fill=tk.X, padx=2, pady=2)

        desc_lacc = tk.Label(
            buttons_frame,
            text="Estrae pezzi con LACCATO/TINTO/VERNICIATURA/EFFETTO ROCCIA (col. AK-AO)",
            font=('Arial', 9, 'italic'),
            bg='#f0f0f0',
            fg='#666666'
        )
        desc_lacc.pack()

        # PULSANTE 4 - SEZIONATURA
        btn2_frame = tk.Frame(buttons_frame, bg='#70AD47', relief=tk.RAISED, bd=2)
        btn2_frame.pack(fill=tk.X, pady=8)
        
        btn2 = tk.Button(
            btn2_frame,
            text="📋 SEZIONATURA",
            command=self.raggruppa_sezione,
            font=('Arial', 13, 'bold'),
            bg='#70AD47',
            fg='white',
            cursor='hand2',
            height=2
        )
        btn2.pack(fill=tk.X, padx=2, pady=2)
        
        desc2 = tk.Label(
            buttons_frame,
            text="Raggruppa pezzi con L, H, Materiale identici",
            font=('Arial', 9, 'italic'),
            bg='#f0f0f0',
            fg='#666666'
        )
        desc2.pack()
        
        # PULSANTE 4 - COMPLETO
        btn3_frame = tk.Frame(buttons_frame, bg='#FF6B35', relief=tk.RAISED, bd=2)
        btn3_frame.pack(fill=tk.X, pady=8)
        
        btn3 = tk.Button(
            btn3_frame,
            text="⚡ FERRAMENTA + ILLUMINAZIONE + LACCATURA + SEZIONATURA",
            command=self.elaborazione_completa,
            font=('Arial', 14, 'bold'),
            bg='#FF6B35',
            fg='white',
            cursor='hand2',
            height=2
        )
        btn3.pack(fill=tk.X, padx=2, pady=2)
        
        desc3 = tk.Label(
            buttons_frame,
            text="⭐ Esegue ENTRAMBE le operazioni → 1 file con tutto",
            font=('Arial', 9, 'bold'),
            bg='#f0f0f0',
            fg='#FF6B35'
        )
        desc3.pack()
        
        # Footer
        footer = tk.Label(
            main_frame,
            text="✅ File salvati in Downloads • ✅ File originale mai modificato • ✅ Timestamp automatico",
            font=('Arial', 8),
            bg='#f0f0f0',
            fg='#999999'
        )
        footer.pack(side=tk.BOTTOM, pady=(15, 0))
        
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Seleziona il file Excel",
            filetypes=[("File Excel", "*.xlsx *.xlsm"), ("Tutti i file", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
    
    def check_file(self):
        if not self.file_path.get():
            messagebox.showerror("Errore", "Seleziona prima un file Excel!")
            return False
        return True
    
    def salva_in_downloads(self, wb, suffisso):
        """Salva il workbook in Downloads con timestamp"""
        downloads = str(Path.home() / "Downloads")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(self.file_path.get()))[0]
        filename = f"{base_name}_{suffisso}_{timestamp}.xlsx"
        path = os.path.join(downloads, filename)
        
        # IMPORTANTE: Salva come .xlsx per mantenere tutto tranne le macro
        wb.save(path)
        return path, filename
    
    def crea_copia_completa(self, suffisso):
        """Crea una copia COMPLETA del file originale in Downloads mantenendo TUTTO"""
        import shutil
        
        downloads = str(Path.home() / "Downloads")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(self.file_path.get()))[0]
        ext = os.path.splitext(self.file_path.get())[1]  # .xlsx o .xlsm
        filename = f"{base_name}_{suffisso}_{timestamp}{ext}"
        dest_path = os.path.join(downloads, filename)
        
        # Copia TUTTO il file (mantiene menu a tendina, collegamenti, tutto!)
        shutil.copy2(self.file_path.get(), dest_path)
        
        # Apri la copia - IMPORTANTE: keep_vba=True per file .xlsm
        wb = openpyxl.load_workbook(dest_path, data_only=True, keep_vba=True)
        
        return wb, dest_path, filename
    
    def apri_file_sicuro(self):
        """Apre il file creando una COPIA temporanea per non toccare MAI l'originale"""
        import shutil
        import tempfile
        
        # Crea file temporaneo
        original_path = self.file_path.get()
        temp_dir = tempfile.gettempdir()
        temp_filename = f"temp_excel_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_path = os.path.join(temp_dir, temp_filename)
        
        # Copia il file originale nel temporaneo
        shutil.copy2(original_path, temp_path)
        
        # Apri la COPIA mantenendo TUTTO (data_only=False per mantenere formule e collegamenti)
        wb = openpyxl.load_workbook(temp_path, data_only=False, keep_vba=True, keep_links=True)
        
        # Ritorna workbook e path temporaneo (da cancellare dopo)
        return wb, temp_path
    
    def pulisci_temp(self, temp_path):
        """Cancella il file temporaneo dopo l'uso"""
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass  # Non importa se non riesce a cancellare

    def trova_foglio_generale(self, wb):
        """Trova il foglio GENERALE; se non esiste, usa il primo foglio."""
        for ws in wb.worksheets:
            if ws.title.strip().upper() == "GENERALE":
                return ws
        return wb.worksheets[0]

    def trova_range_sezione(self, ws, titolo):
        """Cerca un titolo di sezione (es. 'FERRAMENTA', 'ILLUMINAZIONE') in colonna A
        e restituisce (riga_inizio_dati, riga_fine_dati).
        - Inizio dati = riga del titolo + 2 (salta titolo + riga intestazioni colonne)
        - Fine dati = riga precedente al successivo 'TOTALE' (o max_row se non c'è)
        Ritorna (None, None) se il titolo non viene trovato."""
        titolo_upper = titolo.strip().upper()
        riga_titolo = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and str(val).strip().upper() == titolo_upper:
                riga_titolo = r
                break
        if not riga_titolo:
            return None, None
        start = riga_titolo + 2
        end = ws.max_row
        for r in range(start, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and str(val).strip().upper() == 'TOTALE':
                end = r - 1
                break
        return start, end

    def crea_fogli_fornitori(self):
        """FUNZIONE 1: Crea fogli per ogni fornitore"""
        
        if not self.check_file():
            return
        
        try:
            # PRIMA: Crea copia completa del file in Downloads
            wb, dest_path, filename = self.crea_copia_completa("FORNITORI")

            ws_main = self.trova_foglio_generale(wb)

            start, end = self.trova_range_sezione(ws_main, "FERRAMENTA")
            if not start:
                messagebox.showerror("Errore", "Sezione 'FERRAMENTA' non trovata nel foglio GENERALE!")
                return

            fornitori = {}

            for r in range(start, end + 1):
                desc = str(ws_main.cell(r, 3).value or "").strip()
                forn = str(ws_main.cell(r, 5).value or "").strip()
                
                if forn and desc:
                    if forn not in fornitori:
                        fornitori[forn] = []
                    
                    cod = str(ws_main.cell(r, 2).value or "").strip()
                    
                    try:
                        qta = float(ws_main.cell(r, 4).value or 0)
                    except:
                        qta = 0
                    
                    try:
                        prz = float(ws_main.cell(r, 6).value or 0)
                    except:
                        prz = 0
                    
                    if qta > 0:
                        fornitori[forn].append({'cod': cod, 'desc': desc, 'qta': qta, 'prz': prz})
            
            if not fornitori:
                messagebox.showwarning("Attenzione", f"Nessun fornitore trovato (righe {start}-{end})")
                return
            
            for forn, articoli in fornitori.items():
                forn_clean = forn
                for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    forn_clean = forn_clean.replace(char, '_')
                forn_clean = forn_clean[:31]
                
                if forn_clean in wb.sheetnames:
                    del wb[forn_clean]
                
                ws = wb.create_sheet(title=forn_clean)
                
                headers = ['CODICE', 'DESCRIZIONE', 'Q.TA', '€/CAD', '€ TOT.']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.font = Font(bold=True, size=12)
                    c.fill = PatternFill(start_color='D8E4BC', fill_type='solid')
                    c.alignment = Alignment(horizontal='center')
                
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 65
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 14
                
                agg = {}
                for art in articoli:
                    if art['desc'] in agg:
                        agg[art['desc']]['qta'] += art['qta']
                    else:
                        agg[art['desc']] = art.copy()
                
                row = 2
                for art in agg.values():
                    ws.cell(row, 1, art['cod']).number_format = '@'
                    ws.cell(row, 2, art['desc'])
                    ws.cell(row, 3, art['qta'])
                    ws.cell(row, 4, art['prz']).number_format = '#,##0.000 €'
                    ws.cell(row, 5, f"=C{row}*D{row}").number_format = '#,##0.00 €'
                    row += 1
                
                ws.cell(row, 1, 'TOTALE').font = Font(bold=True, size=12)
                ws.cell(row, 3, f"=SUM(C2:C{row-1})").font = Font(bold=True)
                ws.cell(row, 5, f"=SUM(E2:E{row-1})").number_format = '#,##0.00 €'
                ws.cell(row, 5).font = Font(bold=True)
                
                for col in range(1, 6):
                    ws.cell(row, col).fill = PatternFill(start_color='D8E4BC', fill_type='solid')
            
            # Salva il file (già in Downloads!)
            wb.save(dest_path)
            
            messagebox.showinfo(
                "✅ Completato!",
                f"Fogli creati: {len(fornitori)}\n\n📁 {filename}"
            )
            
        except PermissionError:
            messagebox.showerror("Errore", "❌ File aperto in Excel!\n\nChiudilo e riprova.")
        except Exception as e:
            messagebox.showerror("Errore", f"{str(e)}")
    
    def raggruppa_sezione(self):
        """FUNZIONE 2: Raggruppa sezione"""
        
        if not self.check_file():
            return
        
        try:
            # PRIMA: Crea copia completa del file in Downloads
            wb, dest_path, filename = self.crea_copia_completa("RAGGRUPPATO")
            
            ws_gen = None
            sez_row, sez_col = 0, 0
            
            for ws in wb.worksheets:
                if ws.title == "SEZ. RAGGRUPPATA":
                    continue
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value or "").strip().upper() == "SEZIONATURA":
                            ws_gen = ws
                            sez_row = cell.row
                            sez_col = cell.column
                            break
                    if ws_gen:
                        break
                if ws_gen:
                    break
            
            if not ws_gen:
                messagebox.showerror("Errore", "SEZIONATURA non trovata!")
                return
            
            data_start = 0
            for r in range(sez_row + 2, ws_gen.max_row + 1):
                val = ws_gen.cell(r, sez_col).value
                if val and str(val).strip():
                    val_str = str(val).strip().replace('.', '').replace(',', '')
                    if val_str.isdigit() or val_str.replace('-', '').isdigit():
                        data_start = r
                        break
            
            if not data_start:
                messagebox.showerror("Errore", "Nessun dato trovato!")
                return
            
            # Cerca marcatore END in colonna A per limitare la sezionatura
            last_row = ws_gen.max_row
            for r in range(data_start, ws_gen.max_row + 1):
                val_a = ws_gen.cell(r, 1).value
                if val_a and str(val_a).strip().upper() == "END":
                    last_row = r - 1
                    break

            raggruppati = {}
            order = []

            for r in range(data_start, last_row + 1):
                L = ws_gen.cell(r, sez_col).value
                H = ws_gen.cell(r, sez_col + 1).value
                QTA = ws_gen.cell(r, sez_col + 2).value or 0
                MAT = str(ws_gen.cell(r, sez_col + 3).value or "")
                PEZZO = str(ws_gen.cell(r, sez_col + 4).value or "")
                CODICE = str(ws_gen.cell(r, sez_col + 5).value or "")
                MOBILE = str(ws_gen.cell(r, sez_col + 6).value or "")

                if not L and not H and not MAT:
                    continue

                key = f"{L}|{H}|{MAT}"

                if key not in raggruppati:
                    raggruppati[key] = {
                        'L': str(L or ""), 'H': str(H or ""), 'QTA': 0,
                        'MAT': MAT, 'CODICE': "", 'PEZZO': "", 'MOBILE': ""
                    }
                    order.append(key)

                try:
                    qta_num = int(QTA) if QTA else 0
                except:
                    qta_num = 0

                raggruppati[key]['QTA'] += qta_num

                if CODICE:
                    if raggruppati[key]['CODICE']:
                        raggruppati[key]['CODICE'] += " + " + CODICE
                    else:
                        raggruppati[key]['CODICE'] = CODICE

                if PEZZO:
                    if raggruppati[key]['PEZZO']:
                        raggruppati[key]['PEZZO'] += " | " + PEZZO
                    else:
                        raggruppati[key]['PEZZO'] = PEZZO

                if MOBILE and MOBILE not in raggruppati[key]['MOBILE']:
                    if raggruppati[key]['MOBILE']:
                        raggruppati[key]['MOBILE'] += " | " + MOBILE
                    else:
                        raggruppati[key]['MOBILE'] = MOBILE

            if "SEZ. RAGGRUPPATA" in wb.sheetnames:
                del wb["SEZ. RAGGRUPPATA"]

            ws_dst = wb.create_sheet("SEZ. RAGGRUPPATA")
            
            ws_dst.merge_cells('A1:G1')
            titolo = ws_dst.cell(1, 1)
            titolo.value = f"SEZIONATURA RAGGRUPPATA - {ws_gen.title}"
            titolo.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
            titolo.fill = PatternFill(start_color='1F3864', fill_type='solid')
            titolo.alignment = Alignment(horizontal='center')
            ws_dst.row_dimensions[1].height = 22
            
            headers = ['L.(MM)', 'H.(MM)', "Q.TA'", 'MATERIALE', 'CODICE/I', 'PEZZO/I', 'MOBILE/I']
            for col, header in enumerate(headers, 1):
                cell = ws_dst.cell(2, col, header)
                cell.font = Font(name='Arial', bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1F4E79', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin', color='BFBFBF'),
                    right=Side(style='thin', color='BFBFBF'),
                    top=Side(style='thin', color='BFBFBF'),
                    bottom=Side(style='thin', color='BFBFBF')
                )
            
            ws_dst.column_dimensions['A'].width = 10
            ws_dst.column_dimensions['B'].width = 10
            ws_dst.column_dimensions['C'].width = 8
            ws_dst.column_dimensions['D'].width = 32
            ws_dst.column_dimensions['E'].width = 22
            ws_dst.column_dimensions['F'].width = 50
            ws_dst.column_dimensions['G'].width = 30
            
            row = 3
            for idx, key in enumerate(order):
                data = raggruppati[key]
                is_merged = ' + ' in data['CODICE']
                
                values = [data['L'], data['H'], data['QTA'], data['MAT'], 
                         data['CODICE'], data['PEZZO'], data['MOBILE']]
                
                for col, val in enumerate(values, 1):
                    cell = ws_dst.cell(row, col, val)
                    cell.font = Font(name='Arial', size=9)
                    cell.border = Border(
                        left=Side(style='thin', color='BFBFBF'),
                        right=Side(style='thin', color='BFBFBF'),
                        top=Side(style='thin', color='BFBFBF'),
                        bottom=Side(style='thin', color='BFBFBF')
                    )
                    
                    if is_merged:
                        cell.fill = PatternFill(start_color='D6E4F0', fill_type='solid')
                        cell.font = Font(name='Arial', size=9, bold=True, color='1F4E79')
                    elif idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F7FB', fill_type='solid')
                    
                    if col <= 3:
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.alignment = Alignment(horizontal='left')
                
                ws_dst.row_dimensions[row].height = 14
                row += 1
            
            row += 1
            ws_dst.merge_cells(f'A{row}:G{row}')
            nota = ws_dst.cell(row, 1)
            nota.value = "Righe in blu = pezzi accorpati → Q.TA sommata, Codici con +"
            nota.font = Font(italic=True, size=9, color='1F4E79')
            
            # Salva il file (già in Downloads!)
            wb.save(dest_path)
            
            messagebox.showinfo(
                "✅ Completato!",
                f"Righe: {len(order)}\nFoglio: {ws_gen.title}\n\n📁 {filename}"
            )
            
        except PermissionError:
            messagebox.showerror("Errore", "❌ File aperto in Excel!\n\nChiudilo e riprova.")
        except Exception as e:
            messagebox.showerror("Errore", f"{str(e)}")
    
    def elaborazione_completa(self):
        """FA TUTTE E 3 LE OPERAZIONI IN UN UNICO FILE"""
        
        if not self.check_file():
            return
        
        try:
            # PRIMA: Crea copia completa del file in Downloads
            wb, dest_path, filename = self.crea_copia_completa("COMPLETO")
            ws_main = self.trova_foglio_generale(wb)

            # === PARTE 1: FERRAMENTA ===
            fornitori = {}

            start_f, end_f = self.trova_range_sezione(ws_main, "FERRAMENTA")
            range_f = range(start_f, end_f + 1) if start_f else range(0)

            for r in range_f:
                desc = str(ws_main.cell(r, 3).value or "").strip()
                forn = str(ws_main.cell(r, 5).value or "").strip()
                
                if forn and desc:
                    if forn not in fornitori:
                        fornitori[forn] = []
                    
                    cod = str(ws_main.cell(r, 2).value or "").strip()
                    try:
                        qta = float(ws_main.cell(r, 4).value or 0)
                    except:
                        qta = 0
                    try:
                        prz = float(ws_main.cell(r, 6).value or 0)
                    except:
                        prz = 0
                    
                    if qta > 0:
                        fornitori[forn].append({'cod': cod, 'desc': desc, 'qta': qta, 'prz': prz})
            
            fogli_fornitori = 0
            for forn, articoli in fornitori.items():
                forn_clean = forn
                for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    forn_clean = forn_clean.replace(char, '_')
                forn_clean = forn_clean[:31]
                
                if forn_clean in wb.sheetnames:
                    del wb[forn_clean]
                
                ws = wb.create_sheet(title=forn_clean)
                
                headers = ['CODICE', 'DESCRIZIONE', 'Q.TA', '€/CAD', '€ TOT.']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.font = Font(bold=True, size=12)
                    c.fill = PatternFill(start_color='D8E4BC', fill_type='solid')
                    c.alignment = Alignment(horizontal='center')
                
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 65
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 14
                
                agg = {}
                for art in articoli:
                    if art['desc'] in agg:
                        agg[art['desc']]['qta'] += art['qta']
                    else:
                        agg[art['desc']] = art.copy()
                
                row = 2
                for art in agg.values():
                    ws.cell(row, 1, art['cod']).number_format = '@'
                    ws.cell(row, 2, art['desc'])
                    ws.cell(row, 3, art['qta'])
                    ws.cell(row, 4, art['prz']).number_format = '#,##0.000 €'
                    ws.cell(row, 5, f"=C{row}*D{row}").number_format = '#,##0.00 €'
                    row += 1
                
                ws.cell(row, 1, 'TOTALE').font = Font(bold=True, size=12)
                ws.cell(row, 3, f"=SUM(C2:C{row-1})").font = Font(bold=True)
                ws.cell(row, 5, f"=SUM(E2:E{row-1})").number_format = '#,##0.00 €'
                ws.cell(row, 5).font = Font(bold=True)
                
                for col in range(1, 6):
                    ws.cell(row, col).fill = PatternFill(start_color='D8E4BC', fill_type='solid')
                
                fogli_fornitori += 1
            
            # === PARTE 2: ILLUMINAZIONE ===
            fornitori_ill = {}
            ws_ill = ws_main  # già foglio GENERALE

            start_i, end_i = self.trova_range_sezione(ws_ill, "ILLUMINAZIONE")
            range_i = range(start_i, end_i + 1) if start_i else range(0)

            for r in range_i:
                desc = str(ws_ill.cell(r, 3).value or "").strip()
                forn = str(ws_ill.cell(r, 5).value or "").strip()
                
                if forn and desc:
                    if forn not in fornitori_ill:
                        fornitori_ill[forn] = []
                    
                    cod = str(ws_ill.cell(r, 2).value or "").strip()  # COLONNA 2 per codice
                    try:
                        qta = float(ws_ill.cell(r, 4).value or 0)
                    except:
                        qta = 0
                    try:
                        prz = float(ws_ill.cell(r, 6).value or 0)
                    except:
                        prz = 0
                    
                    if qta > 0:
                        fornitori_ill[forn].append({'cod': cod, 'desc': desc, 'qta': qta, 'prz': prz})
            
            fogli_illuminazione = 0
            for forn, articoli in fornitori_ill.items():
                forn_clean = forn
                for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    forn_clean = forn_clean.replace(char, '_')
                sheet_name = (forn_clean + "_ILL")[:31]
                
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                ws = wb.create_sheet(title=sheet_name)
                
                headers = ['CODICE', 'DESCRIZIONE', 'Q.TA', '€/CAD', '€ TOT.']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.font = Font(bold=True, size=12)
                    c.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
                    c.alignment = Alignment(horizontal='center')
                
                ws.column_dimensions['A'].width = 22
                ws.column_dimensions['B'].width = 65
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 14
                
                agg = {}
                for art in articoli:
                    if art['desc'] in agg:
                        agg[art['desc']]['qta'] += art['qta']
                    else:
                        agg[art['desc']] = art.copy()
                
                row = 2
                for art in agg.values():
                    ws.cell(row, 1, art['cod']).number_format = '@'
                    ws.cell(row, 2, art['desc'])
                    ws.cell(row, 3, art['qta'])
                    ws.cell(row, 4, art['prz']).number_format = '#,##0.000 €'
                    ws.cell(row, 5, art['qta'] * art['prz']).number_format = '#,##0.00 €'
                    row += 1
                
                ws.cell(row, 1, 'TOTALE').font = Font(bold=True, size=12)
                ws.cell(row, 3, f"=SUM(C2:C{row-1})").font = Font(bold=True)
                ws.cell(row, 5, f"=SUM(E2:E{row-1})").number_format = '#,##0.00 €'
                ws.cell(row, 5).font = Font(bold=True)
                
                for col in range(1, 6):
                    ws.cell(row, col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
                
                fogli_illuminazione += 1
            
            # === PARTE 3: LACCATURA (un foglio per fornitore) ===
            righe_laccatura = 0
            fogli_laccatura = 0
            finiture_lacc, order_lacc = self._trova_finiture_laccatura(ws_main)
            if order_lacc:
                righe_lacc = self._estrai_pezzi_laccatura(ws_main, finiture_lacc, order_lacc)
                if righe_lacc:
                    pezzi_agg_lacc = self._aggrega_pezzi_laccatura(ws_main, righe_lacc)
                    order_attive_lacc = self._filtra_finiture_attive(pezzi_agg_lacc, order_lacc)
                    forn_map_lacc = self._trova_fornitori_finiture(ws_main)
                    gruppi_lacc, ordine_forn_lacc = self._raggruppa_finiture_per_fornitore(
                        order_attive_lacc, forn_map_lacc)
                    for forn in ordine_forn_lacc:
                        fin_forn = gruppi_lacc[forn]
                        pezzi_forn = [p for p in pezzi_agg_lacc
                                      if any(n.upper() in str(p['mat'] or '').upper() for n in fin_forn)]
                        if not pezzi_forn:
                            continue
                        forn_clean = forn
                        for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                            forn_clean = forn_clean.replace(char, '_')
                        sheet_name = (f'LACCATURA_{forn_clean}')[:31]
                        titolo = f'{forn} — {" + ".join(fin_forn)}'
                        self._scrivi_foglio_laccatura(wb, pezzi_forn, fin_forn,
                                                      sheet_name=sheet_name, titolo=titolo)
                        fogli_laccatura += 1
                    righe_laccatura = len(pezzi_agg_lacc)

            # === PARTE 4: SEZIONE ===
            ws_gen = None
            sez_row, sez_col = 0, 0
            
            for ws in wb.worksheets:
                if ws.title == "SEZ. RAGGRUPPATA":
                    continue
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value or "").strip().upper() == "SEZIONATURA":
                            ws_gen = ws
                            sez_row = cell.row
                            sez_col = cell.column
                            break
                    if ws_gen:
                        break
                if ws_gen:
                    break
            
            righe_sezione = 0
            
            if ws_gen and sez_col > 0:
                data_start = 0
                for r in range(sez_row + 2, ws_gen.max_row + 1):
                    val = ws_gen.cell(r, sez_col).value
                    if val and str(val).strip():
                        val_str = str(val).strip().replace('.', '').replace(',', '')
                        if val_str.isdigit() or val_str.replace('-', '').isdigit():
                            data_start = r
                            break
                
                if data_start:
                    # Cerca marcatore END in colonna A per limitare la sezionatura
                    last_row = ws_gen.max_row
                    for r in range(data_start, ws_gen.max_row + 1):
                        val_a = ws_gen.cell(r, 1).value
                        if val_a and str(val_a).strip().upper() == "END":
                            last_row = r - 1
                            break
                    
                    raggruppati = {}
                    order = []
                    
                    for r in range(data_start, last_row + 1):
                        L = ws_gen.cell(r, sez_col).value
                        H = ws_gen.cell(r, sez_col + 1).value
                        QTA = ws_gen.cell(r, sez_col + 2).value or 0
                        MAT = str(ws_gen.cell(r, sez_col + 3).value or "")
                        PEZZO = str(ws_gen.cell(r, sez_col + 4).value or "")
                        CODICE = str(ws_gen.cell(r, sez_col + 5).value or "")
                        MOBILE = str(ws_gen.cell(r, sez_col + 6).value or "")
                        
                        if not L and not H and not MAT:
                            continue
                        
                        key = f"{L}|{H}|{MAT}"
                        
                        if key not in raggruppati:
                            raggruppati[key] = {
                                'L': str(L or ""), 'H': str(H or ""), 'QTA': 0,
                                'MAT': MAT, 'CODICE': "", 'PEZZO': "", 'MOBILE': ""
                            }
                            order.append(key)
                        
                        try:
                            qta_num = int(QTA) if QTA else 0
                        except:
                            qta_num = 0
                        
                        raggruppati[key]['QTA'] += qta_num
                        
                        if CODICE:
                            if raggruppati[key]['CODICE']:
                                raggruppati[key]['CODICE'] += " + " + CODICE
                            else:
                                raggruppati[key]['CODICE'] = CODICE
                        
                        if PEZZO:
                            if raggruppati[key]['PEZZO']:
                                raggruppati[key]['PEZZO'] += " | " + PEZZO
                            else:
                                raggruppati[key]['PEZZO'] = PEZZO
                        
                        if MOBILE and MOBILE not in raggruppati[key]['MOBILE']:
                            if raggruppati[key]['MOBILE']:
                                raggruppati[key]['MOBILE'] += " | " + MOBILE
                            else:
                                raggruppati[key]['MOBILE'] = MOBILE
                    
                    if "SEZ. RAGGRUPPATA" in wb.sheetnames:
                        del wb["SEZ. RAGGRUPPATA"]
                    
                    ws_dst = wb.create_sheet("SEZ. RAGGRUPPATA")
                    
                    ws_dst.merge_cells('A1:G1')
                    titolo = ws_dst.cell(1, 1)
                    titolo.value = f"SEZIONATURA RAGGRUPPATA - {ws_gen.title}"
                    titolo.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
                    titolo.fill = PatternFill(start_color='1F3864', fill_type='solid')
                    titolo.alignment = Alignment(horizontal='center')
                    ws_dst.row_dimensions[1].height = 22
                    
                    headers = ['L.(MM)', 'H.(MM)', "Q.TA'", 'MATERIALE', 'CODICE/I', 'PEZZO/I', 'MOBILE/I']
                    for col, header in enumerate(headers, 1):
                        cell = ws_dst.cell(2, col, header)
                        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='1F4E79', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                    
                    ws_dst.column_dimensions['A'].width = 10
                    ws_dst.column_dimensions['B'].width = 10
                    ws_dst.column_dimensions['C'].width = 8
                    ws_dst.column_dimensions['D'].width = 32
                    ws_dst.column_dimensions['E'].width = 22
                    ws_dst.column_dimensions['F'].width = 50
                    ws_dst.column_dimensions['G'].width = 30
                    
                    row = 3
                    for idx, key in enumerate(order):
                        data = raggruppati[key]
                        is_merged = ' + ' in data['CODICE']
                        
                        values = [data['L'], data['H'], data['QTA'], data['MAT'], 
                                 data['CODICE'], data['PEZZO'], data['MOBILE']]
                        
                        for col, val in enumerate(values, 1):
                            cell = ws_dst.cell(row, col, val)
                            cell.font = Font(name='Arial', size=9)
                            cell.border = Border(
                                left=Side(style='thin', color='BFBFBF'),
                                right=Side(style='thin', color='BFBFBF'),
                                top=Side(style='thin', color='BFBFBF'),
                                bottom=Side(style='thin', color='BFBFBF')
                            )
                            
                            if is_merged:
                                cell.fill = PatternFill(start_color='D6E4F0', fill_type='solid')
                                cell.font = Font(name='Arial', size=9, bold=True, color='1F4E79')
                            elif idx % 2 == 0:
                                cell.fill = PatternFill(start_color='F2F7FB', fill_type='solid')
                            
                            if col <= 3:
                                cell.alignment = Alignment(horizontal='center')
                            else:
                                cell.alignment = Alignment(horizontal='left')
                        
                        ws_dst.row_dimensions[row].height = 14
                        row += 1
                    
                    righe_sezione = len(order)
                    
                    row += 1
                    ws_dst.merge_cells(f'A{row}:G{row}')
                    nota = ws_dst.cell(row, 1)
                    nota.value = "Righe in blu = pezzi accorpati → Q.TA sommata, Codici con +"
                    nota.font = Font(italic=True, size=9, color='1F4E79')
            
            # === SALVA ===
            wb.save(dest_path)
            
            msg = f"⚡ ELABORAZIONE COMPLETA!\n\n"
            msg += f"📊 Fogli ferramenta: {fogli_fornitori}\n"
            msg += f"💡 Fogli illuminazione: {fogli_illuminazione}\n"
            msg += f"🎨 Fogli laccatura: {fogli_laccatura} ({righe_laccatura} pezzi)\n"
            if righe_sezione > 0:
                msg += f"📋 Righe sezione: {righe_sezione}\n"
            else:
                msg += f"⚠️ Sezione non trovata\n"
            msg += f"\n📁 {filename}"
            
            messagebox.showinfo("✅ Fatto!", msg)
            
        except PermissionError:
            messagebox.showerror("Errore", "❌ File aperto in Excel!\n\nChiudilo e riprova.")
        except Exception as e:
            messagebox.showerror("Errore", f"{str(e)}")
    
    def crea_fogli_illuminazione(self):
        """FUNZIONE ILLUMINAZIONE: Crea fogli illuminazione dalle righe 222-240"""
        
        if not self.check_file():
            return
        
        try:
            # PRIMA: Crea copia completa del file in Downloads
            wb, dest_path, filename = self.crea_copia_completa("ILLUMINAZIONE")
            
            ws_main = self.trova_foglio_generale(wb)

            start, end = self.trova_range_sezione(ws_main, "ILLUMINAZIONE")
            if not start:
                messagebox.showerror("Errore", "Sezione 'ILLUMINAZIONE' non trovata nel foglio GENERALE!")
                return

            fornitori = {}

            for r in range(start, end + 1):
                desc = str(ws_main.cell(r, 3).value or "").strip()
                forn = str(ws_main.cell(r, 5).value or "").strip()
                
                if forn and desc:
                    if forn not in fornitori:
                        fornitori[forn] = []
                    
                    cod = str(ws_main.cell(r, 2).value or "").strip()  # COLONNA 2 per codice
                    
                    try:
                        qta = float(ws_main.cell(r, 4).value or 0)
                    except:
                        qta = 0
                    
                    try:
                        prz = float(ws_main.cell(r, 6).value or 0)
                    except:
                        prz = 0
                    
                    if qta > 0:
                        fornitori[forn].append({'cod': cod, 'desc': desc, 'qta': qta, 'prz': prz})
            
            if not fornitori:
                messagebox.showwarning("Attenzione", f"Nessun fornitore trovato (righe {start}-{end})")
                return
            
            for forn, articoli in fornitori.items():
                # Nome foglio con suffisso _ILL
                forn_clean = forn
                for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    forn_clean = forn_clean.replace(char, '_')
                sheet_name = (forn_clean + "_ILL")[:31]  # Max 31 caratteri
                
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                ws = wb.create_sheet(title=sheet_name)
                
                headers = ['CODICE', 'DESCRIZIONE', 'Q.TA', '€/CAD', '€ TOT.']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.font = Font(bold=True, size=12)
                    c.fill = PatternFill(start_color='FFF2CC', fill_type='solid')  # Giallo chiaro
                    c.alignment = Alignment(horizontal='center')
                
                ws.column_dimensions['A'].width = 22
                ws.column_dimensions['B'].width = 65
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 14
                
                # Aggrega articoli per descrizione
                agg = {}
                for art in articoli:
                    if art['desc'] in agg:
                        agg[art['desc']]['qta'] += art['qta']
                    else:
                        agg[art['desc']] = art.copy()
                
                row = 2
                for art in agg.values():
                    ws.cell(row, 1, art['cod']).number_format = '@'
                    ws.cell(row, 2, art['desc'])
                    ws.cell(row, 3, art['qta'])
                    ws.cell(row, 4, art['prz']).number_format = '#,##0.000 €'
                    ws.cell(row, 5, art['qta'] * art['prz']).number_format = '#,##0.00 €'
                    row += 1
                
                # Riga totale
                ws.cell(row, 1, 'TOTALE').font = Font(bold=True, size=12)
                ws.cell(row, 3, f"=SUM(C2:C{row-1})").font = Font(bold=True)
                ws.cell(row, 5, f"=SUM(E2:E{row-1})").number_format = '#,##0.00 €'
                ws.cell(row, 5).font = Font(bold=True)
                
                for col in range(1, 6):
                    ws.cell(row, col).fill = PatternFill(start_color='FFF2CC', fill_type='solid')
            
            # Salva il file (già in Downloads!)
            wb.save(dest_path)
            
            messagebox.showinfo(
                "✅ Completato!",
                f"Fogli illuminazione creati: {len(fornitori)}\n\n📁 {filename}"
            )
            
        except PermissionError:
            messagebox.showerror("Errore", "❌ File aperto in Excel!\n\nChiudilo e riprova.")
        except Exception as e:
            messagebox.showerror("Errore", f"{str(e)}")

    # === Keyword per riconoscere le finiture nella riga 3 del GENERALE ===
    LACCATURA_KEYWORDS = ['LACCATO', 'EFFETTO ROCCIA', 'TINTO', 'VERNICIAT']
    # Keyword per identificare le colonne BORDO (per calcolo CART.RA AL MQ)
    BORDO_KEYWORD = 'BORDO'

    def _trova_bordi(self, ws_gen):
        """Scorre la riga 3 del foglio GENERALE e trova tutte le colonne BORDO.
        Estrae l'altezza H del bordo dal nome (es. 'BORDO ABS H.23 ...' → H=23).
        Ritorna lista di tuple (col_idx, H_mm) — solo colonne con H riconoscibile.
        Le colonne BORDO contengono ML totali; per MQ → ML * H/1000."""
        bordi = []
        for c in range(10, ws_gen.max_column + 1):
            val = ws_gen.cell(3, c).value
            if not val:
                continue
            s = str(val).strip()
            su = s.upper()
            if not su.startswith(self.BORDO_KEYWORD):
                continue
            # Estrai H (es. H.23, H 28, H.28, H28)
            m = re.search(r'\bH[\s\.]*(\d+)\b', su)
            if not m:
                continue
            h_mm = int(m.group(1))
            bordi.append((c, h_mm))
        return bordi

    def _trova_finiture_laccatura(self, ws_gen):
        """Scorre la riga 3 del foglio GENERALE e trova tutte le colonne che
        contengono le keyword finitura (LACCATO / EFFETTO ROCCIA / TINTO / VERNICIATO).
        Raggruppa per nome base (senza suffisso MQ/ML) e ritorna:
          - finiture: dict {nome_base: {'mq_orig': col, 'ml_orig': col}}
          - order: lista di nomi base nell'ordine di apparizione."""
        finiture = {}
        order = []
        for c in range(10, ws_gen.max_column + 1):
            val = ws_gen.cell(3, c).value
            if not val:
                continue
            s = str(val).strip()
            su = s.upper()
            if not any(kw in su for kw in self.LACCATURA_KEYWORDS):
                continue

            # Estrai nome base e tipo (MQ/ML)
            if re.search(r'\s+MQ\s*$', su):
                nome = re.sub(r'\s+MQ\s*$', '', s, flags=re.IGNORECASE).strip()
                tipo = 'MQ'
            elif re.search(r'\s+ML\s*$', su):
                nome = re.sub(r'\s+ML\s*$', '', s, flags=re.IGNORECASE).strip()
                tipo = 'ML'
            else:
                nome = s
                tipo = 'MQ'  # nessun suffisso → tratto come MQ
            nome = re.sub(r'\s+', ' ', nome).strip()

            if nome not in finiture:
                finiture[nome] = {'mq_orig': None, 'ml_orig': None}
                order.append(nome)
            if tipo == 'MQ':
                finiture[nome]['mq_orig'] = c
            else:
                finiture[nome]['ml_orig'] = c
        return finiture, order

    def _estrai_pezzi_laccatura(self, ws_gen, finiture=None, order=None):
        """Ritorna le righe del foglio GENERALE che hanno almeno un valore numerico
        > 0 in una delle colonne finitura individuate da _trova_finiture_laccatura.
        Se finiture/order non sono passati, vengono calcolati."""
        if finiture is None or order is None:
            finiture, order = self._trova_finiture_laccatura(ws_gen)

        fin_cols = []
        for n in order:
            if finiture[n]['mq_orig']:
                fin_cols.append(finiture[n]['mq_orig'])
            if finiture[n]['ml_orig']:
                fin_cols.append(finiture[n]['ml_orig'])

        righe = []
        for r in range(4, ws_gen.max_row + 1):
            for c in fin_cols:
                v = ws_gen.cell(r, c).value
                if isinstance(v, (int, float)) and v != 0:
                    righe.append(r)
                    break
        return righe

    def _aggrega_pezzi_laccatura(self, ws_gen, righe, bordi=None):
        """Raggruppa per (L, H, SP, MATERIALE, TIPOLOGIA, LATI):
        - somma le Q.TA
        - concatena CODICE con ' + ', MOBILE/PEZZO con ' | '
        - somma 'mq_bordi' = sum(ML_bordo_i * H_i/1000) per ogni colonna BORDO
          (serve per la colonna CART.RA AL MQ)
        Ritorna lista di dict nell'ordine di prima apparizione."""
        if bordi is None:
            bordi = self._trova_bordi(ws_gen)
        agg = {}
        order = []
        for r in righe:
            cod = str(ws_gen.cell(r, 1).value or '').strip()
            mob = str(ws_gen.cell(r, 2).value or '').strip()
            mat = ws_gen.cell(r, 3).value
            try:
                qta = float(ws_gen.cell(r, 4).value or 0)
            except (ValueError, TypeError):
                qta = 0
            pez = str(ws_gen.cell(r, 5).value or '').strip()
            h = ws_gen.cell(r, 6).value
            l = ws_gen.cell(r, 7).value
            sp = ws_gen.cell(r, 8).value
            tip = ws_gen.cell(r, 9).value
            lati = ws_gen.cell(r, 37).value

            # MQ dei bordi del pezzo = somma su ogni colonna bordo di (ML * H/1000)
            mq_bordi_row = 0.0
            for (col_b, h_mm) in bordi:
                vb = ws_gen.cell(r, col_b).value
                if isinstance(vb, (int, float)) and vb:
                    mq_bordi_row += float(vb) * h_mm / 1000.0

            key = (str(l), str(h), str(sp), str(mat), str(tip), str(lati))

            if key not in agg:
                agg[key] = {
                    'cod': cod, 'mob': mob, 'mat': mat, 'qta': qta, 'pez': pez,
                    'h': h, 'l': l, 'sp': sp, 'tip': tip, 'lati': lati,
                    'count': 1,
                    'mq_bordi': mq_bordi_row,
                }
                order.append(key)
            else:
                agg[key]['qta'] += qta
                agg[key]['count'] += 1
                agg[key]['mq_bordi'] += mq_bordi_row
                if cod and cod not in agg[key]['cod'].split(' + '):
                    agg[key]['cod'] = (agg[key]['cod'] + ' + ' + cod) if agg[key]['cod'] else cod
                if mob and mob not in agg[key]['mob'].split(' | '):
                    agg[key]['mob'] = (agg[key]['mob'] + ' | ' + mob) if agg[key]['mob'] else mob
                if pez and pez not in agg[key]['pez'].split(' | '):
                    agg[key]['pez'] = (agg[key]['pez'] + ' | ' + pez) if agg[key]['pez'] else pez

        return [agg[k] for k in order]

    def _filtra_finiture_attive(self, pezzi_agg, order):
        """Ritorna solo i nomi di finitura il cui testo è contenuto nel
        MATERIALE di almeno un pezzo aggregato (quelle che daranno almeno un valore > 0)."""
        attive = []
        for nome in order:
            nome_u = nome.upper()
            for p in pezzi_agg:
                if nome_u in str(p['mat'] or '').upper():
                    attive.append(nome)
                    break
        return attive

    def _trova_fornitori_finiture(self, ws_gen):
        """Legge la sezione VERNICIATURA del foglio GENERALE e ritorna
        un dict {NOME_FINITURA_UPPER: FORNITORE}. Estrae il nome base
        rimuovendo i suffissi MQ/ML in coda."""
        start, end = self.trova_range_sezione(ws_gen, 'VERNICIATURA')
        if not start:
            return {}

        mapping = {}
        for r in range(start, end + 1):
            desc = ws_gen.cell(r, 3).value
            forn = ws_gen.cell(r, 5).value
            if not desc or not forn:
                continue
            s = str(desc).strip()
            if not s:
                continue
            su = s.upper()
            if re.search(r'\s+MQ\s*$', su):
                nome = re.sub(r'\s+MQ\s*$', '', s, flags=re.IGNORECASE).strip()
            elif re.search(r'\s+ML\s*$', su):
                nome = re.sub(r'\s+ML\s*$', '', s, flags=re.IGNORECASE).strip()
            else:
                nome = s
            nome = re.sub(r'\s+', ' ', nome).strip()
            mapping[nome.upper()] = str(forn).strip()
        return mapping

    def _raggruppa_finiture_per_fornitore(self, order_attive, mapping):
        """Raggruppa le finiture per fornitore. Ritorna dict {fornitore: [finiture]}
        preservando l'ordine di apparizione."""
        gruppi = {}
        ordine = []
        for nome in order_attive:
            forn = mapping.get(nome.upper(), 'ALTRO')
            if forn not in gruppi:
                gruppi[forn] = []
                ordine.append(forn)
            gruppi[forn].append(nome)
        return gruppi, ordine

    def _scrivi_foglio_laccatura(self, wb, pezzi_agg, order_attive,
                                  sheet_name="LACCATURA", titolo=None):
        """Crea un foglio di laccatura con layout dinamico:
        - 11 colonne base (CODICE..SP + TIPOLOGIA + LATI + MQ TOTALI)
        - 2 colonne per ogni finitura ATTIVA (MQ + ML) con FORMULE auto-contenute.
        Le righe accorpate (più pezzi con stesse misure/tipologia/lati) sono evidenziate.
        `sheet_name`: nome del foglio (es. 'LACCATURA_FAS').
        `titolo`: titolo da mettere in riga 1 (default generico)."""
        if titolo is None:
            titolo = 'PEZZI CON LACCATURA / TINTO / VERNICIATURA / EFFETTO ROCCIA'
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        base_headers = ['CODICE', 'MOBILE', 'MATERIALE', 'Q.TA',
                        'H.(CM)', 'L.(CM)', 'SP.(CM)',
                        'U.M.', 'LATI DA\nLACCARE', 'MQ TOTALI']
        n_base = len(base_headers)

        # Mappa colonne finiture: (col_idx, tipo, nome)
        fin_headers = []
        fin_cols_map = []
        col_idx = n_base + 1
        for nome in order_attive:
            fin_headers.append(f'{nome}\nMQ')
            fin_cols_map.append((col_idx, 'MQ', nome))
            col_idx += 1
            fin_headers.append(f'{nome}\nML')
            fin_cols_map.append((col_idx, 'ML', nome))
            col_idx += 1

        # Colonne CART.RA in coda: sommano TUTTE le finiture attive (laccato/tinto/
        # verniciato/effetto roccia/alluminio). MQ include anche i bordi:
        # ML_bordo * H_bordo/1000.
        cart_mq_cols = [ci for (ci, tp, nm) in fin_cols_map if tp == 'MQ']
        cart_ml_cols = [ci for (ci, tp, nm) in fin_cols_map if tp == 'ML']
        cart_ml_col = n_base + len(fin_headers) + 1
        cart_mq_col = n_base + len(fin_headers) + 2
        cart_headers = ['CART.RA\nAL ML', 'CART.RA\nAL MQ']
        total_cols = n_base + len(fin_headers) + len(cart_headers)

        # Titolo
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        t = ws.cell(1, 1, titolo)
        t.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
        t.fill = PatternFill(start_color='1F3864', fill_type='solid')
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        all_headers = base_headers + fin_headers + cart_headers
        n_fin = len(fin_headers)
        for col, h in enumerate(all_headers, 1):
            c = ws.cell(2, col, h)
            c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
            if col <= n_base:
                c.fill = PatternFill(start_color='1F4E79', fill_type='solid')
            elif col <= n_base + n_fin:
                idx_pair = (col - n_base - 1) // 2
                color = '7030A0' if idx_pair % 2 == 0 else '9C4FBA'
                c.fill = PatternFill(start_color=color, fill_type='solid')
            else:
                # Colonne CART.RA in verde scuro
                c.fill = PatternFill(start_color='548235', fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border
        ws.row_dimensions[2].height = 42

        widths_base = [12, 22, 38, 6, 9, 9, 8, 12, 10, 10]
        for i, w in enumerate(widths_base, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for i in range(n_base + 1, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Righe dati
        row = 3
        for p in pezzi_agg:
            is_merged = p.get('count', 1) > 1

            base_vals = [
                p['cod'],            # A - CODICE
                p['mob'],            # B - MOBILE
                p['mat'],            # C - MATERIALE
                p['qta'],            # D - QTA
                (p['h']/10 if isinstance(p['h'], (int, float)) else p['h']),    # E - H (cm)
                (p['l']/10 if isinstance(p['l'], (int, float)) else p['l']),    # F - L (cm)
                (p['sp']/10 if isinstance(p['sp'], (int, float)) else p['sp']), # G - SP (cm)
                p['tip'],            # H - U.M. (MQ/ML)
                p['lati'],           # I - LATI DA LACCARE
                None,                # J - MQ TOTALI: formula sotto
            ]
            for col, v in enumerate(base_vals, 1):
                c = ws.cell(row, col, v)
                c.font = Font(name='Arial', size=9)
                c.border = border
                if col in (1, 2, 3):  # CODICE/MOBILE/MAT possono essere concatenati → wrap
                    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    c.alignment = Alignment(horizontal='center', vertical='center')

            # MQ TOTALI in colonna J: formula = F/100 * E/100 * D (L * H * QTA), dimensioni in cm
            c = ws.cell(row, 10, f'=F{row}/100*E{row}/100*D{row}')
            c.font = Font(name='Arial', size=9)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.number_format = '#,##0.0000'

            # Formule MQ e ML per ogni finitura attiva (dimensioni in cm)
            # $H=U.M.(MQ/ML), $C=MATERIALE, $E=H, $F=L, $G=SP, $I=LATI, $D=QTA
            for (col_i, tipo, nome) in fin_cols_map:
                nome_safe = nome.replace('"', '""')
                if tipo == 'MQ':
                    # MQ = (faccia*lati + coste (E+F)*SP*2) / 10000 (cm²->m²) * Q.ta ; lati vuoto = 1
                    formula = (f'=IF(AND($H{row}="MQ",ISNUMBER(SEARCH("{nome_safe}",$C{row}))),'
                               f'($E{row}*$F{row}*IF($I{row}="",1,$I{row})+($E{row}+$F{row})*$G{row}*2)/10000*$D{row},"")')
                else:
                    # ML = MAX(H in metri ; minimo 1 m) * Q.ta
                    formula = (f'=IF(AND($H{row}="ML",ISNUMBER(SEARCH("{nome_safe}",$C{row}))),'
                               f'MAX($E{row}/100,1)*$D{row},"")')
                c = ws.cell(row, col_i, formula)
                c.font = Font(name='Arial', size=9)
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.number_format = '#,##0.0000'

            # === CART.RA AL ML = somma di TUTTE le celle ML delle finiture della riga ===
            # SUM() ignora automaticamente le stringhe vuote "" che le formule
            # finiture restituiscono quando la condizione è falsa.
            if cart_ml_cols:
                refs_ml = ','.join(f'{get_column_letter(ci)}{row}' for ci in cart_ml_cols)
                formula_ml = f'=SUM({refs_ml})'
            else:
                formula_ml = '=0'
            c = ws.cell(row, cart_ml_col, formula_ml)
            c.font = Font(name='Arial', size=9, bold=True, color='375623')
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.number_format = '#,##0.0000'

            # === CART.RA AL MQ = somma di TUTTE le MQ finiture + MQ dei bordi ===
            mq_bordi_val = float(p.get('mq_bordi') or 0)
            if cart_mq_cols:
                refs_mq = ','.join(f'{get_column_letter(ci)}{row}' for ci in cart_mq_cols)
                formula_mq = f'=SUM({refs_mq})+{mq_bordi_val}'
            else:
                formula_mq = f'={mq_bordi_val}'
            c = ws.cell(row, cart_mq_col, formula_mq)
            c.font = Font(name='Arial', size=9, bold=True, color='375623')
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.number_format = '#,##0.0000'

            # Sfondo: righe accorpate in azzurro chiaro/blu, altrimenti alternate
            if is_merged:
                fill = PatternFill(start_color='D6E4F0', fill_type='solid')
                bold_font = Font(name='Arial', size=9, bold=True, color='1F4E79')
                for col in range(1, total_cols + 1):
                    cc = ws.cell(row, col)
                    cc.fill = fill
                    # Mantieni le formule ma metti il testo in grassetto blu per le base
                    if col <= n_base:
                        cc.font = bold_font
            elif row % 2 == 1:
                fill = PatternFill(start_color='F2F7FB', fill_type='solid')
                for col in range(1, total_cols + 1):
                    ws.cell(row, col).fill = fill

            row += 1

        # Riga TOTALE
        ws.cell(row, 1, 'TOTALE').font = Font(name='Arial', bold=True, size=11)
        for col in range(1, n_base + 1):
            c = ws.cell(row, col)
            c.fill = PatternFill(start_color='D8E4BC', fill_type='solid')
            c.border = border
        ws.cell(row, 1).alignment = Alignment(horizontal='center')

        # SUM colonna J (MQ TOTALI) + colonne finitura
        c = ws.cell(row, 10, f'=SUM(J3:J{row-1})')
        c.font = Font(name='Arial', bold=True, size=10)
        c.fill = PatternFill(start_color='D8E4BC', fill_type='solid')
        c.alignment = Alignment(horizontal='center')
        c.number_format = '#,##0.0000'
        c.border = border

        for col_i in range(n_base + 1, total_cols + 1):
            letter = get_column_letter(col_i)
            c = ws.cell(row, col_i, f'=SUM({letter}3:{letter}{row-1})')
            c.font = Font(name='Arial', bold=True, size=10)
            c.fill = PatternFill(start_color='D8E4BC', fill_type='solid')
            c.alignment = Alignment(horizontal='center')
            c.number_format = '#,##0.0000'
            c.border = border

        # Nota in fondo
        nota_row = row + 2
        ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=total_cols)
        nota = ws.cell(nota_row, 1, 'Righe in azzurro = pezzi accorpati (stesse misure/tipologia/lati) → CODICE concatenato con "+", Q.TA sommata')
        nota.font = Font(name='Arial', italic=True, size=9, color='1F4E79')

        ws.freeze_panes = 'A3'

    def crea_foglio_laccatura(self):
        """FUNZIONE LACCATURA: Estrae pezzi con valore in AK-AO del foglio GENERALE"""

        if not self.check_file():
            return

        try:
            wb, dest_path, filename = self.crea_copia_completa("LACCATURA")
            ws_gen = self.trova_foglio_generale(wb)

            finiture, order = self._trova_finiture_laccatura(ws_gen)
            if not order:
                messagebox.showerror("Errore",
                    "Nessuna finitura LACCATO/EFFETTO ROCCIA/TINTO/VERNICIATO\n"
                    "trovata nella riga 3 del foglio GENERALE!")
                return

            righe = self._estrai_pezzi_laccatura(ws_gen, finiture, order)
            if not righe:
                messagebox.showwarning("Attenzione",
                    "Nessun pezzo trovato con valore nelle colonne finitura")
                return

            # Aggrega per misure/tipologia/lati identici
            pezzi_agg = self._aggrega_pezzi_laccatura(ws_gen, righe)
            # Tieni solo le finiture che hanno almeno un pezzo corrispondente
            order_attive = self._filtra_finiture_attive(pezzi_agg, order)

            # Trova mapping finitura -> fornitore (dalla sezione VERNICIATURA)
            forn_mapping = self._trova_fornitori_finiture(ws_gen)
            gruppi, ordine_forn = self._raggruppa_finiture_per_fornitore(order_attive, forn_mapping)

            # Crea un foglio LACCATURA_<FORNITORE> per ogni fornitore
            n_fogli = 0
            report_lines = []
            for forn in ordine_forn:
                finiture_forn = gruppi[forn]
                # Pezzi che usano almeno una delle finiture di questo fornitore
                pezzi_forn = [p for p in pezzi_agg
                              if any(n.upper() in str(p['mat'] or '').upper() for n in finiture_forn)]
                if not pezzi_forn:
                    continue

                # Sanifica nome foglio
                forn_clean = forn
                for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                    forn_clean = forn_clean.replace(char, '_')
                sheet_name = (f'LACCATURA_{forn_clean}')[:31]
                titolo = f'{forn} — {" + ".join(finiture_forn)}'

                self._scrivi_foglio_laccatura(wb, pezzi_forn, finiture_forn,
                                              sheet_name=sheet_name, titolo=titolo)
                n_fogli += 1
                report_lines.append(f'• {sheet_name}: {len(pezzi_forn)} pezzi')

            wb.save(dest_path)

            n_accorpati = sum(1 for p in pezzi_agg if p.get('count', 1) > 1)
            escluse = [n for n in order if n not in order_attive]
            msg = (f"Pezzi laccatura: {len(righe)} → {len(pezzi_agg)} (accorpati: {n_accorpati})\n"
                   f"Fogli creati: {n_fogli}\n")
            msg += '\n'.join(report_lines) + '\n'
            if escluse:
                msg += f"\nFiniture escluse (0 pezzi): {', '.join(escluse)}"
            msg += f"\n\n📁 {filename}"
            messagebox.showinfo("✅ Completato!", msg)

        except PermissionError:
            messagebox.showerror("Errore", "❌ File aperto in Excel!\n\nChiudilo e riprova.")
        except Exception as e:
            messagebox.showerror("Errore", f"{str(e)}")

    def run(self):
        self.window.mainloop()

if __name__ == "__main__":
    app = AutomazioniFornitori()
    app.run()
