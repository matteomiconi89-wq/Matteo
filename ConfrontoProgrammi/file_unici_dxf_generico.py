#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
FILE UNICI DXF - VERSIONE GENERICA   (app con un pulsante)
====================================================================
Versione SENZA vincoli di progetto del "Confronto Programmi DXF":
  - funziona su QUALSIASI cartella: legge tutti i .dxf anche nelle
    sottocartelle a qualsiasi livello (ricorsivo); vanno bene anche
    file sciolti direttamente nella cartella scelta
  - ogni sottocartella che contiene .dxf diventa un GRUPPO (etichetta
    libera, modificabile in finestra: non serve un codice camera)
  - due file sono lo stesso programma se hanno stesso NOME PEZZO e
    stessa GEOMETRIA reale (vertici + fori dall'ACIS, non i byte)
  - riepilogo Excel SEMPLICE: niente regole MATERIALE/CNC/BORDATURA,
    niente moltiplicatori quantita' - vale per qualsiasi lavoro
  - i .dxf 2D (senza solido) vengono confrontati lo stesso usando le
    entita' del disegno (linee, cerchi, polilinee)

NOMI FILE IN USCITA:
  - pezzo identico in tutti i gruppi dove compare -> Nome.dxf
  - pezzo con geometrie diverse tra gruppi        -> Nome_GRUPPO.dxf
  I doppioni con stesso nome+geometria si fondono (qty = n. posizioni).

USO:
  - doppio clic (o l'.exe) -> finestra con il pulsante
  - da riga di comando:  py file_unici_dxf_generico.py --run "C:\\...\\cartella"
Richiede (solo per .py): ezdxf, openpyxl  ->  py -m pip install ezdxf openpyxl
"""
import sys, os, re, json, shutil, hashlib, threading, queue, time
from pathlib import Path
from collections import defaultdict

NDEC = 2  # arrotondamento coordinate (mm) per la firma geometrica
OUT_DIRNAME = "PROGRAMMI_UNICI"


# --- file su Dropbox/OneDrive: la sincronizzazione blocca i file appena
# --- creati/cancellati -> ogni operazione va ritentata con attesa crescente
def _copy_retry(src, dst, tries=12):
    for i in range(tries):
        try:
            return shutil.copy2(src, dst)
        except PermissionError:
            time.sleep(0.5 * (i + 1))
    # ultimo tentativo: copia manuale dei byte (aggira il lock di CopyFile2)
    Path(dst).write_bytes(Path(src).read_bytes())


def _unlink_retry(p, tries=8):
    for i in range(tries):
        try:
            Path(p).unlink()
            return True
        except FileNotFoundError:
            return True
        except PermissionError:
            time.sleep(0.5 * (i + 1))
    return False


# ============================ GEOMETRIA ====================================
def model_solid(doc):
    for e in doc.modelspace():
        if e.dxftype() == "3DSOLID":
            return e
    return None


# Pulizia del nome pezzo letto dal blocco (Xdata ESPSOL_NOME): si tolgono il
# suffisso "_bis" e l'eventuale spazzatura di commessa davanti (J-03.02_...).
# Se il nome non ha quei pattern resta com'e': nessun vincolo di formato.
_BIS_SUFFIX = re.compile(r"[\s_.\-]+bis\s*$", re.IGNORECASE)
_JOB_PREFIX = re.compile(r"^\s*[Jj][-_.\s]?\d")          # inizia col codice commessa
_PIECE_START = re.compile(r"\d{1,2}[\s_.\-]+[A-Za-z]")   # numero pezzo + sep + lettera


def clean_name(name):
    if not name:
        return name
    name = _BIS_SUFFIX.sub("", name).strip()
    if _JOB_PREFIX.match(name):                 # c'e' spazzatura davanti al nome
        starts = list(_PIECE_START.finditer(name))
        if starts:                              # taglia fino all'ultimo numero pezzo
            name = name[starts[-1].start():]
        name = re.sub(r"^\s*(\d{1,2})[\s_.\-]+", r"\1_", name)
    return name.strip()


def materiale_di(doc, e):
    """Nome del materiale ACAD del solido, se assegnato davvero (non
    ByLayer/Global): finisce nella colonna MATERIALE del riepilogo.
    Nei disegni senza materiali torna '' e la colonna si compila a mano
    (la SEZIONATURA la eredita comunque via formula)."""
    try:
        h = e.dxf.get("material_handle", None)
        if h:
            m = doc.entitydb.get(h)
            nome = (getattr(m.dxf, "name", "") or "").strip()
            if nome and nome.lower() not in ("global", "bylayer", "byblock"):
                return nome
    except Exception:
        pass
    return ""


def piece_name(e, fallback):
    name = fallback
    trovato = False
    try:
        for c, v in e.get_xdata("ESPSOL_NOME"):
            if c == 1000 and v:
                name = v
                trovato = True
                break
    except Exception:
        pass
    # VARIANTE senza blocchi/Xdata: il NOME sta sul LAYER "NOME -- MATERIALE"
    # (prima dell'ultimo ' -- '); meglio del fallback numerico dell'export
    if not trovato:
        lay = (getattr(e.dxf, "layer", "") or "").strip()
        up = lay.upper()
        if (" -- " in lay
                and not up.startswith(("FERRAMENTA", "FORO_", "SCASSO_",
                                       "NOMI_SOLIDI"))):
            name = lay.rsplit(" -- ", 1)[0].strip()
    return clean_name(name)


def materiale_layer(e, noti=None):
    """Materiale dal NOME DEL LAYER del solido (convenzione esecutivi:
    layer 'F_<materiale>', es. F_LISTELLARE_rovere_SP_25 MM); accetta
    anche un layer che combacia con una voce della lista materiali
    centrale (es. 'LSB bianco sp. 18 mm'). I layer di servizio no."""
    lay = (getattr(e.dxf, "layer", "") or "").strip()
    if not lay or lay in ("0", "Defpoints"):
        return ""
    up = lay.upper()
    if up.startswith(("FERRAMENTA", "FORO_", "SCASSO_", "NOMI_SOLIDI")):
        return ""
    # VARIANTE senza blocchi: il layer e' "NOME_PEZZO -- MATERIALE" ->
    # il materiale e' dopo l'ultimo ' -- '
    if " -- " in lay:
        return lay.split(" -- ")[-1].strip()
    if up.startswith("F_"):
        return lay
    if noti and " ".join(lay.split()).upper() in noti:
        return lay
    return ""


def materiale_xdata(e):
    """Materiale scritto da ESPLODINOMI (Xdata ESPSOL_MAT) quando il
    blocco STP si chiama "NOME -- MATERIALE": la fonte migliore, viaggia
    col 3D."""
    try:
        for c, v in e.get_xdata("ESPSOL_MAT"):
            if c == 1000 and v:
                return v.strip()
    except Exception:
        pass
    return ""


def _sig_3d(pts, ell, ndec=NDEC):
    """Normalizza punti + ellissi (fori) e produce la firma: uguale sia che
    i dati arrivino dal SAB (binario) sia dal SAT (testo)."""
    if not pts:
        return None
    mn = [min(p[i] for p in pts) for i in range(3)]
    mx = [max(p[i] for p in pts) for i in range(3)]
    dims = tuple(sorted((round(mx[i] - mn[i], 1) for i in range(3)), reverse=True))
    npts = sorted(tuple(round(p[i] - mn[i], ndec) for i in range(3)) for p in pts)
    nell = sorted((round(r, 2), tuple(round(c[i] - mn[i], ndec) for i in range(3)))
                  for c, r in ell)
    h = hashlib.sha1(repr((npts, nell)).encode()).hexdigest()[:12]
    return {"dims": list(dims), "nholes": len(ell), "hash": h, "tipo": "3D"}


# --- transform del SAT: AutoCAD scrive il corpo ACIS in coordinate LOCALI
# --- + un record "transform" da applicare (vedi confronta_file_unici_dxf)
def _tr_leggi(numeri):
    if len(numeri) < 12:
        return None
    cx, cy, cz, t = numeri[0:3], numeri[3:6], numeri[6:9], tuple(numeri[9:12])
    ident = (abs(cx[0]-1) < 1e-12 and abs(cy[1]-1) < 1e-12
             and abs(cz[2]-1) < 1e-12
             and all(abs(v) < 1e-12 for v in (cx[1], cx[2], cy[0], cy[2],
                                              cz[0], cz[1]))
             and all(abs(v) < 1e-12 for v in t))
    return None if ident else ((cx, cy, cz), t)


def _tr_punto(tr, p):
    (cx, cy, cz), t = tr
    return (p[0]*cx[0] + p[1]*cy[0] + p[2]*cz[0] + t[0],
            p[0]*cx[1] + p[1]*cy[1] + p[2]*cz[1] + t[1],
            p[0]*cx[2] + p[1]*cy[2] + p[2]*cz[2] + t[2])


def geom_sig(sab, ndec=NDEC):
    """Firma geometrica dall'ACIS binario (SAB, DXF 2013+)."""
    from ezdxf.acis.sab import Decoder
    from ezdxf.acis.const import Tags
    dec = Decoder(sab); dec.read_header()
    pts, ell = [], []
    tr = None
    for rec in dec.read_records():
        if not rec:
            continue
        nm = rec[0].value if isinstance(rec[0].value, str) else ""
        locs = [t.value for t in rec if t.tag == Tags.LOCATION_VEC]
        dirs = [t.value for t in rec if t.tag == Tags.DIRECTION_VEC]
        if nm.startswith("transform") and tr is None:
            numeri = []
            for tk in rec:
                if isinstance(tk.value, (int, float)):
                    numeri.append(float(tk.value))
                elif isinstance(tk.value, (tuple, list)):
                    numeri.extend(float(v) for v in tk.value)
            tr = _tr_leggi(numeri)
        elif nm.startswith("point"):
            pts.extend(locs)
        elif "ellipse" in nm and locs:
            maj = dirs[-1] if dirs else (0, 0, 0)
            ell.append((locs[0], (maj[0]**2 + maj[1]**2 + maj[2]**2) ** 0.5))
    return _sig_dims_vere(pts, ell, tr, ndec)


def geom_sig_sat(sat_lines, ndec=NDEC):
    """Firma geometrica dall'ACIS testuale (SAT, DXF 2000-2010): il parser
    toglie id e puntatori, restano i dati puri.
      point         -> x y z
      ellipse-curve -> centro(3) normale(3) asse_maggiore(3) ...  raggio=|asse|"""
    from ezdxf.acis.sat import parse_sat
    pts, ell = [], []
    tr = None
    for rec in parse_sat(sat_lines).entities:
        try:
            if rec.name == "transform" and tr is None:
                numeri = []
                for v in rec.data:
                    try:
                        numeri.append(float(v))
                    except (TypeError, ValueError):
                        pass
                tr = _tr_leggi(numeri)
            elif rec.name.startswith("point"):
                pts.append(tuple(float(v) for v in rec.data[:3]))
            elif "ellipse" in rec.name:
                d = rec.data
                c = tuple(float(v) for v in d[0:3])
                maj = [float(v) for v in d[6:9]]
                ell.append((c, (maj[0]**2 + maj[1]**2 + maj[2]**2) ** 0.5))
        except (ValueError, IndexError):
            continue
    return _sig_dims_vere(pts, ell, tr, ndec)


def _sig_dims_vere(pts, ell, tr, ndec):
    """FIRMA (hash) dalle coordinate LOCALI (raggruppamento stabile,
    indipendente da come l'export ha girato il pezzo); MISURE dalla
    geometria con la TRANSFORM applicata (quote vere del pezzo)."""
    sig = _sig_3d(pts, ell, ndec)
    if sig and tr and pts:
        pw = [_tr_punto(tr, p) for p in pts]
        mn = [min(p[i] for p in pw) for i in range(3)]
        mx = [max(p[i] for p in pw) for i in range(3)]
        sig["dims"] = sorted((round(mx[i] - mn[i], 1) for i in range(3)),
                             reverse=True)
    return sig


def geom_sig_2d(doc, ndec=NDEC):
    """Firma per .dxf 2D (senza solido): punti delle entita' + cerchi (fori),
    normalizzati allo spigolo min. Cosi' anche i programmi 2D si confrontano."""
    pts, circ, arcs = [], [], []
    for e in doc.modelspace():
        t = e.dxftype()
        try:
            if t == "LINE":
                pts.append(tuple(e.dxf.start)); pts.append(tuple(e.dxf.end))
            elif t == "LWPOLYLINE":
                for x, y, *_ in e.get_points():
                    pts.append((x, y, 0.0))
            elif t == "POLYLINE":
                for v in e.vertices:
                    pts.append(tuple(v.dxf.location))
            elif t == "CIRCLE":
                c = tuple(e.dxf.center)
                circ.append((c, float(e.dxf.radius))); pts.append(c)
            elif t == "ARC":
                c = tuple(e.dxf.center)
                pts.append(c)
                arcs.append((c, float(e.dxf.radius),
                             round(e.dxf.start_angle, 1), round(e.dxf.end_angle, 1)))
            elif t == "ELLIPSE":
                pts.append(tuple(e.dxf.center))
            elif t == "POINT":
                pts.append(tuple(e.dxf.location))
        except Exception:
            continue
    if not pts:
        return None
    pts = [(p[0], p[1], p[2] if len(p) > 2 else 0.0) for p in pts]
    mn = [min(p[i] for p in pts) for i in range(3)]
    mx = [max(p[i] for p in pts) for i in range(3)]
    dims = tuple(sorted((round(mx[i] - mn[i], 1) for i in range(3)), reverse=True))
    norm = lambda c: tuple(round((c[i] if i < len(c) else 0.0) - mn[i], ndec)
                           for i in range(3))
    npts = sorted(tuple(round(p[i] - mn[i], ndec) for i in range(3)) for p in pts)
    ncirc = sorted((round(r, 2), norm(c)) for c, r in circ)
    narcs = sorted((round(r, 2), a1, a2, norm(c)) for c, r, a1, a2 in arcs)
    h = hashlib.sha1(repr((npts, ncirc, narcs)).encode()).hexdigest()[:12]
    return {"dims": list(dims), "nholes": len(ncirc), "hash": h, "tipo": "2D"}


# ============================ RILEVA GRUPPI =================================
def detect_groups(base):
    """Trova TUTTE le cartelle (base compresa, ricorsivo) che contengono
    direttamente dei .dxf. Ogni cartella e' un gruppo con etichetta libera."""
    base = Path(base)
    groups = []
    dirs = [base] + sorted(p for p in base.rglob("*") if p.is_dir())
    labels_used = {}
    for d in dirs:
        rel = d.relative_to(base)
        if OUT_DIRNAME in rel.parts:          # mai leggere l'output precedente
            continue
        dxfs = [f for f in d.glob("*.dxf") if f.is_file()]
        if not dxfs:
            continue
        label = d.name if d != base else "(principale)"
        # etichette doppie (sottocartelle omonime): aggiungi il pezzo di percorso
        if label in labels_used:
            label = "_".join(rel.parts) or label
            k = 2
            while label in labels_used:
                label = f"{label}_{k}"; k += 1
        labels_used[label] = True
        groups.append({"label": label, "rel": str(rel) if str(rel) != "." else "",
                       "path": str(d), "n": len(dxfs)})
    return groups


# ============================ CONFRONTO =====================================
# file centrale dei materiali per la tendina (in Definitivi, un posto
# solo, lo aggiorna l'utente mano a mano); se manca si usa la riserva
MATERIALI_XLSX = [
    r"C:\Users\User\Desktop\CLAUDE\Definitivi\_MATERIALI_SEZIONATURA.xlsx",
    r"C:\Users\User\Dropbox\STEFANO\Matteo\CLAUDE\Definitivi"
    r"\_MATERIALI_SEZIONATURA.xlsx",
    r"C:\Users\User\Dropbox\STEFANO\Matteo\_MATERIALI_SEZIONATURA.xlsx",
]
MATERIALI_RISERVA = [
    "MD PLACCATO 2 LATI sp.19", "MD PLACCATO 2 LATI sp.10",
    "MD PLACCATO 1 LATO + CONTR. sp.16", "MD PLACCATO 1 LATO + CONTR. sp.8",
    "MDF GREZZO sp.19", "MDF GREZZO sp.25", "MULTISTRATO", "NOBILITATO"]
MAGG_TAGLIO = 10        # mm in piu' su L e H in sezionatura (SP no)


def carica_distinte(base, log=print):
    """Cerca *Distinta*.xls* accanto al lavoro (la cartella sopra DXF
    DEFINITIVI e le sue sottocartelle di primo livello) e ne tira fuori
    Codice/Pezzo -> Materiale, con le misure per sciogliere gli omonimi.
    Serve quando STP/DWG non portano il materiale: la colonna MATERIALE
    del riepilogo si compila da sola dalla distinta del mobile."""
    from openpyxl import load_workbook
    voci = []                    # (CODICE, [misure ordinate], materiale)
    radice = Path(base).parent
    cand = [f for pat in ("*[Dd]istinta*.xls*", "*/*[Dd]istinta*.xls*")
            for f in radice.glob(pat) if not f.name.startswith("~$")]
    # --- esportazione "Distinta base" di CATIA: .txt con tabella a | ----
    for f in [x for pat in ("*.txt", "*/*.txt") for x in radice.glob(pat)]:
        try:
            if f.stat().st_size > 2_000_000:
                continue
            testa = f.read_bytes()[:4000].decode("cp1252", "replace")
            if "Distinta base" not in testa or "Materiale" not in testa:
                continue
            n0 = len(voci)
            ic = im = None
            idx = []
            for riga in f.read_bytes().decode("cp1252",
                                              "replace").splitlines():
                if "|" not in riga:
                    continue
                celle = [c.strip() for c in riga.strip().strip("|").split("|")]
                up = [c.upper() for c in celle]
                if "MATERIALE" in up and ("NUMERO PARTE" in up
                                          or "CODICE" in up):
                    ic = (up.index("NUMERO PARTE") if "NUMERO PARTE" in up
                          else up.index("CODICE"))
                    im = up.index("MATERIALE")
                    idx = [up.index(k) for k in
                           ("LUNGHEZZA", "LARGHEZZA", "SPESSORE") if k in up]
                    continue
                if ic is None or len(celle) <= max(ic, im):
                    continue
                cod, mat = celle[ic], celle[im]
                if not cod or not mat:
                    continue

                def _num(v):
                    try:
                        return float(str(v).replace(",", "."))
                    except (TypeError, ValueError):
                        return None

                dd = sorted((x for x in (_num(celle[i]) for i in idx
                                         if i < len(celle))
                             if x is not None), reverse=True)
                voci.append((cod.upper(), dd, mat, "catia"))
            if len(voci) > n0:
                log(f"  distinta materiali (CATIA): {f.name} "
                    f"({len(voci) - n0} voci)")
        except Exception as ex:
            log(f"  ! distinta {f.name} non leggibile ({ex})")
    # CATIA tronca i materiali alla larghezza colonna: se il pezzo di
    # nome combacia con UNA sola voce della lista centrale, si completa
    try:
        pieni = carica_materiali()

        def _chiave(v):
            return " ".join(v.split()).upper()

        for i, v in enumerate(voci):
            if any(_chiave(v[2]) == _chiave(p) for p in pieni):
                continue
            trovati = [p for p in pieni
                       if _chiave(p).startswith(_chiave(v[2]))]
            if len(trovati) > 1 and v[1]:
                # piu' varianti (sp.8/18/25...): decide lo SPESSORE
                sp = f"{min(v[1]):g}"
                stretti = [p for p in trovati
                           if re.search(rf"(?<!\d){re.escape(sp)}(?!\d)", p)]
                if len(stretti) == 1:
                    trovati = stretti
            if len(trovati) == 1:
                voci[i] = (v[0], v[1], trovati[0]) + v[3:]
    except Exception:
        pass

    for f in cand:
        n0 = len(voci)
        try:
            wb = load_workbook(str(f), read_only=True, data_only=True)
            for ws in wb.worksheets:
                righe = ws.iter_rows(values_only=True)
                intest = None
                for r in righe:
                    vals = [str(v).strip().upper() if v else "" for v in r]
                    if "MATERIALE" in vals and ("CODICE" in vals
                                                or "PEZZO" in vals):
                        intest = vals
                        break
                if not intest:
                    continue
                ic = (intest.index("CODICE") if "CODICE" in intest
                      else intest.index("PEZZO"))
                im = intest.index("MATERIALE")
                idx = {}
                for k in ("LUNGHEZZA", "LARGHEZZA", "SPESSORE"):
                    if k in intest:
                        idx[k] = intest.index(k)

                def num(v):
                    try:
                        return float(str(v).replace(",", "."))
                    except (TypeError, ValueError):
                        return None

                for r in righe:
                    cod = r[ic] if len(r) > ic else None
                    mat = r[im] if len(r) > im else None
                    if not (isinstance(cod, str) and cod.strip()
                            and isinstance(mat, str) and mat.strip()):
                        continue
                    dd = sorted((x for x in (num(r[i]) if len(r) > i
                                             else None
                                             for i in idx.values())
                                 if x is not None), reverse=True)
                    voci.append((cod.strip().upper(), dd, mat.strip(),
                                 "xlsx"))
            if len(voci) > n0:
                log(f"  distinta materiali: {f.name} "
                    f"({len(voci) - n0} voci)")
        except Exception as ex:
            log(f"  ! distinta {f.name} non leggibile ({ex})")
    return voci


def materiali_da_distinta(rows, voci, log=print):
    """Compila la colonna MATERIALE (r[7]) delle righe che ne sono senza,
    abbinando PEZZO -> Codice della distinta; con gli omonimi decide la
    riga con le misure piu' vicine (scarta oltre 30 mm complessivi)."""
    if not voci:
        return
    riempiti = 0
    for r in rows:
        if r[7]:
            continue
        nome = r[1].strip().upper()
        cands = [v for v in voci if v[0] == nome]
        if not cands:
            # CATIA tronca i nomi lunghi alla larghezza colonna:
            # vale il codice come INIZIO del nome (se abbastanza lungo)
            cands = [v for v in voci
                     if len(v[0]) >= 15 and nome.startswith(v[0])]
        if not cands:
            continue
        if len(cands) > 1:
            dr = sorted((x for x in (r[4], r[5], r[6]) if x), reverse=True)

            def scarto(v):
                if len(v[1]) != len(dr):
                    return 1e9
                return sum(abs(a - b) for a, b in zip(dr, v[1]))

            def rango(v):
                # a parita' di misure vince la distinta Excel (curata a
                # mano) sulla Distinta base CATIA
                return 0 if (len(v) > 3 and v[3] == "xlsx") else 1

            cands.sort(key=lambda v: (scarto(v), rango(v)))
            if scarto(cands[0]) > 30:
                # misure lontane (pezzo sagomato: l'ingombro non e' la
                # quota del pannello) ma nome esatto: se TUTTE le fonti
                # dicono lo stesso materiale, vale quello
                if len({" ".join(c[2].split()).upper() for c in cands}) == 1:
                    r[7] = cands[0][2]
                    riempiti += 1
                continue
        r[7] = cands[0][2]
        riempiti += 1
    if riempiti:
        log(f"  materiali dalla distinta: {riempiti}/{len(rows)} pezzi")


def carica_materiali():
    """Legge i materiali da TUTTE le copie e le riallinea (fusione):
    un'aggiunta fatta su una qualsiasi finisce anche nelle altre al
    giro dopo. La fusione non cancella mai: per togliere una voce va
    tolta da tutte le copie a mano."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    def chiave(v):
        # doppione anche se cambia maiuscole/minuscole o spazi doppi
        return " ".join(v.split()).upper()

    visti = []
    chiavi = set()
    per_file = {}
    for percorso in MATERIALI_XLSX:
        try:
            wb = load_workbook(percorso, read_only=True, data_only=True)
            ws = wb.active
            elenco = []
            for r in ws.iter_rows(values_only=True):
                v = r[0] if r else None
                if (isinstance(v, str) and v.strip()
                        and chiave(v) != "MATERIALE"):
                    elenco.append(v.strip())
            per_file[percorso] = elenco
            for v in elenco:
                k = chiave(v)
                if k not in chiavi:
                    chiavi.add(k)
                    visti.append(v)
        except Exception:
            per_file[percorso] = None
    if not visti:
        return MATERIALI_RISERVA
    # riallinea le due copie in Definitivi (mai bloccante: una copia
    # aperta in Excel si aggiorna al giro dopo)
    for percorso in MATERIALI_XLSX[:2]:
        if per_file.get(percorso) == visti:
            continue
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "MATERIALI"
            ws.cell(1, 1, "MATERIALE").font = Font(bold=True)
            for i, m in enumerate(visti, 2):
                ws.cell(i, 1, m)
            ws.column_dimensions["A"].width = 60
            wb.save(percorso)
        except Exception:
            pass
    return visti


def write_report(out, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook(); ws = wb.active; ws.title = "PROGRAMMI_UNICI"
    hdr = ["FILE PROGRAMMA", "PEZZO", "GRUPPI", "Q.TA TOT",
           "L (mm)", "H (mm)", "SP (mm)", "MATERIALE", "N.FORI", "TIPO",
           "POSIZIONI ORIGINALI"]
    bold = Font(bold=True); fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(border_style="thin", color="888888")
    bd = Border(thin, thin, thin, thin)
    for c, t in enumerate(hdr, 1):
        cl = ws.cell(1, c, t); cl.font = bold; cl.fill = fill; cl.border = bd
    # ordina per numero pezzo se il nome inizia con un numero, poi per nome
    def sort_key(r):
        m = re.match(r"(\d+)", r[1])
        return (int(m.group(1)) if m else 10**9, r[1].lower(), r[0])
    rows.sort(key=sort_key)
    green = PatternFill("solid", fgColor="E2EFDA")
    for i, r in enumerate(rows, 2):
        # r = (outname, name, gruppi, qty, L, H, SP, materiale, nfori,
        #      tipo, posizioni, uguale_ovunque)
        for c, v in enumerate(r[:11], 1):
            cell = ws.cell(i, c, v); cell.border = bd
            if r[11]:   # pezzo uguale in tutti i gruppi dove compare -> evidenzia
                cell.fill = green
    for col, w in zip("ABCDEFGHIJK", [40, 30, 28, 9, 9, 9, 9, 26, 8, 7, 60]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    # MATERIALE: menu a tendina dal file centrale _MATERIALI_SEZIONATURA
    # (foglio nascosto MATERIALI nel workbook), testo libero ammesso
    mats = carica_materiali()
    if rows:
        wm = wb.create_sheet("MATERIALI")
        for i, m in enumerate(mats, 1):
            wm.cell(i, 1, m)
        wm.sheet_state = "hidden"
        dv = DataValidation(type="list",
                            formula1=f"=MATERIALI!$A$1:$A${len(mats)}",
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"H2:H{len(rows) + 1}")

    # --- SEZIONATURA: riga per riga (niente raggruppamenti), tutto in
    # FORMULA sul riepilogo: compili/correggi MATERIALE (o misure) su
    # PROGRAMMI_UNICI e qui si aggiorna da solo. L e H TAGLIO = quota
    # pezzo + MAGG_TAGLIO (10 mm); lo spessore resta quello vero.
    sz = wb.create_sheet("SEZIONATURA")
    hdr2 = ["FILE PROGRAMMA", "PEZZO", "L TAGLIO (mm)", "H TAGLIO (mm)",
            "SP (mm)", "Q.TA", "MATERIALE"]
    for c, t in enumerate(hdr2, 1):
        cl = sz.cell(1, c, t); cl.font = bold; cl.fill = fill; cl.border = bd
    col_pu = {"PEZZO": ("B", ""), "L TAGLIO (mm)": ("E", f"+{MAGG_TAGLIO}"),
              "H TAGLIO (mm)": ("F", f"+{MAGG_TAGLIO}"),
              "SP (mm)": ("G", ""), "Q.TA": ("D", ""), "MATERIALE": ("H", "")}
    for i, r in enumerate(rows, 2):
        sz.cell(i, 1, r[0]).border = bd
        for c, t in enumerate(hdr2[1:], 2):
            col, extra = col_pu[t]
            f = (f"=INDEX(PROGRAMMI_UNICI!{col}:{col},"
                 f"MATCH($A{i},PROGRAMMI_UNICI!$A:$A,0)){extra}")
            cell = sz.cell(i, c, f); cell.border = bd
    for col, w in zip("ABCDEFG", [40, 30, 13, 13, 9, 7, 26]):
        sz.column_dimensions[col].width = w
    sz.freeze_panes = "A2"
    dst = out / "_RIEPILOGO_programmi_unici.xlsx"
    try:
        wb.save(dst)
    except PermissionError:
        dst = out / "_RIEPILOGO_programmi_unici_NEW.xlsx"; wb.save(dst)
    return dst


def run_comparison(base, groups, log=print):
    import ezdxf
    recs = {}
    noti = {" ".join(m.split()).upper() for m in carica_materiali()}
    total = sum(g["n"] for g in groups); done = 0
    for gr in groups:
        for p in sorted(Path(gr["path"]).glob("*.dxf")):
            try:
                doc = ezdxf.readfile(str(p))
            except Exception as ex:
                log(f"  ! {gr['label']}/{p.name}: non leggibile ({ex})"); continue
            e = model_solid(doc)
            if e is not None:
                g = None
                try:                               # DXF 2013+: ACIS binario
                    if e.sab:
                        g = geom_sig(e.sab)
                except Exception:
                    g = None
                if g is None:
                    try:                           # DXF 2000-2010: ACIS testo
                        if e.sat:
                            g = geom_sig_sat(e.sat)
                    except Exception as ex:
                        log(f"  ! {gr['label']}/{p.name}: ACIS non leggibile ({ex})")
                name = piece_name(e, p.stem) if g else None
            else:
                g = geom_sig_2d(doc)               # niente solido: provo in 2D
                name = clean_name(p.stem) if g else None
            if g is None:
                log(f"  ! {gr['label']}/{p.name}: geometria vuota, saltato"); continue
            g["name"] = name
            g["mat"] = ((materiale_di(doc, e) or materiale_xdata(e)
                         or materiale_layer(e, noti))
                        if e is not None else "")
            g["label"] = gr["label"]
            g["src"] = str(p)                      # percorso reale del file
            recs[f"{gr['rel']}|{p.name}"] = g      # chiave per CARTELLA (univoca)
            done += 1
            if done % 5 == 0 or done == total:
                log(f"  letti {done}/{total} file...")
        log(f"  gruppo {gr['label']} completato ({gr['n']} file)")

    groups_by = defaultdict(list); name_hashes = defaultdict(set)
    for k, g in recs.items():
        fn = k.split("|", 1)[1]
        groups_by[(g["name"], g["hash"])].append((g["label"], fn, k))
        name_hashes[g["name"]].add(g["hash"])

    out = Path(base) / OUT_DIRNAME
    out.mkdir(exist_ok=True)
    safe = lambda s: re.sub(r'[<>:"/\\|?*]', "_", str(s)).strip()

    # --- piano dei file di destinazione --------------------------------------
    manifest, rows, plan = {}, [], {}
    for (name, h), members in groups_by.items():
        labels_cov = sorted(set(lb for lb, fn, k in members))
        need = len(name_hashes[name]) > 1          # gruppo nel nome solo se serve
        suffix = "_" + "+".join(safe(lb) for lb in labels_cov) if need else ""
        outname = f"{safe(name)}{suffix}.dxf"
        if len(outname) > 120:                     # etichette lunghe: accorcia
            outname = f"{safe(name)}_{h[:8]}.dxf"
        # nome univoco: se due pezzi DIVERSI ricadono sullo stesso nome, numera
        if outname in plan:
            stem, ext = os.path.splitext(outname); j = 2
            while f"{stem}_{j}{ext}" in plan:
                j += 1
            outname = f"{stem}_{j}{ext}"
        # sorgente presa dal PERCORSO REALE memorizzato (no ambiguita')
        src_lb, src_fn, src_key = sorted(members)[0]
        g0 = recs[src_key]
        plan[outname] = Path(g0["src"])
        manifest[outname] = {"name": name, "groups": labels_cov, "qty": len(members),
                             "dims": g0["dims"], "nholes": g0["nholes"],
                             "tipo": g0["tipo"],
                             "members": [(lb, fn) for lb, fn, k in members]}
        rows.append([outname, name, "+".join(labels_cov),
                     len(members), g0["dims"][0], g0["dims"][1], g0["dims"][2],
                     g0.get("mat", ""), g0["nholes"], g0["tipo"],
                     "; ".join(f"{lb}/{fn}" for lb, fn, k in sorted(members)),
                     not need])   # ultimo campo: pezzo uguale in tutti i gruppi

    # --- sincronizzazione MINIMA: tocca solo cio' che manca o e' diverso ------
    # (cosi' file gia' giusti ma aperti in AutoCAD/bloccati da Dropbox
    #  non vengono riscritti e non fanno fallire il lavoro)
    bloccati = []
    for p in out.glob("*.dxf"):                       # rimuovi i file non previsti
        if p.name not in plan:
            if not _unlink_retry(p):
                bloccati.append(p.name)
                log(f"  ! {p.name}: non previsto ma bloccato, lascialo o cancellalo a mano")
    for outname, src in plan.items():                 # copia mancanti/diversi
        dst = out / outname
        try:
            if dst.exists() and dst.stat().st_size == src.stat().st_size \
               and dst.read_bytes() == src.read_bytes():
                continue                              # gia' identico: non toccare
        except PermissionError:
            pass
        try:
            _copy_retry(src, dst)
        except PermissionError:
            bloccati.append(outname)
            log(f"  ! {outname}: BLOCCATO (aperto in AutoCAD?), non aggiornato")
        except FileNotFoundError:
            bloccati.append(outname)
            log(f"  ! {outname}: sorgente non trovato ({src}), saltato")

    (Path(base) / "_manifest.json").write_text(json.dumps(manifest))
    materiali_da_distinta(rows, carica_distinte(base, log), log)
    report = write_report(out, rows)
    return {"unici": len(manifest), "file_letti": len(recs),
            "uguali_tutti": sum(1 for r in rows if r[11]),
            "out": str(out), "report": report.name, "bloccati": bloccati}


# ============================ GUI ==========================================
def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, ttk, messagebox, simpledialog
    root = tk.Tk()
    root.title("File Unici DXF - versione generica")
    root.geometry("820x600")
    state = {"base": None, "groups": []}
    q = queue.Queue()

    head = tk.Label(root, text="File Unici DXF  (qualsiasi cartella, ricorsivo)",
                    font=("Segoe UI", 14, "bold"))
    head.pack(anchor="w", padx=12, pady=(10, 0))

    top = tk.Frame(root); top.pack(fill="x", padx=12, pady=6)
    tk.Label(top, text="1) Cartella da analizzare (legge tutti i .dxf, anche nelle sottocartelle):",
             font=("Segoe UI", 10, "bold")).pack(anchor="w")
    row = tk.Frame(top); row.pack(fill="x", pady=4)
    base_var = tk.StringVar()
    tk.Entry(row, textvariable=base_var, state="readonly").pack(side="left", fill="x", expand=True)

    def choose():
        d = filedialog.askdirectory(title="Scegli la cartella con i file DXF")
        if d:
            base_var.set(d); state["base"] = d
            state["groups"] = detect_groups(d); refresh()
            if not state["groups"]:
                messagebox.showwarning("Nessun DXF",
                    "In questa cartella (e sottocartelle) non ho trovato file .dxf.")
    tk.Button(row, text="Sfoglia...", command=choose).pack(side="left", padx=6)

    tk.Label(top, text="2) Gruppi rilevati  (doppio clic su una riga per cambiare l'etichetta del gruppo):",
             font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(8, 0))
    tree = ttk.Treeview(top, columns=("gruppo", "cartella", "n"), show="headings", height=7)
    for c, t, w in (("gruppo", "Gruppo", 160), ("cartella", "Sottocartella", 440), ("n", "N. DXF", 70)):
        tree.heading(c, text=t); tree.column(c, width=w, anchor="w")
    tree.pack(fill="x", pady=4)

    def refresh():
        tree.delete(*tree.get_children())
        for i, g in enumerate(state["groups"]):
            tree.insert("", "end", iid=str(i),
                        values=(g["label"], g["rel"] or "(cartella principale)", g["n"]))

    def edit_label(_ev):
        iid = tree.focus()
        if not iid:
            return
        i = int(iid)
        new = simpledialog.askstring("Etichetta gruppo",
            f"Etichetta per la cartella:\n{state['groups'][i]['rel'] or '(cartella principale)'}",
            initialvalue=state["groups"][i]["label"])
        if new and new.strip():
            state["groups"][i]["label"] = new.strip(); refresh()
    tree.bind("<Double-1>", edit_label)

    runbtn = tk.Button(root, text="▶  CREA FILE UNICI",
                       bg="#2e7d32", fg="white", font=("Segoe UI", 13, "bold"),
                       height=2, command=lambda: start())
    runbtn.pack(fill="x", padx=12, pady=8)

    logbox = tk.Text(root, height=14, wrap="word")
    logbox.pack(fill="both", expand=True, padx=12, pady=(0, 10))

    def append(msg):
        logbox.insert("end", msg + "\n"); logbox.see("end")

    def worker():
        try:
            res = run_comparison(state["base"], state["groups"],
                                 log=lambda m: q.put(("log", m)))
            q.put(("done", res))
        except Exception:
            import traceback; q.put(("err", traceback.format_exc()))

    def poll():
        try:
            while True:
                kind, payload = q.get_nowait()
                if kind == "log":
                    append(payload)
                elif kind == "done":
                    append(f"\n=== FATTO: {payload['unici']} file unici "
                           f"({payload['uguali_tutti']} uguali in tutti i gruppi) ===")
                    append(f"Cartella:  {payload['out']}")
                    append(f"Riepilogo: {payload['report']}")
                    runbtn.config(state="normal", text="▶  CREA FILE UNICI")
                    try: os.startfile(payload["out"])
                    except Exception: pass
                    msg = f"{payload['unici']} file unici creati in:\n{payload['out']}"
                    if payload.get("bloccati"):
                        msg += ("\n\nATTENZIONE - file bloccati non aggiornati "
                                "(chiudili in AutoCAD e rilancia):\n  "
                                + "\n  ".join(payload["bloccati"]))
                        messagebox.showwarning("Fatto (con avvisi)", msg)
                    else:
                        messagebox.showinfo("Fatto", msg)
                elif kind == "err":
                    append("ERRORE:\n" + payload)
                    runbtn.config(state="normal", text="▶  CREA FILE UNICI")
                    messagebox.showerror("Errore", payload.splitlines()[-1])
        except queue.Empty:
            pass
        root.after(150, poll)

    def start():
        if not state["base"] or not state["groups"]:
            messagebox.showwarning("Attenzione",
                "Scegli prima la cartella con i file DXF."); return
        runbtn.config(state="disabled", text="Elaborazione in corso...")
        logbox.delete("1.0", "end"); append("Avvio confronto...")
        threading.Thread(target=worker, daemon=True).start()

    root.after(150, poll)
    root.mainloop()


# ============================ MAIN =========================================
if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--run":
        base = sys.argv[2]
        logf = open(Path(base) / "_confronto_log.txt", "w", encoding="utf-8")

        def _log(msg):
            print(msg)
            logf.write(str(msg) + "\n"); logf.flush()

        try:
            groups = detect_groups(base)
            _log(f"Gruppi: {[(g['label'], g['n']) for g in groups]}")
            _log(f"RESULT: {run_comparison(base, groups, log=_log)}")
        except Exception:
            import traceback
            _log("ERRORE:\n" + traceback.format_exc())
            sys.exit(1)
        finally:
            logf.close()
    else:
        launch_gui()
