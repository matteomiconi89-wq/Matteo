#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
CONFRONTO PROGRAMMI DXF  ->  PROGRAMMI UNICI   (app con un pulsante)
====================================================================
Confronta i programmi CNC .dxf di piu' camere/cartelle, trova quelli
geometricamente IDENTICI (misura + forature reali, non i byte) e crea
una cartella PROGRAMMI_UNICI con i soli unici + un riepilogo Excel.

COME FUNZIONA IL CONFRONTO:
  per ogni .dxf legge il solido del Model Space e ne estrae la geometria
  vera dall'ACIS (vertici 'point' + fori 'ellipse'), la normalizza
  (trasla allo spigolo, arrotonda, ordina) e la confronta. Cosi' due pezzi
  identici esportati separatamente da STEP risultano uguali, ignorando
  data/ora e ID interni che cambiano sempre.

NOMI FILE:
  - pezzo identico ovunque  -> Nome.dxf            (niente camera)
  - pezzo diverso per camera -> Nome_A115.dxf ...  (camera in coda)
  I doppioni con stesso nome+geometria si fondono (qty = n. posizioni).

USO:
  - doppio clic (o l'.exe) -> finestra con il pulsante
  - da riga di comando:  py confronta_file_unici_dxf.py --run "C:\\...\\cartella"
Richiede (solo per .py): ezdxf, openpyxl  ->  py -m pip install ezdxf openpyxl
"""
import sys, os, re, json, shutil, hashlib, threading, queue, time
from pathlib import Path
from collections import defaultdict

NDEC = 2  # arrotondamento coordinate (mm) per la firma geometrica


# --- file su Dropbox/OneDrive: la sincronizzazione blocca i file appena
# --- creati/cancellati -> ogni operazione va ritentata con attesa crescente
def _copy_retry(src, dst, tries=12):
    for i in range(tries):
        try:
            return shutil.copy2(src, dst)
        except PermissionError:
            time.sleep(0.5 * (i + 1))
    # ultimo tentativo: copia manuale dei byte (aggira il lock di CopyFile2)
    Path(dst).write_bytes(Path(src).read_bytes())


def _unlink_retry(p, tries=8):
    for i in range(tries):
        try:
            Path(p).unlink()
            return True
        except FileNotFoundError:
            return True
        except PermissionError:
            time.sleep(0.5 * (i + 1))
    return False


# ============================ GEOMETRIA ====================================
def model_solids(doc):
    """TUTTI i 3DSOLID del file: i frontali cassetto uniti (SX+CX+DX) sono
    file con 3 solidi e vanno confrontati sull'insieme, non sul primo."""
    return [e for e in doc.modelspace() if e.dxftype() == "3DSOLID"]


# Pulizia del nome pezzo letto dal blocco (Xdata ESPSOL_NOME). Il nome "vero" e'
# sempre "NN_Descrizione" (NN = numero pezzo, 1-2 cifre, es. 01_Fianco_SX_EXT).
# Davanti puo' esserci spazzatura di commessa/revisione di forma varia, es.:
#   "J-03.02_05_Sottotop"             -> "05_Sottotop"
#   "J-03.02_R_100_01 _Fianco_SX_EXT" -> "01_Fianco_SX_EXT"
# Regola: se il nome inizia col codice commessa (J...), si butta tutto cio' che
# precede il numero pezzo (preso come l'ULTIMO numero di 1-2 cifre seguito da
# separatore e da una LETTERA; i numeri della spazzatura -03,02,100- sono seguiti
# da altri numeri o token, non da una parola).
# In coda puo' esserci il suffisso "_bis" (es. "16_Sponda_DX_bis"): va SEMPRE
# tolto dal nome. NB: il confronto resta per GEOMETRIA -> se il pezzo bis e'
# identico al non-bis si fondono in 1 riga, se diverso restano 2 righe (la camera
# si vede in POSIZIONI ORIGINALI).
_BIS_SUFFIX = re.compile(r"[\s_.\-]+bis\s*$", re.IGNORECASE)
_JOB_PREFIX = re.compile(r"^\s*[Jj][-_.\s]?\d")          # inizia col codice commessa
_PIECE_START = re.compile(r"\d{1,2}[\s_.\-]+[A-Za-z]")   # numero pezzo + sep + lettera


def clean_name(name):
    if not name:
        return name
    name = _BIS_SUFFIX.sub("", name).strip()
    if _JOB_PREFIX.match(name):                 # c'e' spazzatura davanti al nome
        starts = list(_PIECE_START.finditer(name))
        if starts:                              # taglia fino all'ultimo numero pezzo
            name = name[starts[-1].start():]
    # normalizza il separatore subito dopo il numero pezzo: "01 _F" -> "01_F"
    name = re.sub(r"^\s*(\d{1,2})[\s_.\-]+", r"\1_", name)
    return name.strip()


def piece_name(e, fallback):
    name = fallback
    try:
        for c, v in e.get_xdata("ESPSOL_NOME"):
            if c == 1000 and v:
                name = v
                break
    except Exception:
        pass
    return clean_name(name)


def _sig_3d(pts, ell, ndec=NDEC):
    """Normalizza punti + ellissi (fori) e produce la firma."""
    if not pts:
        return None
    mn = [min(p[i] for p in pts) for i in range(3)]
    mx = [max(p[i] for p in pts) for i in range(3)]
    dims = tuple(sorted((round(mx[i] - mn[i], 1) for i in range(3)), reverse=True))
    npts = sorted(tuple(round(p[i] - mn[i], ndec) for i in range(3)) for p in pts)
    nell = sorted((round(r, 2), tuple(round(c[i] - mn[i], ndec) for i in range(3)))
                  for c, r in ell)
    h = hashlib.sha1(repr((npts, nell)).encode()).hexdigest()[:12]
    return {"dims": list(dims), "nholes": len(ell), "hash": h}


# --- transform del SAT: AutoCAD scrive il corpo ACIS in coordinate LOCALI
# --- + un record "transform" (colonne assi + traslazione) da APPLICARE,
# --- senno' rispaziature/rotazioni dell'export si perdono (i frontali
# --- uniti leggevano fuga 2.5 invece di 3.2 e il totale corto di 1.4)
def _tr_leggi(numeri):
    if len(numeri) < 12:
        return None
    cx, cy, cz, t = numeri[0:3], numeri[3:6], numeri[6:9], tuple(numeri[9:12])
    ident = (abs(cx[0]-1) < 1e-12 and abs(cy[1]-1) < 1e-12
             and abs(cz[2]-1) < 1e-12
             and all(abs(v) < 1e-12 for v in (cx[1], cx[2], cy[0], cy[2],
                                              cz[0], cz[1]))
             and all(abs(v) < 1e-12 for v in t))
    return None if ident else ((cx, cy, cz), t)


def _tr_punto(tr, p):
    (cx, cy, cz), t = tr
    return (p[0]*cx[0] + p[1]*cy[0] + p[2]*cz[0] + t[0],
            p[0]*cx[1] + p[1]*cy[1] + p[2]*cz[1] + t[1],
            p[0]*cx[2] + p[1]*cy[2] + p[2]*cz[2] + t[2])


def _raw_sab(sab):
    """Vertici + fori dall'ACIS binario (SAB, DXF 2013+)."""
    from ezdxf.acis.sab import Decoder
    from ezdxf.acis.const import Tags
    dec = Decoder(sab); dec.read_header()
    pts, ell = [], []
    tr = None
    for rec in dec.read_records():
        if not rec:
            continue
        nm = rec[0].value if isinstance(rec[0].value, str) else ""
        locs = [t.value for t in rec if t.tag == Tags.LOCATION_VEC]
        dirs = [t.value for t in rec if t.tag == Tags.DIRECTION_VEC]
        if nm.startswith("transform") and tr is None:
            numeri = []
            for tk in rec:
                if isinstance(tk.value, (int, float)):
                    numeri.append(float(tk.value))
                elif isinstance(tk.value, (tuple, list)):
                    numeri.extend(float(v) for v in tk.value)
            tr = _tr_leggi(numeri)
        elif nm.startswith("point"):
            pts.extend(locs)
        elif "ellipse" in nm and locs:
            maj = dirs[-1] if dirs else (0, 0, 0)
            ell.append((locs[0], (maj[0]**2 + maj[1]**2 + maj[2]**2) ** 0.5))
    return pts, ell, tr


def _raw_sat(sat_lines):
    """Vertici + fori dall'ACIS testuale (SAT, DXF 2000-2010)."""
    from ezdxf.acis.sat import parse_sat
    pts, ell = [], []
    tr = None
    for rec in parse_sat(sat_lines).entities:
        try:
            if rec.name == "transform" and tr is None:
                numeri = []
                for v in rec.data:
                    try:
                        numeri.append(float(v))
                    except (TypeError, ValueError):
                        pass
                tr = _tr_leggi(numeri)
            elif rec.name.startswith("point"):
                pts.append(tuple(float(v) for v in rec.data[:3]))
            elif "ellipse" in rec.name:
                d = rec.data
                c = tuple(float(v) for v in d[0:3])
                maj = [float(v) for v in d[6:9]]
                ell.append((c, (maj[0]**2 + maj[1]**2 + maj[2]**2) ** 0.5))
        except (ValueError, IndexError):
            continue
    return pts, ell, tr


def geom_sig_solids(solids, ndec=NDEC):
    """Firma geometrica dell'INSIEME dei solidi del file (1 o piu').
    La FIRMA (hash) usa le coordinate LOCALI dell'ACIS: sono uguali per i
    pezzi duplicati e non dipendono da come l'export ha girato il pezzo
    (raggruppamento identico a come e' sempre stato). Le MISURE invece
    vengono dalla geometria con la TRANSFORM applicata: sono le quote VERE
    (es. frontali uniti rispaziati a fuga 3.2 -> totale 2645.4, non 2644)."""
    pts, ell, pts_w = [], [], []
    for e in solids:
        if e.sab:
            p, l, tr = _raw_sab(e.sab)
        elif e.sat:
            p, l, tr = _raw_sat(e.sat)
        else:
            continue
        pts.extend(p); ell.extend(l)
        pts_w.extend([_tr_punto(tr, q) for q in p] if tr else p)
    sig = _sig_3d(pts, ell, ndec)
    if sig and pts_w:
        mn = [min(p[i] for p in pts_w) for i in range(3)]
        mx = [max(p[i] for p in pts_w) for i in range(3)]
        sig["dims"] = sorted((round(mx[i] - mn[i], 1) for i in range(3)),
                             reverse=True)
    return sig


# ============================ RILEVA CAMERE =================================
def room_code(folder_name):
    """Ricava il codice camera (es. B204, A115) dal nome cartella."""
    for part in re.split(r"[ _\-.]+", folder_name):
        if re.fullmatch(r"[A-Za-z]{1,3}\d{2,4}", part):
            return part.upper()
    return folder_name


def detect_rooms(base):
    rooms = []
    for sub in sorted(Path(base).iterdir()):
        if sub.is_dir() and sub.name != "PROGRAMMI_UNICI":
            dxfs = list(sub.glob("*.dxf"))
            if dxfs:
                rooms.append({"code": room_code(sub.name), "folder": sub.name,
                              "path": str(sub), "n": len(dxfs)})
    return rooms


# ============================ REGOLE CNC/BORDATURA ==========================
# Compilazione automatica delle colonne MATERIALE / CNC / BORDATURA in base
# al NUMERO del pezzo (le prime cifre del nome, es. "09_Frontale..." -> 09).
# Default = regole apprese dal riepilogo compilato a mano dall'utente (21032).
# Ogni lavoro puo' avere il suo file _REGOLE_CNC.xlsx nella cartella del
# lavoro: se non c'e' viene creato coi default, e l'utente lo puo' MODIFICARE.
REGOLE_DEFAULT = {
    "01": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "SQUAD+MASTER",   "NO BORDO"),
    "02": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "SQUAD+MASTER",   "NO BORDO"),
    "03": ("MD PLACCATO 1 LATO + CONTR. sp.16 (finito 16)", "SQUAD+MASTER",   "NO BORDO"),
    "04": ("MD PLACCATO 1 LATO + CONTR. sp.16 (finito 16)", "SQUAD+MASTER",   "NO BORDO"),
    "05": ("MDF GREZZO sp.19",                              "HOMAG",          "NO BORDO"),
    "06": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "HOMAG",          "4 LATI + BORDO 6/10 SU SCASSI"),
    "07": ("MD PLACCATO 1 LATO + CONTR. sp.16 (finito 16)", "HOMAG",          "NO BORDO"),
    "08": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "HOMAG",          "4 LATI + BORDO 6/10 SU SCASSI"),
    "09": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "HOMAG",          "4 LATI"),
    "10": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "SOLO SQUADRATA", "4 LATI"),
    "11": ("MDF GREZZO sp.19",                              "SOLO SQUADRATA", "NO BORDO"),
    "12": ("",                                              "",               ""),
    "14": ("MDF GREZZO sp.19",                              "MASTER",         "NO BORDO"),
    "15": ("MD PLACCATO 2 LATI sp.10 (finito 11)",          "SQUAD+BORD MASTER", "2 LATI LUNGHI"),
    "16": ("MD PLACCATO 2 LATI sp.10 (finito 11)",          "SQUAD+BORD MASTER", "2 LATI LUNGHI"),
    "17": ("MD PLACCATO 2 LATI sp.10 (finito 11)",          "SQUAD+BORD MASTER", "2 LATI LUNGHI"),
    "18": ("MD PLACCATO 2 LATI sp.10 (finito 11)",          "SQUAD+BORD MASTER", "2 LATI LUNGHI"),
    "19": ("MD PLACCATO 2 LATI sp.10 (finito 11)",          "SQUAD+BORD MASTER", "2 LATI LUNGHI"),
    "20": ("MD PLACCATO 1 LATO + CONTR. sp.8 (finito 9)",   "SOLO SQUADRATA", "NO BORDO"),
    "21": ("MD PLACCATO 1 LATO + CONTR. sp.16 (finito 16)", "SQUAD+MASTER",   "NO BORDO"),
    "22": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "MASTER",         "NO BORDO"),
    "24": ("MDF GREZZO sp.19",                              "HOMAG",          "NO BORDO"),
    "25": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "SQUAD+MASTER",   "NO BORDO"),
    "31": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "SQUAD+MASTER",   "NO BORDO"),
    "41": ("MD PLACCATO 2 LATI sp.19 (finito 19)",          "SQUAD+MASTER",   "NO BORDO"),
}

# Pezzi il cui singolo programma rappresenta PIU' pezzi fisici: la Q.TA TOT
# nel riepilogo va moltiplicata per questo fattore.
QTY_MULT = {"24": 4, "25": 2}


def _camera_rl(folder):
    """True se il nome della cartella camera ha il suffisso _R o _L (DX/SX).
    Usato dal pezzo 10_Base: se la camera e' _R/_L il CNC diventa HOMAG."""
    return bool(re.search(r"_[RL](?=[_\-. ]|$)", folder or "", re.IGNORECASE))


def _norm_num(digits):
    """Numero pezzo normalizzato a 2 cifre togliendo gli zeri davanti:
    '031'->'31', '041'->'41', '01'->'01', '9'->'09'. Cosi' i pezzi scritti a
    3 cifre (031, 041) combaciano con le regole a 2 cifre (31, 41)."""
    s = str(digits).strip().lstrip("0") or "0"
    return s.zfill(2)


def num_pezzo(name):
    """'09_Frontale_Cassetto_SX' -> '09', '031_Fianco_SX_INT' -> '31'
    (None se il nome non inizia col numero)."""
    m = re.match(r"\s*(\d+)", name or "")
    return _norm_num(m.group(1)) if m else None


def load_regole(base, log=print):
    """Carica _REGOLE_CNC.xlsx dal lavoro; se manca lo crea coi default."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    p = Path(base) / "_REGOLE_CNC.xlsx"
    if p.exists():
        try:
            wb = load_workbook(p, data_only=True)
            ws = wb.active
            regole = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                num = _norm_num(str(row[0]).split(".")[0])  # 9 / 9.0 / '09' / '031'->'31'
                regole[num] = (str(row[1] or ""), str(row[2] or ""), str(row[3] or ""))
            wb.close()
            log(f"  regole CNC/bordatura: {p.name} ({len(regole)} numeri)")
            return regole
        except Exception as ex:
            log(f"  ! {p.name} illeggibile ({ex}): uso i default")
            return dict(REGOLE_DEFAULT)
    # primo giro su questo lavoro: creo il file regole coi default
    wb = Workbook(); ws = wb.active; ws.title = "REGOLE"
    bold = Font(bold=True); fill = PatternFill("solid", fgColor="D9E1F2")
    for c, t in enumerate(["NUMERO", "MATERIALE", "CNC", "BORDATURA"], 1):
        cl = ws.cell(1, c, t); cl.font = bold; cl.fill = fill
    for r, (num, (mat, cnc, bordo)) in enumerate(sorted(REGOLE_DEFAULT.items()), 2):
        ws.cell(r, 1, num); ws.cell(r, 2, mat); ws.cell(r, 3, cnc); ws.cell(r, 4, bordo)
    for col, w in zip("ABCD", [10, 46, 18, 32]):
        ws.column_dimensions[col].width = w
    try:
        wb.save(p)
        log(f"  creato {p.name}: MODIFICALO per cambiare le regole del lavoro")
    except Exception:
        pass
    return dict(REGOLE_DEFAULT)


# ============================ CONFRONTO =====================================
def write_report(out, rows, regole=None, room_rl=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    regole = regole or {}
    room_rl = room_rl or {}
    wb = Workbook(); ws = wb.active; ws.title = "PROGRAMMI_UNICI"
    hdr = ["FILE PROGRAMMA", "PEZZO", "MATERIALE", "CAMERE", "Q.TA TOT",
           "L (mm)", "H (mm)", "SP (mm)", "N.FORI", "CNC", "BORDATURA",
           "POSIZIONI ORIGINALI"]
    bold = Font(bold=True); fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(border_style="thin", color="888888")
    bd = Border(thin, thin, thin, thin)
    for c, t in enumerate(hdr, 1):
        cl = ws.cell(1, c, t); cl.font = bold; cl.fill = fill; cl.border = bd
    rows.sort(key=lambda r: (int(re.match(r"(\d+)", r[1]).group(1))
                             if re.match(r"(\d+)", r[1]) else 999, r[0]))
    green = PatternFill("solid", fgColor="E2EFDA")
    for i, r in enumerate(rows, 2):
        # r = (outname, name, camere, qty, L, H, SP, nfori, posizioni, uguale_ovunque)
        num = num_pezzo(r[1])
        mat, cnc, bordo = regole.get(num, ("", "", ""))
        # 10_Base: CNC -> HOMAG se una camera della riga ha suffisso _R/_L
        if num == "10" and any(room_rl.get(c) for c in str(r[2]).split("+")):
            cnc = "HOMAG"
        # alcuni pezzi: 1 programma = piu' pezzi fisici -> moltiplica la Q.TA TOT
        qty = r[3]
        if isinstance(qty, (int, float)):
            qty = qty * QTY_MULT.get(num, 1)
        riga = [r[0], r[1], mat, r[2], qty, r[4], r[5], r[6], r[7],
                cnc, bordo, r[8]]
        for c, v in enumerate(riga, 1):
            cell = ws.cell(i, c, v); cell.border = bd
            if len(r) > 9 and r[9]:   # pezzo uguale in tutte le camere -> evidenzia
                cell.fill = green
    for col, w in zip("ABCDEFGHIJKL", [34, 26, 44, 28, 9, 9, 9, 9, 8, 16, 30, 30]):
        ws.column_dimensions[col].width = w
    dst = out / "_RIEPILOGO_programmi_unici.xlsx"
    try:
        wb.save(dst)
    except PermissionError:
        dst = out / "_RIEPILOGO_programmi_unici_NEW.xlsx"; wb.save(dst)
    return dst


def run_comparison(base, rooms, log=print):
    import ezdxf
    recs = {}
    total = sum(r["n"] for r in rooms); done = 0
    for r in rooms:
        for p in sorted(Path(r["path"]).glob("*.dxf")):
            try:
                doc = ezdxf.readfile(str(p))
            except Exception as ex:
                log(f"  ! {r['code']}/{p.name}: non leggibile ({ex})"); continue
            solids = model_solids(doc)
            if not solids:
                log(f"  ! {r['code']}/{p.name}: nessun solido, saltato"); continue
            try:
                g = geom_sig_solids(solids)
            except Exception as ex:
                log(f"  ! {r['code']}/{p.name}: ACIS non leggibile ({ex})"); continue
            if g is None:
                log(f"  ! {r['code']}/{p.name}: geometria vuota, saltato"); continue
            g["name"] = piece_name(solids[0], p.stem)
            g["code"] = r["code"]
            g["src"] = str(p)                      # percorso reale del file
            recs[f"{r['folder']}|{p.name}"] = g    # chiave per CARTELLA (univoca)
            done += 1
            if done % 5 == 0 or done == total:
                log(f"  letti {done}/{total} file...")
        log(f"  camera {r['code']} completata")

    groups = defaultdict(list); name_hashes = defaultdict(set)
    for k, g in recs.items():
        fn = k.split("|", 1)[1]
        groups[(g["name"], g["hash"])].append((g["code"], fn, k))
        name_hashes[g["name"]].add(g["hash"])

    out = Path(base) / "PROGRAMMI_UNICI"
    out.mkdir(exist_ok=True)
    safe = lambda s: re.sub(r'[<>:"/\\|?*]', "_", s).strip()

    # --- piano dei file di destinazione --------------------------------------
    manifest, rows, plan = {}, [], {}
    for (name, h), members in groups.items():
        rooms_cov = sorted(set(rm for rm, fn, k in members))
        need = len(name_hashes[name]) > 1           # camera nel nome solo se serve
        suffix = "_" + "+".join(rooms_cov) if need else ""
        outname = f"{safe(name)}{suffix}.dxf"
        # nome univoco: se due pezzi DIVERSI ricadono sullo stesso nome, numera
        if outname in plan:
            stem, ext = os.path.splitext(outname); j = 2
            while f"{stem}_{j}{ext}" in plan:
                j += 1
            outname = f"{stem}_{j}{ext}"
        # sorgente presa dal PERCORSO REALE memorizzato (no ambiguita' di codice)
        src_code, src_fn, src_key = sorted(members)[0]
        g0 = recs[src_key]
        plan[outname] = Path(g0["src"])
        manifest[outname] = {"name": name, "rooms": rooms_cov, "qty": len(members),
                             "dims": g0["dims"], "nholes": g0["nholes"],
                             "members": [(rm, fn) for rm, fn, k in members]}
        rows.append((outname, name, "+".join(rooms_cov),
                     len(members), g0["dims"][0], g0["dims"][1], g0["dims"][2],
                     g0["nholes"],
                     "; ".join(f"{rm}/{fn}" for rm, fn, k in sorted(members)),
                     not need))   # ultimo campo: pezzo uguale in tutte le camere dove compare

    # --- sincronizzazione MINIMA: tocca solo cio' che manca o e' diverso ------
    # (cosi' file gia' giusti ma aperti in AutoCAD/bloccati da Dropbox
    #  non vengono riscritti e non fanno fallire il lavoro)
    bloccati = []
    for p in out.glob("*.dxf"):                       # rimuovi i file non previsti
        if p.name not in plan:
            if not _unlink_retry(p):
                bloccati.append(p.name)
                log(f"  ! {p.name}: non previsto ma bloccato, lascialo o cancellalo a mano")
    for outname, src in plan.items():                 # copia mancanti/diversi
        dst = out / outname
        try:
            if dst.exists() and dst.stat().st_size == src.stat().st_size \
               and dst.read_bytes() == src.read_bytes():
                continue                              # gia' identico: non toccare
        except PermissionError:
            pass
        try:
            _copy_retry(src, dst)
        except PermissionError:
            bloccati.append(outname)
            log(f"  ! {outname}: BLOCCATO (aperto in AutoCAD?), non aggiornato")
        except FileNotFoundError:
            bloccati.append(outname)
            log(f"  ! {outname}: sorgente non trovato ({src}), saltato")

    (Path(base) / "_manifest.json").write_text(json.dumps(manifest))
    regole = load_regole(base, log)
    room_rl = {}                                   # codice camera -> ha suffisso _R/_L?
    for rm in rooms:
        room_rl[rm["code"]] = room_rl.get(rm["code"], False) or _camera_rl(rm["folder"])
    report = write_report(out, rows, regole, room_rl)
    return {"unici": len(manifest), "file_letti": len(recs),
            "uguali_tutte": sum(1 for r in rows if len(r) > 9 and r[9]),
            "out": str(out), "report": report.name, "bloccati": bloccati}


# ============================ GUI ==========================================
def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, ttk, messagebox, simpledialog
    root = tk.Tk()
    root.title("Confronto Programmi DXF  ->  Programmi Unici")
    root.geometry("780x600")
    state = {"base": None, "rooms": []}
    q = queue.Queue()

    head = tk.Label(root, text="Confronto Programmi DXF",
                    font=("Segoe UI", 14, "bold"))
    head.pack(anchor="w", padx=12, pady=(10, 0))

    top = tk.Frame(root); top.pack(fill="x", padx=12, pady=6)
    tk.Label(top, text="1) Cartella del lavoro (quella che contiene le cartelle DXF delle camere):",
             font=("Segoe UI", 10, "bold")).pack(anchor="w")
    row = tk.Frame(top); row.pack(fill="x", pady=4)
    base_var = tk.StringVar()
    tk.Entry(row, textvariable=base_var, state="readonly").pack(side="left", fill="x", expand=True)

    def choose():
        d = filedialog.askdirectory(title="Scegli la cartella del lavoro")
        if d:
            base_var.set(d); state["base"] = d
            state["rooms"] = detect_rooms(d); refresh()
            if not state["rooms"]:
                messagebox.showwarning("Nessuna camera",
                    "In questa cartella non ho trovato sottocartelle con file .dxf.")
    tk.Button(row, text="Sfoglia...", command=choose).pack(side="left", padx=6)

    tk.Label(top, text="2) Camere rilevate  (doppio clic su una riga per correggere il codice camera):",
             font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(8, 0))
    tree = ttk.Treeview(top, columns=("camera", "cartella", "n"), show="headings", height=7)
    for c, t, w in (("camera", "Camera", 90), ("cartella", "Cartella", 470), ("n", "N. DXF", 70)):
        tree.heading(c, text=t); tree.column(c, width=w, anchor="w")
    tree.pack(fill="x", pady=4)

    def refresh():
        tree.delete(*tree.get_children())
        for i, r in enumerate(state["rooms"]):
            tree.insert("", "end", iid=str(i), values=(r["code"], r["folder"], r["n"]))

    def edit_code(_ev):
        iid = tree.focus()
        if not iid:
            return
        i = int(iid)
        new = simpledialog.askstring("Codice camera",
            f"Codice camera per la cartella:\n{state['rooms'][i]['folder']}",
            initialvalue=state["rooms"][i]["code"])
        if new and new.strip():
            state["rooms"][i]["code"] = new.strip(); refresh()
    tree.bind("<Double-1>", edit_code)

    runbtn = tk.Button(root, text="▶  CREA PROGRAMMI UNICI",
                       bg="#2e7d32", fg="white", font=("Segoe UI", 13, "bold"),
                       height=2, command=lambda: start())
    runbtn.pack(fill="x", padx=12, pady=8)

    logbox = tk.Text(root, height=14, wrap="word")
    logbox.pack(fill="both", expand=True, padx=12, pady=(0, 10))

    def append(msg):
        logbox.insert("end", msg + "\n"); logbox.see("end")

    def worker():
        try:
            res = run_comparison(state["base"], state["rooms"],
                                 log=lambda m: q.put(("log", m)))
            q.put(("done", res))
        except Exception:
            import traceback; q.put(("err", traceback.format_exc()))

    def poll():
        try:
            while True:
                kind, payload = q.get_nowait()
                if kind == "log":
                    append(payload)
                elif kind == "done":
                    append(f"\n=== FATTO: {payload['unici']} programmi unici "
                           f"({payload['uguali_tutte']} uguali in tutte le camere) ===")
                    append(f"Cartella:  {payload['out']}")
                    append(f"Riepilogo: {payload['report']}")
                    runbtn.config(state="normal", text="▶  CREA PROGRAMMI UNICI")
                    try: os.startfile(payload["out"])
                    except Exception: pass
                    msg = f"{payload['unici']} programmi unici creati in:\n{payload['out']}"
                    if payload.get("bloccati"):
                        msg += ("\n\nATTENZIONE - file bloccati non aggiornati "
                                "(chiudili in AutoCAD e rilancia):\n  "
                                + "\n  ".join(payload["bloccati"]))
                        messagebox.showwarning("Fatto (con avvisi)", msg)
                    else:
                        messagebox.showinfo("Fatto", msg)
                elif kind == "err":
                    append("ERRORE:\n" + payload)
                    runbtn.config(state="normal", text="▶  CREA PROGRAMMI UNICI")
                    messagebox.showerror("Errore", payload.splitlines()[-1])
        except queue.Empty:
            pass
        root.after(150, poll)

    def start():
        if not state["base"] or not state["rooms"]:
            messagebox.showwarning("Attenzione",
                "Scegli prima la cartella del lavoro con le camere."); return
        runbtn.config(state="disabled", text="Elaborazione in corso...")
        logbox.delete("1.0", "end"); append("Avvio confronto...")
        threading.Thread(target=worker, daemon=True).start()

    root.after(150, poll)
    root.mainloop()


# ============================ MAIN =========================================
if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--run":
        base = sys.argv[2]
        logf = open(Path(base) / "_confronto_log.txt", "w", encoding="utf-8")

        def _log(msg):
            print(msg)
            logf.write(str(msg) + "\n"); logf.flush()

        try:
            rooms = detect_rooms(base)
            _log(f"Camere: {[(r['code'], r['n']) for r in rooms]}")
            _log(f"RESULT: {run_comparison(base, rooms, log=_log)}")
        except Exception:
            import traceback
            _log("ERRORE:\n" + traceback.format_exc())
            sys.exit(1)
        finally:
            logf.close()
    else:
        launch_gui()
