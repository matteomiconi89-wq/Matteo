"""
genera_sezionatura_cp_21032 - exe esterno (senza aprire Excel).

Parte da un file BOM "origine" (foglio unico con colonne:
Set produzione, Stanza, Modello, Quantita', Numero parte, Materiale,
Lunghezza, Larghezza, Spessore) e crea un NUOVO file "_LAVORATO" con:
  - colonne A-H standard + P = commessa (ricavata dal Modello)
  - colonne J-O / Q con le FORMULE (visibili e controllabili)
  - foglio "sezionatura" consolidato (valori) raggruppato per
    materiale / numero parte / lunghezza / larghezza
  - formattazione: intestazioni colorate, bordi, filtri, blocca riquadri

I frontali cassetto SX+DX+CX vengono uniti in un'unica riga, anche con
suffisso "_bis" o prefisso modello (es. J-03.11_09_Frontale_Cassetto_SX).

Accetta anche il _RIEPILOGO_programmi_unici.xlsx (output di
confronta_file_unici_dxf): in quel caso le righe dei frontali cassetto
vengono esplose per camera (colonna CAMERE, es. "A102+B302+C101") e
l'unione SX+DX+CX avviene camera per camera; nel BOM vanno i VALORI
calcolati (niente formule, l'abbinamento per camera non e' esprimibile
con SUMIFS su righe deduplicate).

Excel NON viene mai aperto (usa openpyxl). L'origine resta intatto.
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# --- regole (come il file con le formule) ---
MAT_ORDER = [
    "MD PLACCATO 1 LATO + CONTR. sp.16 (finito 16)",
    "MD PLACCATO 1 LATO + CONTR. sp.8 (finito 9)",
    "MD PLACCATO 1 LATO + CONTR. sp.19 (finito 19)",
    "MD PLACCATO 2 LATI sp.10 (finito 11)",
    "MD PLACCATO 2 LATI sp.19 (finito 19)",
    "MDF GREZZO sp.19",
    "MDF GREZZO sp.25",
]
ABBONDO = 10
SX_MARK = "Frontale_Cassetto_SX"
DX_MARK = "Frontale_Cassetto_DX"
CX_MARK = "Frontale_Cassetto_CX"

HDR_BOM = {
    1: "Set produzione", 2: "Stanza", 3: "Quantita'", 4: "Numero parte",
    5: "Materiale", 6: "Lunghezza (mm)", 7: "Larghezza (mm)", 8: "Spessore (mm)",
    10: "Lunghezza (mm)", 11: "Larghezza (mm)", 12: "Quantita'", 13: "Materiale",
    14: "Numero parte", 15: "Stanza", 17: "Set produzione", 18: "Spessore (mm)",
}
HDR_SEZ = {
    1: "Lunghezza (mm)", 2: "Larghezza (mm)", 3: "Quantita'", 4: "Materiale",
    5: "Numero parte", 6: "Stanza", 8: "Set produzione", 9: "Spessore (mm)",
}
WIDTHS_BOM = {"A": 12, "B": 10, "C": 9, "D": 30, "E": 40, "F": 13, "G": 13,
              "H": 12, "I": 3, "J": 13, "K": 13, "L": 9, "M": 40, "N": 30,
              "O": 18, "P": 14, "Q": 12, "R": 13}
WIDTHS_SEZ = {"A": 13, "B": 13, "C": 9, "D": 40, "E": 30, "F": 62, "H": 16, "I": 13}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF15BFCA")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _rank(mat):
    return MAT_ORDER.index(mat) if mat in MAT_ORDER else 999


def _num(x):
    return x if isinstance(x, (int, float)) else 0


def _norm(h):
    return str(h).strip().lower() if h is not None else ""


def _role(part):
    p = part or ""
    # frontale GIA' unito (es. 09_Frontale_Cassetto_SX_CX_DX da ESPSOL,
    # o DX_CX_SX fuso a mano): niente nuova unione, e' un pezzo normale
    if sum(1 for t in ("SX", "DX", "CX") if t in p) >= 2:
        return None
    if SX_MARK in p:
        return "SX"
    if DX_MARK in p:
        return "DX"
    if CX_MARK in p:
        return "CX"
    return None


def modello_to_commessa(mod):
    """ 'J-03.01' -> '___J03_01' (separatore ___ per la colonna Stanza). """
    if mod in (None, ""):
        return ""
    return "___" + str(mod).strip().replace("-", "").replace(".", "_")


# ---------- formule (generalizzate per varianti _bis / prefisso) ----------
def _f_sx(r):
    return f'ISNUMBER(SEARCH("{SX_MARK}",$D{r}))'


def _f_dxcx(r):
    return (f'OR(ISNUMBER(SEARCH("{DX_MARK}",$D{r})),'
            f'ISNUMBER(SEARCH("{CX_MARK}",$D{r})))')


def _sumifs(mark, r):
    return f'SUMIFS($F:$F,$D:$D,"*{mark}*",$B:$B,$B{r},$A:$A,$A{r})'


def _countifs(mark, r):
    return f'COUNTIFS($D:$D,"*{mark}*",$B:$B,$B{r},$A:$A,$A{r})'


def _formule(r):
    sx = _f_sx(r)
    dxcx = _f_dxcx(r)
    J = (f'=IF({sx},(F{r}+{ABBONDO})'
         f'+({_sumifs(DX_MARK, r)}+{ABBONDO}*{_countifs(DX_MARK, r)})'
         f'+({_sumifs(CX_MARK, r)}+{ABBONDO}*{_countifs(CX_MARK, r)}),'
         f'IF({dxcx},"",F{r}+{ABBONDO}))')
    K = f'=IF({dxcx},"",G{r}+{ABBONDO})'
    L = f'=IF({dxcx},"",C{r})'
    M = f'=IF({dxcx},"",E{r})'
    N = (f'=IF({sx},SUBSTITUTE($D{r},"{SX_MARK}","{SX_MARK}_DX_CX"),'
         f'IF({dxcx},"",$D{r}))')
    O = f'=IF({dxcx},"",$B{r}&$P{r})'
    Q = f'=IF({dxcx},"",$A{r})'
    return {10: J, 11: K, 12: L, 13: M, 14: N, 15: O, 17: Q}


def _find_header_row(ws):
    """Ritorna (riga_intestazioni, modo): 'bom' oppure 'riepilogo'."""
    for r in range(1, 8):
        for c in range(1, 5):
            v = _norm(ws.cell(r, c).value)
            if v.startswith("set produzione"):
                return r, "bom"
            if v.startswith("file programma"):
                return r, "riepilogo"
    raise ValueError("Intestazioni non trovate: ne' 'Set produzione' (BOM)"
                     " ne' 'FILE PROGRAMMA' (riepilogo programmi unici).")


def _map_columns(ws, hrow):
    cols = {}
    c = 1
    while True:
        h = ws.cell(hrow, c).value
        if h is None or str(h).strip() == "":
            break
        cols[_norm(h)] = c
        c += 1

    def find(*prefixes):
        for key, col in cols.items():
            for p in prefixes:
                if key.startswith(p):
                    return col
        return None

    f = {
        "set": find("set produzione", "set"), "stanza": find("stanza"),
        "modello": find("modello"), "qta": find("quantit"),
        "parte": find("numero parte", "parte"), "materiale": find("materiale"),
        "lung": find("lunghezza"), "larg": find("larghezza"), "sp": find("spessore"),
    }
    obblig = ("set", "stanza", "qta", "parte", "materiale", "lung", "larg")
    mancanti = [k for k in obblig if f[k] is None]
    if mancanti:
        raise ValueError("Colonne mancanti nell'origine: " + ", ".join(mancanti))
    return f


def _map_columns_riepilogo(ws, hrow):
    cols = {}
    c = 1
    while True:
        h = ws.cell(hrow, c).value
        if h is None or str(h).strip() == "":
            break
        cols[_norm(h)] = c
        c += 1

    def find(*prefixes):
        for key, col in cols.items():
            for p in prefixes:
                if key.startswith(p):
                    return col
        return None

    f = {
        "parte": find("pezzo"), "materiale": find("materiale"),
        "camere": find("camere"), "qta": find("q.ta", "q.tà", "quantit"),
        "lung": find("l (", "l("), "larg": find("h (", "h("),
        "sp": find("sp (", "sp(", "spessore"),
    }
    obblig = ("parte", "materiale", "camere", "qta", "lung", "larg")
    mancanti = [k for k in obblig if f[k] is None]
    if mancanti:
        raise ValueError("Colonne mancanti nel riepilogo: " + ", ".join(mancanti))
    return f


def _set_from_path(path):
    """Ricava il set produzione dalla cartella (es. ...\\SET_19_20_21_22\\...)."""
    for part in os.path.normpath(os.path.abspath(path)).split(os.sep):
        if part.upper().startswith("SET"):
            return part
    return ""


def process_source(src_path, out_path):
    wb_in = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb_in[wb_in.sheetnames[0]]
    hrow, mode = _find_header_row(ws)
    title = ws.cell(1, 1).value if hrow > 1 else None

    recs = []
    if mode == "bom":
        f = _map_columns(ws, hrow)
        for r in range(hrow + 1, ws.max_row + 1):
            parte = ws.cell(r, f["parte"]).value
            if parte in (None, ""):
                continue
            rec = {
                "set": ws.cell(r, f["set"]).value,
                "stanza": ws.cell(r, f["stanza"]).value,
                "modello": ws.cell(r, f["modello"]).value if f["modello"] else None,
                "qta": ws.cell(r, f["qta"]).value,
                "parte": parte,
                "materiale": ws.cell(r, f["materiale"]).value,
                "lung": ws.cell(r, f["lung"]).value,
                "larg": ws.cell(r, f["larg"]).value,
                "sp": ws.cell(r, f["sp"]).value if f["sp"] else None,
            }
            rec["commessa"] = modello_to_commessa(rec["modello"])
            recs.append(rec)
    else:  # riepilogo programmi unici
        f = _map_columns_riepilogo(ws, hrow)
        set_prod = _set_from_path(src_path)
        for r in range(hrow + 1, ws.max_row + 1):
            parte = ws.cell(r, f["parte"]).value
            if parte in (None, ""):
                continue
            base = {
                "set": set_prod, "modello": None, "commessa": "",
                "parte": parte,
                "materiale": ws.cell(r, f["materiale"]).value,
                "lung": ws.cell(r, f["lung"]).value,
                "larg": ws.cell(r, f["larg"]).value,
                "sp": ws.cell(r, f["sp"]).value if f["sp"] else None,
            }
            camere = ws.cell(r, f["camere"]).value
            cams = [c.strip() for c in str(camere).split("+")
                    if c.strip()] if camere not in (None, "") else []
            if _role(parte) and cams:
                # frontale cassetto: una riga per camera, cosi' l'unione
                # SX+DX+CX avviene camera per camera
                for cam in cams:
                    recs.append(dict(base, stanza=cam, qta=1))
            else:
                recs.append(dict(base, stanza=camere,
                                 qta=ws.cell(r, f["qta"]).value))

        # controllo: ogni camera con DX o CX deve avere anche l'SX,
        # altrimenti quelle lunghezze andrebbero perse in silenzio
        con_sx = {r["stanza"] for r in recs if _role(r["parte"]) == "SX"}
        senza_sx = sorted({r["stanza"] for r in recs
                           if _role(r["parte"]) in ("DX", "CX")
                           and r["stanza"] not in con_sx})
        if senza_sx:
            raise ValueError("Camere con frontale DX/CX ma senza SX: "
                             + ", ".join(map(str, senza_sx)))

    # somme frontali cassetto DX/CX per (set, stanza)  -- come fa la formula
    dxs, dxc, cxs, cxc = {}, {}, {}, {}
    for rec in recs:
        role = _role(rec["parte"])
        k = (rec["set"], rec["stanza"])
        if role == "DX":
            dxs[k] = dxs.get(k, 0) + _num(rec["lung"]); dxc[k] = dxc.get(k, 0) + 1
        elif role == "CX":
            cxs[k] = cxs.get(k, 0) + _num(rec["lung"]); cxc[k] = cxc.get(k, 0) + 1

    # valori interni (servono per la sezionatura; nel BOM vanno le formule)
    for rec in recs:
        role = _role(rec["parte"])
        k = (rec["set"], rec["stanza"])
        if role in ("DX", "CX"):
            rec.update(J=None, K=None, L=None, M=None, N=None, O=None, Q=None)
            continue
        if role == "SX":
            J = ((_num(rec["lung"]) + ABBONDO)
                 + (dxs.get(k, 0) + ABBONDO * dxc.get(k, 0))
                 + (cxs.get(k, 0) + ABBONDO * cxc.get(k, 0)))
            N = rec["parte"].replace(SX_MARK, SX_MARK + "_DX_CX")
        else:
            J = _num(rec["lung"]) + ABBONDO
            N = rec["parte"]
        rec.update(J=J, K=_num(rec["larg"]) + ABBONDO, L=rec["qta"],
                   M=rec["materiale"], N=N,
                   O=f"{rec['stanza']}{rec['commessa']}", Q=rec["set"])

    # consolidamento sezionatura
    groups, order = {}, []
    for rec in recs:
        if not rec["N"]:
            continue
        key = (rec["M"], rec["N"], rec["J"], rec["K"])
        if key not in groups:
            groups[key] = {"q": 0, "st": [], "set": []}
            order.append(key)
        gg = groups[key]
        gg["q"] += _num(rec["L"])
        if rec["O"] not in gg["st"]:
            gg["st"].append(rec["O"])
        if str(rec["Q"]) not in gg["set"]:
            gg["set"].append(str(rec["Q"]))
    sez_rows = []
    for (mat, parte, lung, larg) in order:
        gg = groups[(mat, parte, lung, larg)]
        sez_rows.append([lung, larg, gg["q"], mat, parte,
                         ", ".join(gg["st"]), "", ", ".join(gg["set"])])
    sez_rows.sort(key=lambda r: (_rank(r[3]), r[0]))

    # ---------------- scrittura nuovo workbook ----------------
    wb = openpyxl.Workbook()
    bom = wb.active
    bom.title = "BOM unificata"
    sez = wb.create_sheet("sezionatura")

    if title:
        c = bom.cell(1, 1)
        c.value = title
        c.font = TITLE_FONT
    for col, h in HDR_BOM.items():
        cell = bom.cell(3, col)
        cell.value = h
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    rr = 4
    for rec in recs:
        bom.cell(rr, 1).value = rec["set"]
        bom.cell(rr, 2).value = rec["stanza"]
        bom.cell(rr, 3).value = rec["qta"]
        bom.cell(rr, 4).value = rec["parte"]
        bom.cell(rr, 5).value = rec["materiale"]
        bom.cell(rr, 6).value = rec["lung"]
        bom.cell(rr, 7).value = rec["larg"]
        bom.cell(rr, 8).value = rec["sp"]
        bom.cell(rr, 16).value = rec["commessa"] or None      # P = commessa
        if mode == "bom":
            for col, formula in _formule(rr).items():         # J-O / Q = formule
                bom.cell(rr, col).value = formula
        else:                                                 # riepilogo: valori
            for col, key in ((10, "J"), (11, "K"), (12, "L"), (13, "M"),
                             (14, "N"), (15, "O"), (17, "Q")):
                bom.cell(rr, col).value = rec[key]
        rr += 1
    last_bom = rr - 1

    for col, h in HDR_SEZ.items():
        cell = sez.cell(1, col)
        cell.value = h
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(sez_rows):
        for col, val in enumerate(row, start=1):
            sez.cell(2 + i, col).value = val
    last_sez = 1 + len(sez_rows)

    # bordi sull'area tabella
    for r in range(3, last_bom + 1):
        for c in list(range(1, 9)) + list(range(10, 18)):
            bom.cell(r, c).border = BORDER
    for r in range(1, last_sez + 1):
        for c in list(range(1, 7)) + [8, 9]:
            sez.cell(r, c).border = BORDER

    # larghezze, filtri, blocca riquadri
    for letter, w in WIDTHS_BOM.items():
        bom.column_dimensions[letter].width = w
    for letter, w in WIDTHS_SEZ.items():
        sez.column_dimensions[letter].width = w
    bom.auto_filter.ref = f"A3:R{last_bom}"
    sez.auto_filter.ref = f"A1:I{last_sez}"
    bom.freeze_panes = "A4"
    sez.freeze_panes = "A2"

    wb.save(out_path)
    return len(recs), len(sez_rows)


def _out_path_for(src_path):
    d = os.path.dirname(src_path)
    base = os.path.splitext(os.path.basename(src_path))[0]
    cand = os.path.join(d, base + "_LAVORATO.xlsx")
    i = 2
    while os.path.exists(cand):
        cand = os.path.join(d, f"{base}_LAVORATO_{i}.xlsx")
        i += 1
    return cand


def _is_locked(path):
    lock = os.path.join(os.path.dirname(path), "~$" + os.path.basename(path))
    return os.path.exists(lock)


def main():
    if len(sys.argv) >= 3 and sys.argv[1] == "--selftest":
        try:
            out = sys.argv[3] if len(sys.argv) >= 4 else _out_path_for(sys.argv[2])
            n, k = process_source(sys.argv[2], out)
            print(f"OK righe={n} sezionatura={k} out={out}")
            sys.exit(0)
        except Exception as e:
            print(f"ERRORE {type(e).__name__}: {e}")
            sys.exit(1)

    root = tk.Tk()
    root.withdraw()

    paths = [a for a in sys.argv[1:] if a and not a.startswith("-")]
    if not paths:
        p = filedialog.askopenfilename(
            title="Scegli il file da elaborare (BOM origine o _RIEPILOGO_programmi_unici)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Tutti i file", "*.*")],
        )
        if not p:
            return
        paths = [p]

    esiti = []
    for src in paths:
        nome = os.path.basename(src)
        if _is_locked(src):
            esiti.append(f"[APERTO IN EXCEL]  {nome}  -> chiudilo e riprova")
            continue
        try:
            out = _out_path_for(src)
            n, k = process_source(src, out)
            esiti.append(f"[OK]  {nome}\n      -> {os.path.basename(out)}"
                         f"  ({n} righe, {k} righe sezionatura)")
        except Exception as e:
            esiti.append(f"[ERRORE]  {nome}  -> {type(e).__name__}: {e}")

    ok = sum(1 for e in esiti if e.startswith("[OK]"))
    titolo = "Fatto" if ok == len(esiti) else "Completato con avvisi"
    messagebox.showinfo(titolo, "\n".join(esiti))


if __name__ == "__main__":
    main()
